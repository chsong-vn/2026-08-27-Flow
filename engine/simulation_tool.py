"""
Flow Chemistry Simulation Tool

⚠ 구 워크플로 재현 — 타이밍 신뢰 금지 (engine/test_detailed_timing.py 상단 배너 참조.
  실행 진실원은 strict_engine. 2026-08-13 개편 미반영.)

@codesyncer-context: 파라미터 입력 → Excel 생성 도구
- 단일 조건 시뮬레이션
- 다중 조건 비교 (파라미터 스윕)
- 최적 조건 탐색

Usage:
    python -m engine.simulation_tool              # 대화형 모드
    python -m engine.simulation_tool --gui        # GUI 모드
    python -m engine.simulation_tool --sweep      # 파라미터 스윕 모드
"""

import os
import sys
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Dict, List, Tuple, Optional
import json

# 상위 모듈 임포트
from engine.test_detailed_timing import (
    GroupConfig, ReagentConfig, ExperimentParams, SystemParams,
    run_detailed_simulation, export_to_excel, calculate_flows,
    print_summary_table, HAS_OPENPYXL
)

if HAS_OPENPYXL:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, Reference


# ═══════════════════════════════════════════════════════════════════════════════
# 1. 설정 저장/로드
# ═══════════════════════════════════════════════════════════════════════════════

CONFIG_DIR = os.path.join(os.path.dirname(__file__), "..", "temp", "simulation_configs")
os.makedirs(CONFIG_DIR, exist_ok=True)


def save_config(name: str, groups: List[GroupConfig], experiment: ExperimentParams,
                params: SystemParams) -> str:
    """설정을 JSON으로 저장"""
    config = {
        "groups": [
            {
                "name": g.name,
                "reagent": {
                    "name": g.reagent.name,
                    "port": g.reagent.port,
                    "concentration": g.reagent.concentration,
                    "equivalents": g.reagent.equivalents
                },
                "syringe_volume": g.syringe_volume,
                "refill_rate": g.refill_rate
            }
            for g in groups
        ],
        "experiment": {
            "temperature": experiment.temperature,
            "residence_time": experiment.residence_time,
            "reaction_volume": experiment.reaction_volume,
            "volume_per_tube": experiment.volume_per_tube,
            "start_tube": experiment.start_tube
        },
        "params": {
            "reactor_volume": params.reactor_volume,
            "post_reactor_volume": params.post_reactor_volume,
            "collection_line_volume": params.collection_line_volume,
            "dead_volume_per_line": params.dead_volume_per_line
        }
    }

    filepath = os.path.join(CONFIG_DIR, f"{name}.json")
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

    return filepath


def load_config(name: str) -> Tuple[List[GroupConfig], ExperimentParams, SystemParams]:
    """JSON에서 설정 로드"""
    filepath = os.path.join(CONFIG_DIR, f"{name}.json")
    with open(filepath, "r", encoding="utf-8") as f:
        config = json.load(f)

    groups = [
        GroupConfig(
            name=g["name"],
            reagent=ReagentConfig(
                name=g["reagent"]["name"],
                port=g["reagent"]["port"],
                concentration=g["reagent"]["concentration"],
                equivalents=g["reagent"]["equivalents"]
            ),
            syringe_volume=g["syringe_volume"],
            refill_rate=g["refill_rate"]
        )
        for g in config["groups"]
    ]

    experiment = ExperimentParams(
        temperature=config["experiment"]["temperature"],
        residence_time=config["experiment"]["residence_time"],
        reaction_volume=config["experiment"]["reaction_volume"],
        volume_per_tube=config["experiment"]["volume_per_tube"],
        start_tube=config["experiment"].get("start_tube", 1)
    )

    params = SystemParams(
        reactor_volume=config["params"]["reactor_volume"],
        post_reactor_volume=config["params"]["post_reactor_volume"],
        collection_line_volume=config["params"]["collection_line_volume"],
        dead_volume_per_line=config["params"]["dead_volume_per_line"]
    )

    return groups, experiment, params


def list_saved_configs() -> List[str]:
    """저장된 설정 목록"""
    files = [f.replace(".json", "") for f in os.listdir(CONFIG_DIR) if f.endswith(".json")]
    return files


# ═══════════════════════════════════════════════════════════════════════════════
# 2. 대화형 입력
# ═══════════════════════════════════════════════════════════════════════════════

def input_float(prompt: str, default: float = None) -> float:
    """float 입력 받기"""
    if default is not None:
        prompt = f"{prompt} [{default}]: "
    else:
        prompt = f"{prompt}: "

    while True:
        val = input(prompt).strip()
        if val == "" and default is not None:
            return default
        try:
            return float(val)
        except ValueError:
            print("  숫자를 입력하세요.")


def input_int(prompt: str, default: int = None) -> int:
    """int 입력 받기"""
    if default is not None:
        prompt = f"{prompt} [{default}]: "
    else:
        prompt = f"{prompt}: "

    while True:
        val = input(prompt).strip()
        if val == "" and default is not None:
            return default
        try:
            return int(val)
        except ValueError:
            print("  정수를 입력하세요.")


def input_groups() -> List[GroupConfig]:
    """그룹 설정 입력"""
    print("\n" + "=" * 60)
    print("GROUP CONFIGURATION")
    print("=" * 60)

    num_groups = input_int("펌프 그룹 개수", 3)
    groups = []

    for i in range(num_groups):
        name = f"Group {chr(65 + i)}"  # Group A, B, C...
        print(f"\n--- {name} ---")

        reagent_name = input(f"  시약 이름 [Reagent {chr(65 + i)}]: ").strip()
        if not reagent_name:
            reagent_name = f"Reagent {chr(65 + i)}"

        port = input_int(f"  12-way 포트 번호", 2 + i)
        concentration = input_float(f"  농도 (M)", 1.0 if i == 0 else 0.5)
        equivalents = input_float(f"  몰비 (eq)", 1.0)
        syringe_vol = input_float(f"  시린지 용량 (mL)", 10.0)
        refill_rate = input_float(f"  리필 속도 (mL/min)", 40.0)

        groups.append(GroupConfig(
            name=name,
            reagent=ReagentConfig(
                name=reagent_name,
                port=port,
                concentration=concentration,
                equivalents=equivalents
            ),
            syringe_volume=syringe_vol,
            refill_rate=refill_rate
        ))

    return groups


def input_experiment() -> ExperimentParams:
    """실험 파라미터 입력"""
    print("\n" + "=" * 60)
    print("EXPERIMENT PARAMETERS")
    print("=" * 60)

    temperature = input_float("반응 온도 (°C)", 80.0)
    residence_time = input_float("체류 시간 (min)", 10.0)
    reaction_volume = input_float("반응량 (mL)", 5.0)
    volume_per_tube = input_float("튜브당 수거량 (mL)", 1.5)
    start_tube = input_int("시작 튜브 번호", 1)

    return ExperimentParams(
        temperature=temperature,
        residence_time=residence_time,
        reaction_volume=reaction_volume,
        volume_per_tube=volume_per_tube,
        start_tube=start_tube
    )


def input_system_params() -> SystemParams:
    """시스템 파라미터 입력"""
    print("\n" + "=" * 60)
    print("SYSTEM PARAMETERS")
    print("=" * 60)

    use_default = input("기본 시스템 파라미터 사용? (Y/n): ").strip().lower()
    if use_default != "n":
        return SystemParams()

    reactor_vol = input_float("반응기 부피 (mL)", 7.854)
    post_reactor_vol = input_float("반응기 후 공통 구간 (mL)", 2.0)
    collection_line_vol = input_float("수거 라인 부피 (mL)", 1.0)
    dead_vol = input_float("라인당 데드볼륨 (mL)", 0.785)

    return SystemParams(
        reactor_volume=reactor_vol,
        post_reactor_volume=post_reactor_vol,
        collection_line_volume=collection_line_vol,
        dead_volume_per_line=dead_vol
    )


# ═══════════════════════════════════════════════════════════════════════════════
# 3. 파라미터 스윕 (최적화용)
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class SweepResult:
    """스윕 결과"""
    condition_id: int
    params: Dict
    total_time: float
    total_flow: float
    injection_time: float
    num_refills: int
    flows: Dict[str, float]


def run_parameter_sweep(
    groups: List[GroupConfig],
    base_experiment: ExperimentParams,
    params: SystemParams,
    sweep_param: str,
    sweep_values: List[float]
) -> List[SweepResult]:
    """
    파라미터 스윕 실행

    Args:
        sweep_param: 스윕할 파라미터 ("residence_time", "temperature", "reaction_volume", 등)
        sweep_values: 스윕할 값 리스트

    Returns:
        스윕 결과 리스트
    """
    results = []

    for i, value in enumerate(sweep_values):
        # 실험 파라미터 복사 및 수정
        exp_dict = {
            "temperature": base_experiment.temperature,
            "residence_time": base_experiment.residence_time,
            "reaction_volume": base_experiment.reaction_volume,
            "volume_per_tube": base_experiment.volume_per_tube,
            "start_tube": base_experiment.start_tube
        }

        if sweep_param in exp_dict:
            exp_dict[sweep_param] = value

        experiment = ExperimentParams(**exp_dict)

        # 시뮬레이션 실행
        print(f"\n[{i+1}/{len(sweep_values)}] {sweep_param} = {value}")
        events, total_time, flows, tracker = run_detailed_simulation(groups, experiment, params)

        # 리필 횟수 계산
        from engine.test_detailed_timing import EventType
        num_refills = len([e for e in events if e.event_type == EventType.PUMP_WITHDRAW_START])

        # Injection 시간 계산
        total_flow = sum(flows.values())
        injection_time = (experiment.reaction_volume / total_flow) * 60 if total_flow > 0 else 0

        results.append(SweepResult(
            condition_id=i + 1,
            params={sweep_param: value},
            total_time=total_time,
            total_flow=total_flow,
            injection_time=injection_time,
            num_refills=num_refills,
            flows=flows
        ))

    return results


def export_sweep_results(
    results: List[SweepResult],
    groups: List[GroupConfig],
    sweep_param: str,
    filename: str = None
) -> str:
    """스윕 결과를 Excel로 내보내기"""

    if not HAS_OPENPYXL:
        print("openpyxl이 설치되지 않았습니다.")
        return None

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"sweep_{sweep_param}_{timestamp}.xlsx"

    output_dir = os.path.join(os.path.dirname(__file__), "..", "temp")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sweep Results"

    # 헤더 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # 헤더
    headers = ["#", sweep_param, "Total Time (s)", "Total Time (min)",
               "Total Flow (mL/min)", "Injection Time (s)", "Refill Count"]

    for g in groups:
        headers.append(f"{g.name} Flow")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    # 데이터
    for row, r in enumerate(results, 2):
        ws.cell(row=row, column=1, value=r.condition_id)
        ws.cell(row=row, column=2, value=list(r.params.values())[0])
        ws.cell(row=row, column=3, value=round(r.total_time, 1))
        ws.cell(row=row, column=4, value=round(r.total_time / 60, 2))
        ws.cell(row=row, column=5, value=round(r.total_flow, 4))
        ws.cell(row=row, column=6, value=round(r.injection_time, 1))
        ws.cell(row=row, column=7, value=r.num_refills)

        for col, g in enumerate(groups, 8):
            ws.cell(row=row, column=col, value=round(r.flows.get(g.name, 0), 4))

    # 차트 추가
    if len(results) > 1:
        chart = LineChart()
        chart.title = f"{sweep_param} vs Total Time"
        chart.x_axis.title = sweep_param
        chart.y_axis.title = "Total Time (min)"

        data = Reference(ws, min_col=4, min_row=1, max_row=len(results) + 1)
        cats = Reference(ws, min_col=2, min_row=2, max_row=len(results) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 10

        ws.add_chart(chart, f"J2")

    # 열 너비 조정
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

    wb.save(filepath)
    print(f"\n[Excel] 스윕 결과 저장됨: {filepath}")

    return filepath


# ═══════════════════════════════════════════════════════════════════════════════
# 4. 최적 조건 탐색
# ═══════════════════════════════════════════════════════════════════════════════

def find_optimal_conditions(
    groups: List[GroupConfig],
    params: SystemParams,
    target: str = "min_time",
    constraints: Dict = None
) -> Dict:
    """
    최적 조건 탐색

    Args:
        target: 최적화 목표
            - "min_time": 최소 시퀀스 시간
            - "min_refills": 최소 리필 횟수
            - "max_throughput": 최대 처리량 (volume/time)
        constraints: 제약 조건
            - "max_total_time": 최대 허용 시간 (초)
            - "max_refills": 최대 리필 횟수
            - "min_residence_time": 최소 체류 시간 (min)
            - "max_residence_time": 최대 체류 시간 (min)

    Returns:
        최적 조건 및 결과
    """
    if constraints is None:
        constraints = {}

    # 기본 스윕 범위
    rt_range = [
        max(constraints.get("min_residence_time", 1.0), 1.0),
        min(constraints.get("max_residence_time", 30.0), 30.0)
    ]

    # Residence time 스윕
    rt_values = []
    rt = rt_range[0]
    while rt <= rt_range[1]:
        rt_values.append(rt)
        rt += 1.0  # 1분 간격

    print(f"\n최적 조건 탐색 중... (Residence Time: {rt_range[0]} ~ {rt_range[1]} min)")

    base_experiment = ExperimentParams(
        temperature=80.0,
        residence_time=10.0,
        reaction_volume=5.0,
        volume_per_tube=1.5,
        start_tube=1
    )

    results = run_parameter_sweep(groups, base_experiment, params, "residence_time", rt_values)

    # 제약 조건 필터링
    valid_results = results
    if "max_total_time" in constraints:
        valid_results = [r for r in valid_results if r.total_time <= constraints["max_total_time"]]
    if "max_refills" in constraints:
        valid_results = [r for r in valid_results if r.num_refills <= constraints["max_refills"]]

    if not valid_results:
        print("제약 조건을 만족하는 결과가 없습니다.")
        return None

    # 최적 조건 선택
    if target == "min_time":
        optimal = min(valid_results, key=lambda x: x.total_time)
    elif target == "min_refills":
        optimal = min(valid_results, key=lambda x: x.num_refills)
    elif target == "max_throughput":
        # throughput = reaction_volume / total_time
        optimal = max(valid_results, key=lambda x: 5.0 / (x.total_time / 60))
    else:
        optimal = valid_results[0]

    return {
        "optimal": optimal,
        "all_results": results,
        "target": target,
        "constraints": constraints
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 5. 메인 대화형 인터페이스
# ═══════════════════════════════════════════════════════════════════════════════

def interactive_mode():
    """대화형 모드"""
    print("\n" + "=" * 60)
    print("FLOW CHEMISTRY SIMULATION TOOL")
    print("=" * 60)

    while True:
        print("\n[메뉴]")
        print("  1. 새 시뮬레이션 실행")
        print("  2. 저장된 설정 불러오기")
        print("  3. 파라미터 스윕 (조건 비교)")
        print("  4. 최적 조건 탐색")
        print("  5. 빠른 시뮬레이션 (기본값)")
        print("  0. 종료")

        choice = input("\n선택: ").strip()

        if choice == "0":
            print("종료합니다.")
            break

        elif choice == "1":
            # 새 시뮬레이션
            groups = input_groups()
            experiment = input_experiment()
            params = input_system_params()

            # 저장 여부
            save = input("\n이 설정을 저장하시겠습니까? (y/N): ").strip().lower()
            if save == "y":
                name = input("설정 이름: ").strip()
                if name:
                    save_config(name, groups, experiment, params)
                    print(f"설정 저장됨: {name}")

            # 시뮬레이션 실행
            print("\n시뮬레이션 실행 중...")
            events, total_time, flows, tracker = run_detailed_simulation(groups, experiment, params)
            print_summary_table(events, flows, groups, total_time)

            # Excel 저장
            if HAS_OPENPYXL:
                export_to_excel(events, flows, groups, experiment, params, total_time, tracker=tracker)

        elif choice == "2":
            # 저장된 설정 불러오기
            configs = list_saved_configs()
            if not configs:
                print("저장된 설정이 없습니다.")
                continue

            print("\n저장된 설정:")
            for i, name in enumerate(configs, 1):
                print(f"  {i}. {name}")

            idx = input_int("불러올 설정 번호", 1) - 1
            if 0 <= idx < len(configs):
                groups, experiment, params = load_config(configs[idx])
                print(f"\n'{configs[idx]}' 설정 불러옴")

                # 시뮬레이션 실행
                events, total_time, flows, tracker = run_detailed_simulation(groups, experiment, params)
                print_summary_table(events, flows, groups, total_time)

                if HAS_OPENPYXL:
                    export_to_excel(events, flows, groups, experiment, params, total_time, tracker=tracker)

        elif choice == "3":
            # 파라미터 스윕
            print("\n[파라미터 스윕]")
            print("스윕할 파라미터:")
            print("  1. residence_time (체류 시간)")
            print("  2. reaction_volume (반응량)")
            print("  3. temperature (온도)")

            param_choice = input("선택 [1]: ").strip() or "1"
            param_map = {"1": "residence_time", "2": "reaction_volume", "3": "temperature"}
            sweep_param = param_map.get(param_choice, "residence_time")

            start_val = input_float("시작 값", 5.0)
            end_val = input_float("끝 값", 15.0)
            step_val = input_float("간격", 2.0)

            sweep_values = []
            v = start_val
            while v <= end_val:
                sweep_values.append(v)
                v += step_val

            # 기본 그룹 설정
            groups = [
                GroupConfig("Group A", ReagentConfig("Reagent A", 3, 1.0, 1.0), 10.0, 40.0),
                GroupConfig("Group B", ReagentConfig("Reagent B", 5, 0.5, 1.2), 10.0, 40.0),
                GroupConfig("Group C", ReagentConfig("Catalyst", 7, 0.1, 0.05), 5.0, 30.0),
            ]

            base_experiment = ExperimentParams(80.0, 10.0, 5.0, 1.5, 1)
            params = SystemParams()

            results = run_parameter_sweep(groups, base_experiment, params, sweep_param, sweep_values)

            # 결과 출력
            print("\n[스윕 결과 요약]")
            print(f"{'#':<4} {sweep_param:<15} {'Total Time':<15} {'Refills':<10}")
            print("-" * 50)
            for r in results:
                print(f"{r.condition_id:<4} {list(r.params.values())[0]:<15.2f} "
                      f"{r.total_time:<15.1f} {r.num_refills:<10}")

            if HAS_OPENPYXL:
                export_sweep_results(results, groups, sweep_param)

        elif choice == "4":
            # 최적 조건 탐색
            print("\n[최적 조건 탐색]")
            print("최적화 목표:")
            print("  1. 최소 시퀀스 시간")
            print("  2. 최소 리필 횟수")
            print("  3. 최대 처리량")

            target_choice = input("선택 [1]: ").strip() or "1"
            target_map = {"1": "min_time", "2": "min_refills", "3": "max_throughput"}
            target = target_map.get(target_choice, "min_time")

            # 제약 조건
            constraints = {}
            max_time = input("최대 허용 시간 (초, Enter=무제한): ").strip()
            if max_time:
                constraints["max_total_time"] = float(max_time)

            max_refills = input("최대 리필 횟수 (Enter=무제한): ").strip()
            if max_refills:
                constraints["max_refills"] = int(max_refills)

            # 기본 그룹 설정
            groups = [
                GroupConfig("Group A", ReagentConfig("Reagent A", 3, 1.0, 1.0), 10.0, 40.0),
                GroupConfig("Group B", ReagentConfig("Reagent B", 5, 0.5, 1.2), 10.0, 40.0),
                GroupConfig("Group C", ReagentConfig("Catalyst", 7, 0.1, 0.05), 5.0, 30.0),
            ]
            params = SystemParams()

            result = find_optimal_conditions(groups, params, target, constraints)

            if result:
                opt = result["optimal"]
                print("\n" + "=" * 60)
                print("최적 조건 발견!")
                print("=" * 60)
                print(f"  Residence Time: {list(opt.params.values())[0]:.1f} min")
                print(f"  Total Time: {opt.total_time:.1f}s ({opt.total_time/60:.2f} min)")
                print(f"  Total Flow: {opt.total_flow:.4f} mL/min")
                print(f"  Refill Count: {opt.num_refills}")

                if HAS_OPENPYXL:
                    export_sweep_results(result["all_results"], groups, "residence_time")

        elif choice == "5":
            # 빠른 시뮬레이션
            print("\n[빠른 시뮬레이션 - 기본값 사용]")

            groups = [
                GroupConfig("Group A", ReagentConfig("Reagent A", 3, 1.0, 1.0), 10.0, 40.0),
                GroupConfig("Group B", ReagentConfig("Reagent B", 5, 0.5, 1.2), 10.0, 40.0),
                GroupConfig("Group C", ReagentConfig("Catalyst", 7, 0.1, 0.05), 5.0, 30.0),
            ]

            experiment = ExperimentParams(80.0, 10.0, 5.0, 1.5, 1)
            params = SystemParams()

            events, total_time, flows, tracker = run_detailed_simulation(groups, experiment, params)
            print_summary_table(events, flows, groups, total_time)

            if HAS_OPENPYXL:
                export_to_excel(events, flows, groups, experiment, params, total_time, tracker=tracker)


# ═══════════════════════════════════════════════════════════════════════════════
# 6. 메인
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if "--sweep" in sys.argv:
        # 파라미터 스윕 모드
        print("파라미터 스윕 모드")
        groups = [
            GroupConfig("Group A", ReagentConfig("Reagent A", 3, 1.0, 1.0), 10.0, 40.0),
            GroupConfig("Group B", ReagentConfig("Reagent B", 5, 0.5, 1.2), 10.0, 40.0),
            GroupConfig("Group C", ReagentConfig("Catalyst", 7, 0.1, 0.05), 5.0, 30.0),
        ]
        base_experiment = ExperimentParams(80.0, 10.0, 5.0, 1.5, 1)
        params = SystemParams()

        rt_values = [5, 7, 10, 12, 15, 20]
        results = run_parameter_sweep(groups, base_experiment, params, "residence_time", rt_values)

        if HAS_OPENPYXL:
            export_sweep_results(results, groups, "residence_time")

    else:
        # 대화형 모드
        interactive_mode()

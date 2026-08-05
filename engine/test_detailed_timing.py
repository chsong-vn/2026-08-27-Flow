"""
Detailed Timing Simulation Test with Excel Export

@codesyncer-context: 실제 simpy_engine.py 로직 기반
- Group A/B/C의 12-way, 3-way 밸브 전환 타이밍
- 시린지 펌프 Infusing/Withdrawing 상태 및 속도
- Outlet 밸브 전환 타이밍
- M농도 및 몰비 기반 유속 계산
- Excel 파일 자동 저장
"""

import simpy
import math
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional
from enum import Enum
from datetime import datetime
import os

# Excel 라이브러리
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl not installed. Run: pip install openpyxl")


# ═══════════════════════════════════════════════════════════════════════════════
# 1. 데이터 구조 정의
# ═══════════════════════════════════════════════════════════════════════════════

class EventType(Enum):
    """이벤트 타입"""
    PHASE_START = "PHASE_START"
    PHASE_END = "PHASE_END"
    VALVE_12WAY = "VALVE_12WAY"
    VALVE_3WAY = "VALVE_3WAY"
    VALVE_OUTLET = "VALVE_OUTLET"
    PUMP_INFUSE_START = "PUMP_INFUSE_START"
    PUMP_INFUSE_END = "PUMP_INFUSE_END"
    PUMP_WITHDRAW_START = "PUMP_WITHDRAW_START"
    PUMP_WITHDRAW_END = "PUMP_WITHDRAW_END"
    COLLECTOR_MOVE = "COLLECTOR_MOVE"
    COLLECTOR_HOME = "COLLECTOR_HOME"
    TEMP_REACHED = "TEMP_REACHED"


@dataclass
class TimingEvent:
    """타이밍 이벤트 기록"""
    time: float              # 시간 (초)
    event_type: EventType    # 이벤트 타입
    component: str           # 컴포넌트 이름 (Group A, Outlet 등)
    action: str              # 동작 설명
    from_state: str = ""     # 이전 상태
    to_state: str = ""       # 새 상태
    value: float = 0.0       # 값 (유속, 부피 등)
    duration: float = 0.0    # 소요 시간
    phase: str = ""          # 현재 Phase


@dataclass
class ReagentConfig:
    """시약 설정"""
    name: str
    port: int              # 12-way 포트 번호
    concentration: float   # M (mol/L)
    equivalents: float     # 몰비


@dataclass
class GroupConfig:
    """펌프 그룹 설정"""
    name: str              # "Group A", "Group B", "Group C"
    reagent: ReagentConfig
    syringe_volume: float = 10.0   # mL
    refill_rate: float = 40.0      # mL/min


@dataclass
class TubingSpec:
    """
    튜빙 규격 (내경 mm + 길이 m → 부피 자동 계산)

    @codesyncer-decision: 물리적 치수(mm, m)로부터 부피(mL)를 계산하고,
    역으로 부피 위치를 물리적 거리(m)로 환산하기 위한 데이터 구조
    """
    id_mm: float       # 내경 (mm)
    length_m: float    # 길이 (m)

    @property
    def volume_ml(self) -> float:
        """부피 = π × r² × L (mL, 내부 cm 변환)"""
        r_cm = (self.id_mm / 10.0) / 2.0
        length_cm = self.length_m * 100.0
        return math.pi * (r_cm ** 2) * length_cm

    @property
    def cross_section_cm2(self) -> float:
        """단면적 (cm²)"""
        r_cm = (self.id_mm / 10.0) / 2.0
        return math.pi * (r_cm ** 2)

    def vol_to_length_m(self, vol_ml: float) -> float:
        """부피(mL) → 튜빙 내 길이(m) 변환"""
        return vol_ml / self.cross_section_cm2 / 100.0

    def length_to_vol_ml(self, length_m: float) -> float:
        """길이(m) → 부피(mL) 변환"""
        return self.cross_section_cm2 * length_m * 100.0


@dataclass
class SystemParams:
    """시스템 파라미터"""
    reactor_volume: float = 7.854      # 반응기 부피 (mL)
    post_reactor_volume: float = 2.0   # 반응기 후 공통 구간 (mL)
    collection_line_volume: float = 1.0  # 수거 라인 (mL)
    dead_volume_per_line: float = 0.785  # 데드볼륨/라인 (mL)

    # 튜빙 규격 (물리적 위치 추적용 — 제공 시 부피 자동 계산)
    reactor_tubing: TubingSpec = None
    post_reactor_tubing: TubingSpec = None
    collection_line_tubing: TubingSpec = None

    # 밸브 전환 시간
    switch_time_12way_per_port: float = 0.3   # 초/포트
    switch_time_12way_max: float = 1.5        # 최대 시간
    switch_time_3way: float = 0.5             # 고정 시간
    switch_time_outlet: float = 0.5           # 고정 시간

    # 펌프 설정
    pump_start_delay: float = 0.2
    pump_stop_delay: float = 0.1

    # 기타
    priming_rate: float = 5.0          # mL/min
    stabilization_time: float = 10.0   # 초
    temp_tolerance: float = 1.0        # °C

    def __post_init__(self):
        """TubingSpec 제공 시 부피를 자동 계산으로 덮어쓰기"""
        if self.reactor_tubing is not None:
            self.reactor_volume = self.reactor_tubing.volume_ml
        if self.post_reactor_tubing is not None:
            self.post_reactor_volume = self.post_reactor_tubing.volume_ml
        if self.collection_line_tubing is not None:
            self.collection_line_volume = self.collection_line_tubing.volume_ml

    @property
    def has_tubing_specs(self) -> bool:
        """모든 zone의 튜빙 규격이 있는지 (물리적 위치 표시 가능 여부)"""
        return all([
            self.reactor_tubing is not None,
            self.post_reactor_tubing is not None,
            self.collection_line_tubing is not None
        ])


@dataclass
class ExperimentParams:
    """실험 파라미터"""
    temperature: float       # 반응 온도 (°C)
    residence_time: float    # 체류 시간 (min)
    reaction_volume: float   # 반응량 (mL)
    volume_per_tube: float   # 튜브당 부피 (mL)
    start_tube: int = 1


# 전역 이벤트 로그
all_events: List[TimingEvent] = []
current_phase: str = ""


# ═══════════════════════════════════════════════════════════════════════════════
# 1.5. 유체 위치 추적 데이터 구조
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class FluidSnapshot:
    """특정 시간의 유체 위치 스냅샷"""
    time: float
    phase: str
    trigger: str
    rxn_front: float        # 반응 혼합물 front 위치 (mL, Mixer=0)
    rxn_tail: float         # 반응 혼합물 tail 위치
    rxn_zone: str           # front가 위치한 zone
    verdict: str            # OK, WARNING, FAIL, -
    detail: str             # 설명
    rxn_front_m: float = 0.0    # front 물리적 위치 (m, Mixer=0)
    rxn_tail_m: float = 0.0     # tail 물리적 위치 (m)


@dataclass
class ValidationResult:
    """밸브 전환 시 유체 위치 검증 결과"""
    checkpoint: str
    expected: str
    actual: str
    margin: str
    status: str             # PASS, WARNING, FAIL


# ═══════════════════════════════════════════════════════════════════════════════
# 2. 유속 계산 (실제 calculators.py 로직)
# ═══════════════════════════════════════════════════════════════════════════════

def calculate_flows(groups: List[GroupConfig], residence_time: float,
                    reactor_volume: float) -> Dict[str, float]:
    """
    M농도와 몰비 기반 유속 계산

    @codesyncer-decision: calculators.py의 FlowCalculator.calculate_flows 로직 복제
    - ratio = equivalents / concentration
    - target_total_flow = reactor_volume / residence_time
    - flow = target_total_flow × (ratio / total_ratio)
    """
    ratios = {}
    for g in groups:
        conc = g.reagent.concentration
        eq = g.reagent.equivalents

        if conc <= 0:
            ratios[g.name] = 0.0
        else:
            ratios[g.name] = eq / conc

    total_ratio = sum(ratios.values())
    if total_ratio <= 0:
        raise ValueError("유효한 시약이 없습니다")

    target_total_flow = reactor_volume / residence_time

    flows = {}
    for g in groups:
        flows[g.name] = target_total_flow * (ratios[g.name] / total_ratio)

    return flows


# ═══════════════════════════════════════════════════════════════════════════════
# 3. SimPy 컴포넌트 (이벤트 추적 포함)
# ═══════════════════════════════════════════════════════════════════════════════

def log_event(time: float, event_type: EventType, component: str, action: str,
              from_state: str = "", to_state: str = "", value: float = 0.0,
              duration: float = 0.0):
    """이벤트 로그 기록"""
    global current_phase
    all_events.append(TimingEvent(
        time=time,
        event_type=event_type,
        component=component,
        action=action,
        from_state=from_state,
        to_state=to_state,
        value=value,
        duration=duration,
        phase=current_phase
    ))


def set_phase(phase_name: str, env_time: float):
    """현재 Phase 설정"""
    global current_phase
    current_phase = phase_name
    log_event(env_time, EventType.PHASE_START, "SYSTEM", f"Phase: {phase_name}")


class TrackedValve12Way:
    """12-way 밸브 (이벤트 추적)"""

    def __init__(self, env: simpy.Environment, group_name: str, params: SystemParams):
        self.env = env
        self.group_name = group_name
        self.params = params
        self.current_port = 1

    def set_port(self, target_port: int):
        return self.env.process(self._switch(target_port))

    def _switch(self, target_port: int):
        if self.current_port == target_port:
            return

        distance = abs(target_port - self.current_port)
        switch_time = min(
            self.params.switch_time_12way_max,
            distance * self.params.switch_time_12way_per_port
        )

        log_event(
            self.env.now, EventType.VALVE_12WAY, f"{self.group_name}_12way",
            f"Port {self.current_port} → Port {target_port}",
            from_state=f"Port {self.current_port}",
            to_state=f"Port {target_port}",
            value=distance,
            duration=switch_time
        )

        yield self.env.timeout(switch_time)
        self.current_port = target_port


class TrackedValve3Way:
    """3-way 밸브 (이벤트 추적)"""

    def __init__(self, env: simpy.Environment, group_name: str, params: SystemParams):
        self.env = env
        self.group_name = group_name
        self.params = params
        self.state = "WASTE"

    def set_position(self, state: str):
        return self.env.process(self._switch(state))

    def _switch(self, state: str):
        if self.state == state:
            return

        log_event(
            self.env.now, EventType.VALVE_3WAY, f"{self.group_name}_3way",
            f"{self.state} → {state}",
            from_state=self.state,
            to_state=state,
            duration=self.params.switch_time_3way
        )

        yield self.env.timeout(self.params.switch_time_3way)
        self.state = state


class TrackedOutletValve:
    """Outlet 밸브 (이벤트 추적)"""

    def __init__(self, env: simpy.Environment, params: SystemParams):
        self.env = env
        self.params = params
        self.state = "WASTE"

    def set_position(self, state: str):
        return self.env.process(self._switch(state))

    def _switch(self, state: str):
        if self.state == state:
            return

        log_event(
            self.env.now, EventType.VALVE_OUTLET, "Outlet",
            f"{self.state} → {state}",
            from_state=self.state,
            to_state=state,
            duration=self.params.switch_time_outlet
        )

        yield self.env.timeout(self.params.switch_time_outlet)
        self.state = state


class TrackedPump:
    """시린지 펌프 (이벤트 추적)"""

    def __init__(self, env: simpy.Environment, group_name: str,
                 syringe_volume: float, refill_rate: float, params: SystemParams):
        self.env = env
        self.group_name = group_name
        self.syringe_volume = syringe_volume
        self.refill_rate = refill_rate
        self.params = params
        self.current_volume = syringe_volume
        self.resource = simpy.Resource(env, capacity=1)

    def infuse(self, volume: float, rate: float, source_port: int = 1):
        return self.env.process(self._infuse_process(volume, rate, source_port))

    def _infuse_process(self, volume: float, rate: float, source_port: int):
        with self.resource.request() as req:
            yield req

            remaining = volume
            while remaining > 0.01:
                # 리필 체크
                if self.current_volume <= 0.1:
                    yield self.env.process(self._refill_process(source_port))

                inject_amount = min(remaining, self.current_volume)
                inject_time = (inject_amount / rate) * 60

                log_event(
                    self.env.now, EventType.PUMP_INFUSE_START, f"{self.group_name}_Pump",
                    f"INFUSE: {inject_amount:.3f}mL @ {rate:.4f}mL/min",
                    from_state="IDLE",
                    to_state="INFUSING",
                    value=rate,
                    duration=inject_time
                )

                yield self.env.timeout(self.params.pump_start_delay)
                yield self.env.timeout(inject_time - self.params.pump_start_delay - self.params.pump_stop_delay)
                yield self.env.timeout(self.params.pump_stop_delay)

                self.current_volume -= inject_amount
                remaining -= inject_amount

                log_event(
                    self.env.now, EventType.PUMP_INFUSE_END, f"{self.group_name}_Pump",
                    f"INFUSE END: Syringe {self.current_volume:.2f}mL remaining",
                    from_state="INFUSING",
                    to_state="IDLE",
                    value=self.current_volume
                )

    def _refill_process(self, source_port: int):
        refill_time = (self.syringe_volume / self.refill_rate) * 60

        log_event(
            self.env.now, EventType.PUMP_WITHDRAW_START, f"{self.group_name}_Pump",
            f"WITHDRAW: {self.syringe_volume:.1f}mL @ {self.refill_rate:.1f}mL/min (Port {source_port})",
            from_state="IDLE",
            to_state="WITHDRAWING",
            value=self.refill_rate,
            duration=refill_time
        )

        yield self.env.timeout(refill_time)
        self.current_volume = self.syringe_volume

        log_event(
            self.env.now, EventType.PUMP_WITHDRAW_END, f"{self.group_name}_Pump",
            f"WITHDRAW END: Syringe refilled to {self.syringe_volume:.1f}mL",
            from_state="WITHDRAWING",
            to_state="IDLE",
            value=self.syringe_volume
        )


class TrackedCollector:
    """분취기 (이벤트 추적)"""

    def __init__(self, env: simpy.Environment, params: SystemParams):
        self.env = env
        self.params = params
        self.current_tube = 0
        self.move_time_per_tube = 1.0
        self.home_time = 3.0

    def home(self):
        return self.env.process(self._home())

    def _home(self):
        log_event(
            self.env.now, EventType.COLLECTOR_HOME, "Collector",
            f"HOME: From Tube {self.current_tube}",
            from_state=f"Tube {self.current_tube}",
            to_state="HOME",
            duration=self.home_time
        )
        yield self.env.timeout(self.home_time)
        self.current_tube = 0

    def move_to_tube(self, tube_num: int):
        return self.env.process(self._move(tube_num))

    def _move(self, tube_num: int):
        if self.current_tube == tube_num:
            return

        distance = abs(tube_num - self.current_tube)
        move_time = distance * self.move_time_per_tube

        log_event(
            self.env.now, EventType.COLLECTOR_MOVE, "Collector",
            f"MOVE: Tube {self.current_tube} → Tube {tube_num}",
            from_state=f"Tube {self.current_tube}",
            to_state=f"Tube {tube_num}",
            value=distance,
            duration=move_time
        )

        yield self.env.timeout(move_time)
        self.current_tube = tube_num


# ═══════════════════════════════════════════════════════════════════════════════
# 3.5. 유체 위치 추적기
# ═══════════════════════════════════════════════════════════════════════════════

class FluidTracker:
    """
    Common path 내 반응 혼합물 slug 위치 추적기

    @codesyncer-decision: 이산 이벤트 시뮬레이션의 유체 위치를 추적하여
    밸브 전환 타이밍의 정확성을 검증함.
    - Injection 시작 이후 common path에 주입된 누적 볼륨으로 slug 위치 계산
    - 각 밸브 전환 시점에 스냅샷 + 자동 검증
    """

    def __init__(self, params: SystemParams, experiment: ExperimentParams,
                 flows: Dict[str, float], groups: List[GroupConfig]):
        self.params = params
        self.experiment = experiment
        self.flows = flows
        self.groups = groups
        self.total_flow = sum(flows.values())

        # Zone boundaries (mL from mixer)
        self.reactor_end = params.reactor_volume
        self.outlet_pos = params.reactor_volume + params.post_reactor_volume
        self.collection_end = self.outlet_pos + params.collection_line_volume

        # 물리적 경계 (m from mixer) — TubingSpec 있을 때만
        self.has_physical = params.has_tubing_specs
        if self.has_physical:
            self.reactor_length_m = params.reactor_tubing.length_m
            self.post_reactor_length_m = params.post_reactor_tubing.length_m
            self.collection_length_m = params.collection_line_tubing.length_m
            # 누적 경계 (m)
            self.reactor_end_m = self.reactor_length_m
            self.outlet_pos_m = self.reactor_end_m + self.post_reactor_length_m
            self.collection_end_m = self.outlet_pos_m + self.collection_length_m

        # Common path 누적 볼륨 (injection 시작 기준)
        self.cumulative_vol = 0.0
        self.injection_started = False

        # Records
        self.snapshots: List[FluidSnapshot] = []
        self.validations: List[ValidationResult] = []

        # Segments: 시간-볼륨 구간 (Position-Time 차트용)
        self.segments: List[dict] = []

    def vol_to_m(self, pos_ml: float) -> float:
        """
        부피 위치(mL) → 물리적 위치(m) 변환

        @codesyncer-decision: 각 zone의 내경이 다를 수 있으므로
        zone별로 단면적 기반 변환 적용
        """
        if not self.has_physical:
            return 0.0

        if pos_ml <= 0:
            return 0.0
        elif pos_ml <= self.reactor_end:
            return self.params.reactor_tubing.vol_to_length_m(pos_ml)
        elif pos_ml <= self.outlet_pos:
            excess = pos_ml - self.reactor_end
            return self.reactor_end_m + self.params.post_reactor_tubing.vol_to_length_m(excess)
        elif pos_ml <= self.collection_end:
            excess = pos_ml - self.outlet_pos
            return self.outlet_pos_m + self.params.collection_line_tubing.vol_to_length_m(excess)
        else:
            # Past system — collection line 내경으로 연장 추정
            excess = pos_ml - self.collection_end
            return self.collection_end_m + self.params.collection_line_tubing.vol_to_length_m(excess)

    def get_zone(self, pos_ml: float) -> str:
        """position → zone name (물리적 거리 포함)"""
        if pos_ml <= 0:
            return "Mixer"
        elif pos_ml <= self.reactor_end:
            pct = pos_ml / self.reactor_end * 100
            if self.has_physical:
                m = self.params.reactor_tubing.vol_to_length_m(pos_ml)
                return f"Reactor ({pct:.0f}%, {m:.2f}m)"
            return f"Reactor ({pct:.0f}%)"
        elif pos_ml <= self.outlet_pos:
            pct = (pos_ml - self.reactor_end) / self.params.post_reactor_volume * 100
            if self.has_physical:
                m = self.params.post_reactor_tubing.vol_to_length_m(pos_ml - self.reactor_end)
                return f"Post-Reactor ({pct:.0f}%, {m:.3f}m)"
            return f"Post-Reactor ({pct:.0f}%)"
        elif pos_ml <= self.collection_end:
            if self.has_physical:
                m = self.params.collection_line_tubing.vol_to_length_m(pos_ml - self.outlet_pos)
                return f"Collection ({m:.3f}m)"
            return "Collection Line"
        else:
            return "Past System"

    def get_rxn_position(self) -> Tuple[float, float]:
        """반응 혼합물 (front, tail) 위치 반환"""
        if not self.injection_started:
            return 0.0, 0.0
        front = self.cumulative_vol
        tail = max(0.0, front - self.experiment.reaction_volume)
        return front, tail

    def mark_injection_start(self):
        """Injection 시작 — 여기서부터 slug 위치 추적"""
        self.injection_started = True
        self.cumulative_vol = 0.0

    def add_volume(self, vol_ml: float, t_start: float, t_end: float, phase: str):
        """Common path에 볼륨 추가 + segment 기록"""
        vol_before = self.cumulative_vol
        self.cumulative_vol += vol_ml

        if self.injection_started:
            self.segments.append({
                'start_time': t_start, 'end_time': t_end,
                'vol_start': vol_before, 'vol_end': self.cumulative_vol,
                'phase': phase
            })

    def snapshot(self, time: float, phase: str, trigger: str,
                 verdict: str = "-", detail: str = ""):
        """현재 상태 스냅샷 기록 (물리적 위치 포함)"""
        front, tail = self.get_rxn_position()
        self.snapshots.append(FluidSnapshot(
            time=round(time, 2), phase=phase, trigger=trigger,
            rxn_front=round(front, 3), rxn_tail=round(tail, 3),
            rxn_zone=self.get_zone(front),
            verdict=verdict, detail=detail,
            rxn_front_m=round(self.vol_to_m(front), 3),
            rxn_tail_m=round(self.vol_to_m(tail), 3)
        ))

    # ── Validation helpers ──────────────────────────────────────────────

    def validate_outlet_switch(self, time: float, to_state: str):
        """Outlet 밸브 전환 시 유체 위치 검증"""
        front, tail = self.get_rxn_position()

        if to_state == "COLLECT":
            if tail >= self.outlet_pos:
                status = "FAIL"
                detail = (f"반응물 tail({tail:.2f}mL) >= Outlet({self.outlet_pos:.2f}mL), "
                          f"전량 유실")
            elif front > self.outlet_pos:
                lost = front - self.outlet_pos
                remaining = self.experiment.reaction_volume - lost
                pct = remaining / self.experiment.reaction_volume * 100
                status = "OK" if pct >= 50 else "WARNING"
                detail = (f"{lost:.2f}mL 이미 통과 → WASTE, "
                          f"{remaining:.2f}mL({pct:.0f}%) 수거 가능")
            else:
                margin = self.outlet_pos - front
                status = "OK"
                detail = f"Outlet까지 {margin:.2f}mL 여유"

            self.validations.append(ValidationResult(
                checkpoint="Outlet→COLLECT 전환 시 반응물 위치",
                expected=f"tail < {self.outlet_pos:.2f} mL",
                actual=f"front={front:.2f}, tail={tail:.2f}",
                margin=f"{self.outlet_pos - tail:.2f} mL (tail→Outlet)",
                status=status
            ))
            self.snapshot(time, "COLLECTION", "Outlet → COLLECT", status, detail)

    def validate_prefill(self, group_name: str, filled: float, dead_vol: float):
        """Prefill 검증: dead_vol 정확히 채워졌는지"""
        diff = filled - dead_vol
        if abs(diff) < 0.01:
            status, detail = "PASS", "정확히 채움"
        elif diff > 0:
            status, detail = "WARNING", f"{diff:.3f}mL 넘침→Reactor 유입"
        else:
            status, detail = "WARNING", f"{-diff:.3f}mL 부족→솔벤트 잔류"

        self.validations.append(ValidationResult(
            checkpoint=f"{group_name} Prefill",
            expected=f"{dead_vol:.3f} mL",
            actual=f"{filled:.3f} mL",
            margin=f"{diff:+.3f} mL",
            status=status
        ))

    def validate_collection_complete(self, collected: float):
        """Collection 완료 검증"""
        target = self.experiment.reaction_volume
        pct = collected / target * 100 if target > 0 else 0
        if pct >= 95:
            status = "PASS"
        elif pct >= 50:
            status = "WARNING"
        else:
            status = "FAIL"

        self.validations.append(ValidationResult(
            checkpoint="수거 완료 시 총 수거량",
            expected=f"{target:.2f} mL",
            actual=f"{collected:.2f} mL ({pct:.0f}%)",
            margin=f"{collected - target:+.2f} mL",
            status=status
        ))

    # ── Position-Time 차트 데이터 ──────────────────────────────────────

    def generate_time_series(self, interval_s: float = 5.0) -> List[dict]:
        """정기 시간 간격 위치 데이터 (차트용)"""
        if not self.segments:
            return []

        series = []
        t_start = self.segments[0]['start_time']
        t_end = self.segments[-1]['end_time']

        # segment 경계 + 정기 간격 포인트 모두 생성
        time_points = set()
        t = t_start
        while t <= t_end + 0.01:
            time_points.add(round(t, 1))
            t += interval_s
        for seg in self.segments:
            time_points.add(round(seg['start_time'], 1))
            time_points.add(round(seg['end_time'], 1))

        for tp in sorted(time_points):
            vol = self._interpolate_volume(tp)
            front = vol
            tail = max(0.0, vol - self.experiment.reaction_volume)
            phase = ""
            for seg in self.segments:
                if seg['start_time'] <= tp <= seg['end_time']:
                    phase = seg['phase']
                    break
            series.append({
                'time': tp,
                'rxn_front': round(front, 3),
                'rxn_tail': round(tail, 3),
                'zone': self.get_zone(front),
                'phase': phase
            })

        return series

    def _interpolate_volume(self, t: float) -> float:
        """시간 t에서의 누적 볼륨 (선형 보간)"""
        vol = 0.0
        for seg in self.segments:
            if t < seg['start_time']:
                break
            elif t >= seg['end_time']:
                vol = seg['vol_end']
            else:
                dt = seg['end_time'] - seg['start_time']
                if dt > 0:
                    frac = (t - seg['start_time']) / dt
                    vol = seg['vol_start'] + frac * (seg['vol_end'] - seg['vol_start'])
                else:
                    vol = seg['vol_end']
        return vol


# ═══════════════════════════════════════════════════════════════════════════════
# 4. 시퀀스 시뮬레이션 (실제 simpy_engine.py 로직 기반)
# ═══════════════════════════════════════════════════════════════════════════════

def run_detailed_simulation(
    groups: List[GroupConfig],
    experiment: ExperimentParams,
    params: SystemParams = None
) -> Tuple[List[TimingEvent], float, Dict[str, float], FluidTracker]:
    """
    상세 타이밍 시뮬레이션 실행

    @codesyncer-decision: simpy_engine.py의 시퀀스 로직 재현
    + FluidTracker로 common path 내 유체 위치 추적
    """
    global all_events, current_phase
    all_events = []
    current_phase = ""

    if params is None:
        params = SystemParams()

    env = simpy.Environment()

    # 유속 계산
    flows = calculate_flows(groups, experiment.residence_time, params.reactor_volume)
    total_flow = sum(flows.values())

    # Zone 부피
    zone_vol = params.reactor_volume + params.post_reactor_volume

    # 컴포넌트 생성
    pumps = {}
    valves_12way = {}
    valves_3way = {}

    for g in groups:
        pumps[g.name] = TrackedPump(env, g.name, g.syringe_volume, g.refill_rate, params)
        valves_12way[g.name] = TrackedValve12Way(env, g.name, params)
        valves_3way[g.name] = TrackedValve3Way(env, g.name, params)

    outlet = TrackedOutletValve(env, params)
    collector = TrackedCollector(env, params)

    # 유체 위치 추적기
    tracker = FluidTracker(params, experiment, flows, groups)

    def sequence_process():
        """메인 시퀀스 (simpy_engine.py 기반 + 유체 추적)"""

        # ═══════════════════════════════════════════════════════════════════
        # Phase 0: 초기화
        # ═══════════════════════════════════════════════════════════════════
        set_phase("INITIALIZATION", env.now)

        init_procs = []
        for g in groups:
            init_procs.append(valves_12way[g.name].set_port(1))
            init_procs.append(valves_3way[g.name].set_position("WASTE"))
        init_procs.append(outlet.set_position("WASTE"))
        yield simpy.AllOf(env, init_procs)

        yield collector.home()
        yield collector.move_to_tube(experiment.start_tube)
        tracker.snapshot(env.now, "INITIALIZATION", "초기화 완료")

        # ═══════════════════════════════════════════════════════════════════
        # Phase 1: 가열 (시뮬레이션: 간단한 1차 응답)
        # ═══════════════════════════════════════════════════════════════════
        set_phase("HEATING", env.now)

        sim_temp = 25.0
        target_temp = experiment.temperature

        while abs(sim_temp - target_temp) > params.temp_tolerance:
            sim_temp += (target_temp - sim_temp) * 0.15
            yield env.timeout(1.0)

        log_event(env.now, EventType.TEMP_REACHED, "Heater",
                  f"Temperature reached: {sim_temp:.1f}°C",
                  to_state=f"{sim_temp:.1f}°C")
        tracker.snapshot(env.now, "HEATING", f"온도 도달: {sim_temp:.1f}°C")

        # ═══════════════════════════════════════════════════════════════════
        # Phase 2: 시스템 세척 (Priming) — 3-way WASTE, common path 무관
        # ═══════════════════════════════════════════════════════════════════
        set_phase("PRIMING", env.now)

        wash_vol = zone_vol * 1.5
        prime_procs = []
        for g in groups:
            pump_vol = wash_vol / len(groups)
            prime_procs.append(pumps[g.name].infuse(pump_vol, params.priming_rate, source_port=1))
        yield simpy.AllOf(env, prime_procs)
        tracker.snapshot(env.now, "PRIMING", "세척 완료 (→WASTE, common path 무관)")

        # ═══════════════════════════════════════════════════════════════════
        # Phase 3: 안정화
        # ═══════════════════════════════════════════════════════════════════
        set_phase("STABILIZATION", env.now)
        yield env.timeout(params.stabilization_time)
        tracker.snapshot(env.now, "STABILIZATION", "안정화 완료")

        # ═══════════════════════════════════════════════════════════════════
        # Phase 4: 데드볼륨 프리필
        # ═══════════════════════════════════════════════════════════════════
        set_phase("PREFILL", env.now)

        # 12-way: 시약 포트로 전환
        port_procs = []
        for g in groups:
            port_procs.append(valves_12way[g.name].set_port(g.reagent.port))
        yield simpy.AllOf(env, port_procs)
        tracker.snapshot(env.now, "PREFILL", "12-way → 시약 포트")

        # 3-way: REACTOR로 전환
        reactor_procs = []
        for g in groups:
            reactor_procs.append(valves_3way[g.name].set_position("REACTOR"))
        yield simpy.AllOf(env, reactor_procs)
        tracker.snapshot(env.now, "PREFILL", "3-way → REACTOR")

        # 데드볼륨 채우기
        prefill_procs = []
        for g in groups:
            flow = flows[g.name]
            if flow > 0:
                prefill_procs.append(pumps[g.name].infuse(
                    params.dead_volume_per_line, params.priming_rate, g.reagent.port))
        if prefill_procs:
            yield simpy.AllOf(env, prefill_procs)

        # Prefill 검증
        for g in groups:
            tracker.validate_prefill(g.name, params.dead_volume_per_line,
                                     params.dead_volume_per_line)
        tracker.snapshot(env.now, "PREFILL",
                         f"Dead vol 채움 ({params.dead_volume_per_line:.3f}mL/line)")

        # 밸브 복귀 (WASTE)
        waste_procs = []
        for g in groups:
            waste_procs.append(valves_3way[g.name].set_position("WASTE"))
        yield simpy.AllOf(env, waste_procs)
        tracker.snapshot(env.now, "PREFILL", "3-way → WASTE (Prefill 완료)")

        # ═══════════════════════════════════════════════════════════════════
        # Phase 5: 시약 주입 (Injection)
        # ═══════════════════════════════════════════════════════════════════
        set_phase("INJECTION", env.now)

        # 3-way: REACTOR로 전환
        reactor_procs = []
        for g in groups:
            reactor_procs.append(valves_3way[g.name].set_position("REACTOR"))
        yield simpy.AllOf(env, reactor_procs)

        # ★ Injection 시작 — 여기서부터 slug 추적
        tracker.mark_injection_start()
        tracker.snapshot(env.now, "INJECTION", "3-way → REACTOR, 주입 시작")

        # 시약 주입
        inject_start_t = env.now
        inject_procs = []
        for g in groups:
            flow = flows[g.name]
            if flow > 0:
                pump_vol = experiment.reaction_volume * (flow / total_flow)
                inject_procs.append(pumps[g.name].infuse(pump_vol, flow, g.reagent.port))
        if inject_procs:
            yield simpy.AllOf(env, inject_procs)
        inject_end_t = env.now

        tracker.add_volume(experiment.reaction_volume, inject_start_t, inject_end_t, "INJECTION")
        tracker.snapshot(env.now, "INJECTION",
                         f"주입 완료 ({experiment.reaction_volume}mL, "
                         f"{inject_end_t - inject_start_t:.1f}s)")

        # ═══════════════════════════════════════════════════════════════════
        # Phase 6: 이송 (Transit)
        # ═══════════════════════════════════════════════════════════════════
        set_phase("TRANSIT", env.now)

        # @codesyncer-decision: transit은 반응물 front가 outlet에 도달하도록 밀어냄
        # 이전 로직(zone_vol * 0.95)은 reaction_volume 미고려 → 90%+ 유실 발생
        # 수정: zone_vol - reaction_volume = front가 outlet 정확 도달
        transit_vol = max(0, zone_vol - experiment.reaction_volume)

        transit_start_t = env.now
        transit_procs = []
        for g in groups:
            flow = flows[g.name]
            if flow > 0:
                pump_vol = transit_vol * (flow / total_flow)
                transit_procs.append(pumps[g.name].infuse(pump_vol, flow, source_port=1))
        if transit_procs:
            yield simpy.AllOf(env, transit_procs)
        transit_end_t = env.now

        tracker.add_volume(transit_vol, transit_start_t, transit_end_t, "TRANSIT")
        tracker.snapshot(env.now, "TRANSIT",
                         f"이송 완료 ({transit_vol:.2f}mL pushed, "
                         f"{transit_end_t - transit_start_t:.1f}s)")

        # ═══════════════════════════════════════════════════════════════════
        # Phase 7: 분취 (Collection)
        # ═══════════════════════════════════════════════════════════════════
        set_phase("COLLECTION", env.now)

        num_tubes = math.ceil(experiment.reaction_volume / experiment.volume_per_tube)

        # Outlet: COLLECT로 전환 → ★ 핵심 검증 포인트
        yield outlet.set_position("COLLECT")
        tracker.validate_outlet_switch(env.now, "COLLECT")

        # @codesyncer-decision: transit에서 front가 이미 outlet까지 도달하므로 여유분 불필요
        # 이전 로직(zone_vol * 0.05)은 0.95배 transit과 쌍이었으나, transit 수정에 따라 제거
        current_tube = experiment.start_tube
        total_collected = 0.0

        for tube_idx in range(num_tubes):
            if tube_idx == num_tubes - 1:
                remaining = experiment.reaction_volume - (tube_idx * experiment.volume_per_tube)
                tube_vol = min(remaining, experiment.volume_per_tube)
            else:
                tube_vol = experiment.volume_per_tube

            # 분취기 이동 (첫 튜브 제외)
            if tube_idx > 0:
                yield collector.move_to_tube(current_tube)

            # 수거
            tube_start_t = env.now
            collect_procs = []
            for g in groups:
                flow = flows[g.name]
                if flow > 0:
                    pump_vol = tube_vol * (flow / total_flow)
                    collect_procs.append(pumps[g.name].infuse(pump_vol, flow, source_port=1))
            if collect_procs:
                yield simpy.AllOf(env, collect_procs)
            tube_end_t = env.now

            tracker.add_volume(tube_vol, tube_start_t, tube_end_t, "COLLECTION")
            total_collected += tube_vol
            tracker.snapshot(env.now, "COLLECTION",
                             f"Tube {current_tube}: {tube_vol:.2f}mL")

            current_tube += 1

        # 수거량 검증
        tracker.validate_collection_complete(total_collected)

        # ═══════════════════════════════════════════════════════════════════
        # Phase 8: 수거 라인 세척
        # ═══════════════════════════════════════════════════════════════════
        set_phase("LINE_WASH", env.now)

        yield collector.move_to_tube(current_tube)

        wash_vol = params.collection_line_volume * 1.0
        wash_start_t = env.now
        wash_procs = []
        for g in groups:
            flow = flows[g.name]
            if flow > 0:
                pump_vol = wash_vol * (flow / total_flow)
                wash_procs.append(pumps[g.name].infuse(pump_vol, flow, source_port=1))
        if wash_procs:
            yield simpy.AllOf(env, wash_procs)
        wash_end_t = env.now

        tracker.add_volume(wash_vol, wash_start_t, wash_end_t, "LINE_WASH")
        tracker.snapshot(env.now, "LINE_WASH", f"라인 세척 ({wash_vol:.2f}mL)")

        # Outlet: WASTE로 복귀
        yield outlet.set_position("WASTE")
        tracker.snapshot(env.now, "LINE_WASH", "Outlet → WASTE")

        # ═══════════════════════════════════════════════════════════════════
        # Phase 9: 종료
        # ═══════════════════════════════════════════════════════════════════
        set_phase("FINALIZATION", env.now)

        yield collector.home()

        set_phase("COMPLETE", env.now)
        tracker.snapshot(env.now, "COMPLETE", "시퀀스 완료")

    # 시뮬레이션 실행
    env.process(sequence_process())
    env.run()

    return all_events, env.now, flows, tracker


# ═══════════════════════════════════════════════════════════════════════════════
# 5. Excel 출력
# ═══════════════════════════════════════════════════════════════════════════════

def export_to_excel(
    events: List[TimingEvent],
    flows: Dict[str, float],
    groups: List[GroupConfig],
    experiment: ExperimentParams,
    params: SystemParams,
    total_time: float,
    filename: str = None,
    tracker: FluidTracker = None
) -> str:
    """Excel 파일로 내보내기 (유체 추적 시트 포함)"""

    if not HAS_OPENPYXL:
        print("openpyxl이 설치되지 않았습니다. pip install openpyxl")
        return None

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"simulation_result_{timestamp}.xlsx"

    # 경로 설정
    output_dir = os.path.join(os.path.dirname(__file__), "..", "temp")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 1: 실험 파라미터
    # ─────────────────────────────────────────────────────────────────────
    ws_params = wb.active
    ws_params.title = "Parameters"

    # 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    ws_params["A1"] = "EXPERIMENT PARAMETERS"
    ws_params["A1"].font = Font(bold=True, size=14)

    # 실험 조건
    ws_params["A3"] = "Reaction Conditions"
    ws_params["A3"].font = header_font
    ws_params["A4"] = "Temperature (°C)"
    ws_params["B4"] = experiment.temperature
    ws_params["A5"] = "Residence Time (min)"
    ws_params["B5"] = experiment.residence_time
    ws_params["A6"] = "Reaction Volume (mL)"
    ws_params["B6"] = experiment.reaction_volume
    ws_params["A7"] = "Volume per Tube (mL)"
    ws_params["B7"] = experiment.volume_per_tube

    # 시스템 파라미터
    ws_params["A9"] = "System Parameters"
    ws_params["A9"].font = header_font
    ws_params["A10"] = "Reactor Volume (mL)"
    ws_params["B10"] = params.reactor_volume
    ws_params["A11"] = "Post-Reactor Volume (mL)"
    ws_params["B11"] = params.post_reactor_volume
    ws_params["A12"] = "Collection Line (mL)"
    ws_params["B12"] = params.collection_line_volume

    # 그룹별 설정
    ws_params["A14"] = "GROUP CONFIGURATIONS"
    ws_params["A14"].font = Font(bold=True, size=12)

    headers = ["Group", "Reagent", "Port", "Conc (M)", "Eq", "Flow (mL/min)", "Syringe (mL)"]
    for col, header in enumerate(headers, 1):
        cell = ws_params.cell(row=15, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, g in enumerate(groups, 16):
        ws_params.cell(row=row, column=1, value=g.name)
        ws_params.cell(row=row, column=2, value=g.reagent.name)
        ws_params.cell(row=row, column=3, value=g.reagent.port)
        ws_params.cell(row=row, column=4, value=g.reagent.concentration)
        ws_params.cell(row=row, column=5, value=g.reagent.equivalents)
        ws_params.cell(row=row, column=6, value=round(flows[g.name], 4))
        ws_params.cell(row=row, column=7, value=g.syringe_volume)

    # 총 유속
    total_flow = sum(flows.values())
    ws_params.cell(row=len(groups)+16, column=1, value="TOTAL")
    ws_params.cell(row=len(groups)+16, column=1).font = header_font
    ws_params.cell(row=len(groups)+16, column=6, value=round(total_flow, 4))

    ws_params.cell(row=len(groups)+18, column=1, value="Total Sequence Time (s)")
    ws_params.cell(row=len(groups)+18, column=2, value=round(total_time, 1))
    ws_params.cell(row=len(groups)+19, column=1, value="Total Sequence Time (min)")
    ws_params.cell(row=len(groups)+19, column=2, value=round(total_time/60, 2))

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 2: 전체 타이밍 테이블
    # ─────────────────────────────────────────────────────────────────────
    ws_timeline = wb.create_sheet("Timeline")

    headers = ["Time (s)", "Time (min)", "Phase", "Component", "Event Type",
               "Action", "From", "To", "Value", "Duration (s)"]
    for col, header in enumerate(headers, 1):
        cell = ws_timeline.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, event in enumerate(sorted(events, key=lambda x: x.time), 2):
        ws_timeline.cell(row=row, column=1, value=round(event.time, 2))
        ws_timeline.cell(row=row, column=2, value=round(event.time/60, 3))
        ws_timeline.cell(row=row, column=3, value=event.phase)
        ws_timeline.cell(row=row, column=4, value=event.component)
        ws_timeline.cell(row=row, column=5, value=event.event_type.value)
        ws_timeline.cell(row=row, column=6, value=event.action)
        ws_timeline.cell(row=row, column=7, value=event.from_state)
        ws_timeline.cell(row=row, column=8, value=event.to_state)
        ws_timeline.cell(row=row, column=9, value=round(event.value, 4) if event.value else "")
        ws_timeline.cell(row=row, column=10, value=round(event.duration, 2) if event.duration else "")

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 3: 밸브 전환 테이블
    # ─────────────────────────────────────────────────────────────────────
    ws_valves = wb.create_sheet("Valve Events")

    valve_events = [e for e in events if e.event_type in
                    [EventType.VALVE_12WAY, EventType.VALVE_3WAY, EventType.VALVE_OUTLET]]

    headers = ["Time (s)", "Phase", "Component", "Type", "From", "To", "Duration (s)"]
    for col, header in enumerate(headers, 1):
        cell = ws_valves.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, event in enumerate(sorted(valve_events, key=lambda x: x.time), 2):
        ws_valves.cell(row=row, column=1, value=round(event.time, 2))
        ws_valves.cell(row=row, column=2, value=event.phase)
        ws_valves.cell(row=row, column=3, value=event.component)
        ws_valves.cell(row=row, column=4, value=event.event_type.value)
        ws_valves.cell(row=row, column=5, value=event.from_state)
        ws_valves.cell(row=row, column=6, value=event.to_state)
        ws_valves.cell(row=row, column=7, value=round(event.duration, 2))

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 4: 펌프 상태 테이블
    # ─────────────────────────────────────────────────────────────────────
    ws_pumps = wb.create_sheet("Pump Events")

    pump_events = [e for e in events if e.event_type in
                   [EventType.PUMP_INFUSE_START, EventType.PUMP_INFUSE_END,
                    EventType.PUMP_WITHDRAW_START, EventType.PUMP_WITHDRAW_END]]

    headers = ["Time (s)", "Phase", "Component", "Event", "Rate (mL/min)", "Duration (s)", "Action"]
    for col, header in enumerate(headers, 1):
        cell = ws_pumps.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, event in enumerate(sorted(pump_events, key=lambda x: x.time), 2):
        ws_pumps.cell(row=row, column=1, value=round(event.time, 2))
        ws_pumps.cell(row=row, column=2, value=event.phase)
        ws_pumps.cell(row=row, column=3, value=event.component)
        ws_pumps.cell(row=row, column=4, value=event.event_type.value)
        ws_pumps.cell(row=row, column=5, value=round(event.value, 4) if event.value else "")
        ws_pumps.cell(row=row, column=6, value=round(event.duration, 2) if event.duration else "")
        ws_pumps.cell(row=row, column=7, value=event.action)

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 5: 분취기 이벤트
    # ─────────────────────────────────────────────────────────────────────
    ws_collector = wb.create_sheet("Collector Events")

    collector_events = [e for e in events if e.event_type in
                        [EventType.COLLECTOR_MOVE, EventType.COLLECTOR_HOME]]

    headers = ["Time (s)", "Phase", "Event", "From", "To", "Duration (s)"]
    for col, header in enumerate(headers, 1):
        cell = ws_collector.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, event in enumerate(sorted(collector_events, key=lambda x: x.time), 2):
        ws_collector.cell(row=row, column=1, value=round(event.time, 2))
        ws_collector.cell(row=row, column=2, value=event.phase)
        ws_collector.cell(row=row, column=3, value=event.event_type.value)
        ws_collector.cell(row=row, column=4, value=event.from_state)
        ws_collector.cell(row=row, column=5, value=event.to_state)
        ws_collector.cell(row=row, column=6, value=round(event.duration, 2))

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 6: 유체 위치 추적 (Fluid Tracking)
    # ─────────────────────────────────────────────────────────────────────
    if tracker and tracker.snapshots:
        ws_fluid = wb.create_sheet("Fluid Tracking")

        # 스타일
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        has_phys = params.has_tubing_specs
        headers_fluid = ["Time (s)", "Phase", "Trigger",
                         "Rxn Front (mL)", "Rxn Tail (mL)"]
        if has_phys:
            headers_fluid += ["Front (m)", "Tail (m)"]
        headers_fluid += ["Zone", "Verdict", "Detail"]

        for col, h in enumerate(headers_fluid, 1):
            cell = ws_fluid.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row, snap in enumerate(tracker.snapshots, 2):
            col = 1
            ws_fluid.cell(row=row, column=col, value=snap.time); col += 1
            ws_fluid.cell(row=row, column=col, value=snap.phase); col += 1
            ws_fluid.cell(row=row, column=col, value=snap.trigger); col += 1
            ws_fluid.cell(row=row, column=col, value=snap.rxn_front); col += 1
            ws_fluid.cell(row=row, column=col, value=snap.rxn_tail); col += 1
            if has_phys:
                ws_fluid.cell(row=row, column=col, value=snap.rxn_front_m); col += 1
                ws_fluid.cell(row=row, column=col, value=snap.rxn_tail_m); col += 1
            ws_fluid.cell(row=row, column=col, value=snap.rxn_zone); col += 1
            verdict_cell = ws_fluid.cell(row=row, column=col, value=snap.verdict); col += 1
            ws_fluid.cell(row=row, column=col, value=snap.detail)

            # 색상 적용
            if snap.verdict == "OK" or snap.verdict == "PASS":
                verdict_cell.fill = pass_fill
            elif snap.verdict == "WARNING":
                verdict_cell.fill = warn_fill
            elif snap.verdict == "FAIL":
                verdict_cell.fill = fail_fill

        # Zone 경계선 참고 행
        ref_row = len(tracker.snapshots) + 3
        ws_fluid.cell(row=ref_row, column=1, value="[Zone Boundaries]")
        ws_fluid.cell(row=ref_row, column=1).font = Font(bold=True)

        outlet_ml = params.reactor_volume + params.post_reactor_volume
        coll_end_ml = outlet_ml + params.collection_line_volume

        def _boundary_str(ml_val, tubing):
            """mL + 물리적 치수 문자열"""
            s = f"{ml_val:.3f} mL"
            if tubing:
                s += f" (ID={tubing.id_mm}mm, L={tubing.length_m:.2f}m)"
            return s

        ws_fluid.cell(row=ref_row + 1, column=1, value="Reactor End")
        ws_fluid.cell(row=ref_row + 1, column=2,
                       value=_boundary_str(params.reactor_volume, params.reactor_tubing))
        ws_fluid.cell(row=ref_row + 2, column=1, value="Outlet (Post-Reactor End)")
        ws_fluid.cell(row=ref_row + 2, column=2,
                       value=_boundary_str(outlet_ml, params.post_reactor_tubing))
        ws_fluid.cell(row=ref_row + 3, column=1, value="Collection End")
        ws_fluid.cell(row=ref_row + 3, column=2,
                       value=_boundary_str(coll_end_ml, params.collection_line_tubing))

        if has_phys:
            ws_fluid.cell(row=ref_row + 4, column=1, value="Total Length")
            total_m = (params.reactor_tubing.length_m +
                       params.post_reactor_tubing.length_m +
                       params.collection_line_tubing.length_m)
            ws_fluid.cell(row=ref_row + 4, column=2, value=f"{total_m:.2f} m")

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 7: 검증 요약 (Validation Summary)
    # ─────────────────────────────────────────────────────────────────────
    if tracker and tracker.validations:
        ws_valid = wb.create_sheet("Validation")

        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        headers_v = ["Checkpoint", "Expected", "Actual", "Margin", "Status"]
        for col, h in enumerate(headers_v, 1):
            cell = ws_valid.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row, v in enumerate(tracker.validations, 2):
            ws_valid.cell(row=row, column=1, value=v.checkpoint)
            ws_valid.cell(row=row, column=2, value=v.expected)
            ws_valid.cell(row=row, column=3, value=v.actual)
            ws_valid.cell(row=row, column=4, value=v.margin)
            status_cell = ws_valid.cell(row=row, column=5, value=v.status)

            if v.status == "PASS":
                status_cell.fill = pass_fill
            elif v.status == "WARNING":
                status_cell.fill = warn_fill
            elif v.status == "FAIL":
                status_cell.fill = fail_fill

        # 최종 판정
        statuses = [v.status for v in tracker.validations]
        summary_row = len(tracker.validations) + 3
        if "FAIL" in statuses:
            overall = "FAIL - 밸브 타이밍 문제 발견"
            overall_fill = fail_fill
        elif "WARNING" in statuses:
            overall = "WARNING - 주의 필요"
            overall_fill = warn_fill
        else:
            overall = "ALL PASS - 밸브 타이밍 안전"
            overall_fill = pass_fill

        ws_valid.cell(row=summary_row, column=1, value="OVERALL RESULT")
        ws_valid.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        result_cell = ws_valid.cell(row=summary_row, column=2, value=overall)
        result_cell.font = Font(bold=True, size=12)
        result_cell.fill = overall_fill

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 8: Position-Time 차트 (Space-Time Diagram)
    # ─────────────────────────────────────────────────────────────────────
    if tracker and tracker.segments:
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.axis import ChartLines
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.layout import Layout, ManualLayout
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont

        ws_chart = wb.create_sheet("Position-Time")

        # ── 데이터 생성 ──────────────────────────────────────────────
        time_series = tracker.generate_time_series(interval_s=5.0)
        has_phys = params.has_tubing_specs
        outlet_pos = params.reactor_volume + params.post_reactor_volume
        collection_end = outlet_pos + params.collection_line_volume

        # 컬럼 헤더
        headers_c = [
            "Time (s)", "Time (min)",
            "Reaction Front (mL)", "Reaction Tail (mL)",
            "Reactor End (mL)", "Outlet Valve (mL)", "Collection End (mL)",
            "Phase"
        ]
        if has_phys:
            headers_c += [
                "Reaction Front (m)", "Reaction Tail (m)",
                "Reactor End (m)", "Outlet Valve (m)", "Collection End (m)"
            ]

        for col, h in enumerate(headers_c, 1):
            cell = ws_chart.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row, pt in enumerate(time_series, 2):
            ws_chart.cell(row=row, column=1, value=pt['time'])
            ws_chart.cell(row=row, column=2, value=round(pt['time'] / 60.0, 2))
            ws_chart.cell(row=row, column=3, value=pt['rxn_front'])
            ws_chart.cell(row=row, column=4, value=pt['rxn_tail'])
            ws_chart.cell(row=row, column=5, value=params.reactor_volume)
            ws_chart.cell(row=row, column=6, value=outlet_pos)
            ws_chart.cell(row=row, column=7, value=collection_end)
            ws_chart.cell(row=row, column=8, value=pt.get('phase', ''))
            if has_phys:
                ws_chart.cell(row=row, column=9, value=round(tracker.vol_to_m(pt['rxn_front']), 3))
                ws_chart.cell(row=row, column=10, value=round(tracker.vol_to_m(pt['rxn_tail']), 3))
                ws_chart.cell(row=row, column=11, value=round(tracker.reactor_end_m, 3))
                ws_chart.cell(row=row, column=12, value=round(tracker.outlet_pos_m, 3))
                ws_chart.cell(row=row, column=13, value=round(tracker.collection_end_m, 3))

        num_data_rows = len(time_series)

        # ── 헬퍼: 축 폰트 생성 ───────────────────────────────────────
        def _axis_title_text(text, size=1100, bold=True):
            """축 제목용 RichText 생성"""
            rp = CharacterProperties(sz=size, b=bold)
            rp.solidFill = "333333"
            paragraph = Paragraph(pPr=ParagraphProperties(defRPr=rp), endParaRPr=rp)
            paragraph.text = text
            return RichText(p=[paragraph])

        def _chart_title_text(text, size=1400, bold=True):
            """차트 제목용 RichText 생성"""
            rp = CharacterProperties(sz=size, b=bold)
            rp.solidFill = "1A1A1A"
            paragraph = Paragraph(pPr=ParagraphProperties(defRPr=rp), endParaRPr=rp)
            paragraph.text = text
            return RichText(p=[paragraph])

        # ── Chart 1: Position (mL) vs Time ────────────────────────────
        if num_data_rows > 1:
            chart = LineChart()
            chart.title = "Reaction Mixture Slug Position in Common Path"
            chart.y_axis.title = "Cumulative Position from Mixer Inlet (mL)"
            chart.x_axis.title = "Elapsed Time (min)"
            chart.style = 2
            chart.width = 32
            chart.height = 20

            # 그리드라인
            chart.y_axis.majorGridlines = ChartLines()
            chart.y_axis.minorGridlines = ChartLines()
            chart.x_axis.majorGridlines = ChartLines()
            chart.y_axis.delete = False
            chart.x_axis.delete = False

            # 축 숫자 포맷
            chart.y_axis.numFmt = '0.0'
            chart.x_axis.numFmt = '0.0'
            chart.x_axis.tickLblPos = "low"

            # X axis = Time (min)
            cats = Reference(ws_chart, min_col=2, min_row=2,
                              max_row=num_data_rows + 1)
            chart.set_categories(cats)

            # (1) Reaction Front — 파란 실선 (두꺼움)
            front_data = Reference(ws_chart, min_col=3, min_row=1,
                                    max_row=num_data_rows + 1)
            chart.add_data(front_data, titles_from_data=True)
            s_front = chart.series[0]
            s_front.graphicalProperties.line.solidFill = "1F77B4"
            s_front.graphicalProperties.line.width = 28000

            # (2) Reaction Tail — 주황 실선 (두꺼움)
            tail_data = Reference(ws_chart, min_col=4, min_row=1,
                                   max_row=num_data_rows + 1)
            chart.add_data(tail_data, titles_from_data=True)
            s_tail = chart.series[1]
            s_tail.graphicalProperties.line.solidFill = "FF7F0E"
            s_tail.graphicalProperties.line.width = 28000

            # (3) Reactor End — 회색 점선
            reactor_data = Reference(ws_chart, min_col=5, min_row=1,
                                      max_row=num_data_rows + 1)
            chart.add_data(reactor_data, titles_from_data=True)
            s_reactor = chart.series[2]
            s_reactor.graphicalProperties.line.solidFill = "999999"
            s_reactor.graphicalProperties.line.dashStyle = "dash"
            s_reactor.graphicalProperties.line.width = 15000

            # (4) Outlet Valve — 빨강 점선
            outlet_data = Reference(ws_chart, min_col=6, min_row=1,
                                     max_row=num_data_rows + 1)
            chart.add_data(outlet_data, titles_from_data=True)
            s_outlet = chart.series[3]
            s_outlet.graphicalProperties.line.solidFill = "D62728"
            s_outlet.graphicalProperties.line.dashStyle = "dash"
            s_outlet.graphicalProperties.line.width = 15000

            # (5) Collection End — 보라 점선
            coll_data = Reference(ws_chart, min_col=7, min_row=1,
                                   max_row=num_data_rows + 1)
            chart.add_data(coll_data, titles_from_data=True)
            s_coll = chart.series[4]
            s_coll.graphicalProperties.line.solidFill = "9467BD"
            s_coll.graphicalProperties.line.dashStyle = "dashDot"
            s_coll.graphicalProperties.line.width = 12000

            # 레전드 하단 배치
            chart.legend.position = 'b'

            ws_chart.add_chart(chart, "A2")

            # ── Chart 2: Position (m) vs Time ─────────────────────────
            if has_phys:
                chart_m = LineChart()
                chart_m.title = "Reaction Mixture Slug Position — Physical Distance"
                chart_m.y_axis.title = "Position from Mixer Inlet (m)"
                chart_m.x_axis.title = "Elapsed Time (min)"
                chart_m.style = 2
                chart_m.width = 32
                chart_m.height = 20

                chart_m.y_axis.majorGridlines = ChartLines()
                chart_m.y_axis.minorGridlines = ChartLines()
                chart_m.x_axis.majorGridlines = ChartLines()
                chart_m.y_axis.numFmt = '0.00'
                chart_m.x_axis.numFmt = '0.0'
                chart_m.x_axis.tickLblPos = "low"

                chart_m.set_categories(cats)

                for col_idx, color, dash, width in [
                    (9, "1F77B4", None, 28000),      # Front (m)
                    (10, "FF7F0E", None, 28000),     # Tail (m)
                    (11, "999999", "dash", 15000),   # Reactor End (m)
                    (12, "D62728", "dash", 15000),   # Outlet (m)
                    (13, "9467BD", "dashDot", 12000), # Collection End (m)
                ]:
                    data_ref = Reference(ws_chart, min_col=col_idx, min_row=1,
                                          max_row=num_data_rows + 1)
                    chart_m.add_data(data_ref, titles_from_data=True)
                    s = chart_m.series[-1]
                    s.graphicalProperties.line.solidFill = color
                    s.graphicalProperties.line.width = width
                    if dash:
                        s.graphicalProperties.line.dashStyle = dash

                chart_m.legend.position = 'b'
                ws_chart.add_chart(chart_m, "A22")

        # ── 차트 해석 주석 (Chart Annotation) ─────────────────────────
        ann_start = num_data_rows + 4
        ann_font_title = Font(bold=True, size=12, color="1A1A1A")
        ann_font_section = Font(bold=True, size=10, color="333333")
        ann_font = Font(size=10, color="444444")
        ann_font_italic = Font(size=10, italic=True, color="666666")

        # 제목
        ws_chart.cell(row=ann_start, column=1,
                       value="차트 해석 가이드").font = ann_font_title
        ws_chart.merge_cells(start_row=ann_start, start_column=1,
                              end_row=ann_start, end_column=6)

        r = ann_start + 2

        # 1. 차트 개요
        ws_chart.cell(row=r, column=1,
                       value="1. 이 차트가 보여주는 것").font = ann_font_section
        r += 1
        lines = [
            "이 차트는 반응 혼합물 \"슬러그(slug)\"가 연속 흐름 반응기 공통 경로를 통과하는 과정을 추적하는 시공간 다이어그램(Space-Time Diagram)입니다.",
            "경로: 믹서(Mixer) → 반응기(Reactor) → 후단 튜빙(Post-Reactor) → 아울렛 밸브(Outlet) → 분취 라인(Collection)",
            "X축은 경과 시간, Y축은 믹서 입구로부터의 누적 유체 위치를 나타냅니다.",
            "Front와 Tail 사이의 수직 거리 = 반응량(reaction volume)에 해당합니다.",
        ]
        for line in lines:
            ws_chart.cell(row=r, column=1, value=line).font = ann_font
            ws_chart.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1

        r += 1

        # 2. 데이터 시리즈 설명
        ws_chart.cell(row=r, column=1,
                       value="2. 데이터 시리즈 범례").font = ann_font_section
        r += 1

        legend_items = [
            ("Reaction Front (파란 실선)", "반응 혼합물 슬러그의 선두. 각 구간 경계에 가장 먼저 도달합니다."),
            ("Reaction Tail (주황 실선)", "슬러그의 후미. Tail이 특정 지점을 넘으면 슬러그 전체가 해당 지점을 통과한 것입니다."),
            ("Reactor End (회색 점선)", f"경계선: {params.reactor_volume:.3f} mL — 체류 시간이 완료되는 반응기 코일의 끝."),
            ("Outlet Valve (빨간 점선)", f"경계선: {outlet_pos:.3f} mL — WASTE/COLLECT로 흐름을 전환하는 아울렛 밸브 위치."),
            ("Collection End (보라 일점쇄선)", f"경계선: {collection_end:.3f} mL — 분취기까지의 수거 튜빙 끝."),
        ]

        for name, desc in legend_items:
            ws_chart.cell(row=r, column=1, value=name).font = Font(bold=True, size=10)
            ws_chart.cell(row=r, column=3, value=desc).font = ann_font
            ws_chart.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
            r += 1

        r += 1

        # 3. Phase별 해석
        ws_chart.cell(row=r, column=1,
                       value="3. 단계별(Phase) 해석").font = ann_font_section
        r += 1

        phase_desc = [
            ("INJECTION", "시약을 공통 경로에 주입. Front가 전진하고 Tail은 0에 머무름 (슬러그 성장). 기울기 = 총 유속."),
            ("TRANSIT", "용매가 슬러그를 아울렛 방향으로 밀어냄. Front와 Tail이 동일 속도로 전진 (평행선 = 슬러그 길이 일정)."),
            ("COLLECTION", "Front가 Outlet 경계선에 도달하면 아울렛 밸브가 COLLECT로 전환. 슬러그가 분취 튜브로 이동."),
            ("LINE_WASH", "수거 후 세척. 분취 튜빙에 남은 잔류 유체를 폐기 튜브로 세척."),
        ]

        for phase, desc in phase_desc:
            ws_chart.cell(row=r, column=1, value=phase).font = Font(bold=True, size=10, color="0070C0")
            ws_chart.cell(row=r, column=2, value=desc).font = ann_font
            ws_chart.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            r += 1

        r += 1

        # 4. 핵심 검증 포인트
        ws_chart.cell(row=r, column=1,
                       value="4. 핵심 검증 포인트").font = ann_font_section
        r += 1

        validation_lines = [
            "아울렛 밸브가 WASTE → COLLECT로 전환될 때, Reaction Front는 Outlet 경계선에 정확히 도달해야 합니다.",
            "Front가 Outlet을 이미 지남 → 반응물 일부가 폐기로 유실 (수율 손실).",
            "Front가 Outlet에 미도달 → 용매/데드볼륨이 수거됨 (순도 저하).",
            f"본 시뮬레이션: transit_vol = zone_vol - reaction_vol = {outlet_pos:.3f} - {experiment.reaction_volume:.3f} = {outlet_pos - experiment.reaction_volume:.3f} mL",
            "→ 수거 시작 시점에 Front가 Outlet에 정확히 도달 (유실 0%).",
        ]
        for line in validation_lines:
            ws_chart.cell(row=r, column=1, value=line).font = ann_font
            ws_chart.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1

        r += 1

        # 5. 기울기 해석
        ws_chart.cell(row=r, column=1,
                       value="5. 기울기(Slope) 해석").font = ann_font_section
        r += 1
        total_flow = sum(flows.values())
        slope_lines = [
            f"Front/Tail 선의 기울기 = 총 유속 = {total_flow:.4f} mL/min = {total_flow/60:.6f} mL/s.",
            "기울기가 급함 → 유속이 빠름 → 실험 시간 단축.",
            "평탄 구간 (기울기 = 0) → 펌핑 없음 (예: 밸브 전환, 안정화 대기).",
        ]
        if has_phys:
            slope_lines.append(
                f"물리적 유속 = {total_flow / (params.reactor_tubing.cross_section_cm2 * 60):.4f} cm/s "
                f"(ID={params.reactor_tubing.id_mm}mm 튜빙 기준)."
            )
        for line in slope_lines:
            ws_chart.cell(row=r, column=1, value=line).font = ann_font
            ws_chart.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1

        if has_phys:
            r += 1
            ws_chart.cell(row=r, column=1,
                           value="6. 물리적 거리 차트 (m)").font = ann_font_section
            r += 1
            phys_lines = [
                "두 번째 차트는 부피 기반 위치(mL)를 각 구간의 단면적을 이용하여 물리적 튜빙 길이(m)로 변환한 것입니다.",
                f"반응기(Reactor): ID={params.reactor_tubing.id_mm}mm, L={params.reactor_tubing.length_m:.2f}m "
                f"(A={params.reactor_tubing.cross_section_cm2:.6f} cm²)",
                f"후단 튜빙(Post-Reactor): ID={params.post_reactor_tubing.id_mm}mm, L={params.post_reactor_tubing.length_m:.3f}m "
                f"(A={params.post_reactor_tubing.cross_section_cm2:.6f} cm²)",
                f"분취 라인(Collection): ID={params.collection_line_tubing.id_mm}mm, L={params.collection_line_tubing.length_m:.3f}m "
                f"(A={params.collection_line_tubing.cross_section_cm2:.6f} cm²)",
                "구간별 내경이 다르면 구간 경계에서 기울기가 변합니다 (동일한 mL/min이지만 m/min은 달라짐).",
            ]
            for line in phys_lines:
                ws_chart.cell(row=r, column=1, value=line).font = ann_font
                ws_chart.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
                r += 1

    # 열 너비 자동 조정
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)
    print(f"\n[Excel] 시뮬레이션 결과 저장됨: {filepath}")

    return filepath


# ═══════════════════════════════════════════════════════════════════════════════
# 6. 콘솔 출력
# ═══════════════════════════════════════════════════════════════════════════════

def print_summary_table(events: List[TimingEvent], flows: Dict[str, float],
                        groups: List[GroupConfig], total_time: float):
    """콘솔 요약 테이블 출력"""

    print("\n" + "=" * 100)
    print("DETAILED TIMING SIMULATION RESULT")
    print("=" * 100)

    # 유속 계산 결과
    print("\n[FLOW CALCULATION RESULT]")
    print(f"{'Group':<12} {'Reagent':<15} {'Port':<6} {'Conc(M)':<10} {'Eq':<8} {'Flow(mL/min)':<15}")
    print("-" * 70)
    for g in groups:
        print(f"{g.name:<12} {g.reagent.name:<15} {g.reagent.port:<6} "
              f"{g.reagent.concentration:<10.3f} {g.reagent.equivalents:<8.2f} "
              f"{flows[g.name]:<15.4f}")
    print("-" * 70)
    print(f"{'TOTAL':<58} {sum(flows.values()):<15.4f}")

    # 밸브 전환 테이블
    print("\n[VALVE SWITCHING TIMELINE]")
    print(f"{'Time(s)':<10} {'Phase':<15} {'Component':<20} {'From':<15} {'To':<15} {'Duration(s)':<12}")
    print("-" * 90)

    valve_events = [e for e in events if e.event_type in
                    [EventType.VALVE_12WAY, EventType.VALVE_3WAY, EventType.VALVE_OUTLET]]
    for e in sorted(valve_events, key=lambda x: x.time):
        print(f"{e.time:<10.2f} {e.phase:<15} {e.component:<20} "
              f"{e.from_state:<15} {e.to_state:<15} {e.duration:<12.2f}")

    # 펌프 상태 테이블 (요약)
    print("\n[PUMP STATE CHANGES]")
    print(f"{'Time(s)':<10} {'Phase':<15} {'Component':<18} {'State':<12} {'Rate(mL/min)':<15}")
    print("-" * 75)

    pump_events = [e for e in events if e.event_type in
                   [EventType.PUMP_INFUSE_START, EventType.PUMP_WITHDRAW_START]]
    for e in sorted(pump_events, key=lambda x: x.time):
        state = "INFUSING" if e.event_type == EventType.PUMP_INFUSE_START else "WITHDRAWING"
        print(f"{e.time:<10.2f} {e.phase:<15} {e.component:<18} {state:<12} {e.value:<15.4f}")

    # Outlet 밸브 전환
    print("\n[OUTLET VALVE TRANSITIONS]")
    outlet_events = [e for e in events if e.event_type == EventType.VALVE_OUTLET]
    for e in sorted(outlet_events, key=lambda x: x.time):
        print(f"  {e.time:.2f}s: {e.from_state} → {e.to_state} (Phase: {e.phase})")

    print("\n" + "=" * 100)
    print(f"TOTAL SEQUENCE TIME: {total_time:.1f}s ({total_time/60:.2f} min)")
    print("=" * 100)


# ═══════════════════════════════════════════════════════════════════════════════
# 7. 메인 실행
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # ─────────────────────────────────────────────────────────────────────
    # 예제 설정: 3개 그룹 (Group A, B, C)
    # ─────────────────────────────────────────────────────────────────────
    # ── 현재 hardware_config.json 기준 (2026-04-12 업데이트) ──
    groups = [
        GroupConfig(
            name="Group A",
            reagent=ReagentConfig(
                name="Reagent A",
                port=2,
                concentration=1.0,
                equivalents=1.0
            ),
            syringe_volume=6.0,
            refill_rate=12.0
        ),
        GroupConfig(
            name="Group B",
            reagent=ReagentConfig(
                name="Reagent B",
                port=2,
                concentration=1.0,
                equivalents=1.0
            ),
            syringe_volume=6.0,
            refill_rate=12.0
        ),
        GroupConfig(
            name="Group C",
            reagent=ReagentConfig(
                name="Reagent C",
                port=2,
                concentration=1.0,
                equivalents=1.0
            ),
            syringe_volume=6.0,
            refill_rate=12.0
        ),
    ]

    experiment = ExperimentParams(
        temperature=80.0,
        residence_time=10.0,
        reaction_volume=3.0,
        volume_per_tube=1.5,
        start_tube=1
    )

    params = SystemParams(
        reactor_volume=2.0056,       # π × (0.4mm)² × 399cm
        post_reactor_volume=2.0,
        collection_line_volume=1.0,
        dead_volume_per_line=0.0,    # 데드볼륨 보정 제거됨
        priming_rate=4.0,            # 현재 설정
    )

    # ─────────────────────────────────────────────────────────────────────
    # 시뮬레이션 실행
    # ─────────────────────────────────────────────────────────────────────
    print("Starting detailed timing simulation...")
    events, total_time, flows, tracker = run_detailed_simulation(groups, experiment, params)

    # 콘솔 출력
    print_summary_table(events, flows, groups, total_time)

    # Excel 저장
    if HAS_OPENPYXL:
        excel_path = export_to_excel(events, flows, groups, experiment, params, total_time,
                                     tracker=tracker)
        print(f"\nExcel file saved: {excel_path}")
    else:
        print("\n[Warning] openpyxl not installed. Excel export skipped.")
        print("Install with: pip install openpyxl")

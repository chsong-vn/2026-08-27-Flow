"""
Calculator Tab - 농도 계산기 (B형: 2단 헤더 그룹 테이블)
CAS 번호 → PubChem 자동 조회 (시약명 + MW)
MW + (mL / g / M) 중 2개 입력 → 나머지 자동 계산
행 수: 그룹 수 × 10 (자동 할당), 그룹별 섹션 구분
Import: 행 위치 기반 자동 그룹/포트 매핑
"""
import urllib.request
import urllib.parse
import json
import math
import os
import re

import pyqtgraph as pg
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QHeaderView, QLineEdit, QApplication,
    QDialog, QGridLayout, QMessageBox, QFrame, QTabWidget,
    QDoubleSpinBox, QComboBox, QSplitter, QScrollArea
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QEvent, QTimer, QFileSystemWatcher
from PyQt5.QtGui import QColor

from ui.colors import DarkPalette as Dark, LightPalette as Light, DarkExtras, LightExtras, T
from ui.theme import get_secondary_button_style, get_primary_button_style

ROWS_PER_GROUP = 10   # 그룹당 행 수 (포트 2~11 = 10개)

# 헤더 그룹 색상
_HDR_INPUT  = "#1e3a5f"
_HDR_RESULT = "#1a3a20"
_ROW_DONE   = "#0d2010"

# 그룹 섹션 색상 (bg, fg) — 전 그룹 통일
_GRP_COLORS = [("#1e252e", "#e0e0e0")] * 6

# 셀 위젯 공통 스타일
_CELL_FS        = T.FS_MD
_CELL_EDIT      = f"font-family: {T.FONT}; font-size: {_CELL_FS}; background: transparent; border: none; padding: 0 4px;"
_CELL_NUM       = f"font-family: {T.FONT_MONO}; font-size: {_CELL_FS}; background: transparent; border: none; padding: 0 4px;"
_CELL_RES       = f"font-family: {T.FONT_MONO}; font-size: {_CELL_FS}; padding: 0 6px;"
# 희석 계산기 입력 칸 — 구분을 위해 하단 보더 + 미세 배경
_CELL_DIL_INPUT = (
    f"font-family: {T.FONT_MONO}; font-size: {_CELL_FS};"
    " background: #1a2540; border: none; border-bottom: 1px solid #3a5a9a;"
    " padding: 0 6px;"
)


class _Row:
    __slots__ = ('name', 'mw', 'density', 'ml', 'g', 'conc', 'eff_conc', 'cas', 'smiles')

    def __init__(self):
        self.name     = ''
        self.cas      = ''     # CAS 번호 (WebGrid 통합)
        self.smiles   = ''     # 구조 (PubChem/연구노트 연동)
        self.mw       = 0.0
        self.density  = None   # g/mL — 액상 시약일 때만 입력 (선택)
        self.ml       = None
        self.g        = None
        self.conc     = None
        self.eff_conc = None


def _parse(text):
    try:
        v = float(text.strip().replace(',', '.'))
        return v if v > 0 else None
    except Exception:
        return None


def _d2t(dr):
    """data index → table row (separator rows 포함)"""
    return (dr // ROWS_PER_GROUP) * (ROWS_PER_GROUP + 1) + 1 + (dr % ROWS_PER_GROUP)


# ── 마우스 투과 QLineEdit (드래그 선택 + 더블클릭 편집) ─────────────

class _DragLineEdit(QLineEdit):
    """마우스 투과 QLineEdit — ReadOnly 상태에서 클릭/드래그를 부모 테이블로 전달,
    더블클릭 시 편집 모드 진입 (엑셀 UX)"""

    def _find_table(self):
        p = self.parent()
        while p and not isinstance(p, QTableWidget):
            p = p.parent()
        return p

    def _cell_at(self, e):
        """마우스 위치에 해당하는 테이블 (row, col) 반환"""
        tbl = self._find_table()
        if not tbl:
            return None, None, None
        vp = tbl.viewport()
        pos = vp.mapFromGlobal(self.mapToGlobal(e.pos()))
        idx = tbl.indexAt(pos)
        if idx.isValid():
            return tbl, idx.row(), idx.column()
        return tbl, None, None

    def mousePressEvent(self, e):
        if not self.isReadOnly():
            super().mousePressEvent(e)
            return
        tbl, row, col = self._cell_at(e)
        if tbl and row is not None:
            tbl.setCurrentCell(row, col)
            tbl.setProperty("_drag_origin", f"{row},{col}")

    def mouseMoveEvent(self, e):
        if not self.isReadOnly():
            super().mouseMoveEvent(e)
            return
        tbl, row, col = self._cell_at(e)
        if tbl and row is not None:
            origin = tbl.property("_drag_origin")
            if origin:
                from PyQt5.QtWidgets import QTableWidgetSelectionRange
                r0, c0 = [int(x) for x in origin.split(",")]
                top, bottom = min(r0, row), max(r0, row)
                left, right = min(c0, col), max(c0, col)
                tbl.clearSelection()
                tbl.setRangeSelected(QTableWidgetSelectionRange(top, left, bottom, right), True)

    def mouseReleaseEvent(self, e):
        if not self.isReadOnly():
            super().mouseReleaseEvent(e)
            return
        tbl = self._find_table()
        if tbl:
            tbl.setProperty("_drag_origin", None)

    def mouseDoubleClickEvent(self, e):
        self.setReadOnly(False)
        self.setFocusPolicy(Qt.StrongFocus)
        self.setFocus()
        self.selectAll()

    def focusOutEvent(self, e):
        self.setReadOnly(True)
        self.setFocusPolicy(Qt.NoFocus)
        self.deselect()
        super().focusOutEvent(e)


# ── CAS 입력 전용 ─────────────────────────────────────────────────

class _CasLineEdit(_DragLineEdit):
    """CAS 입력 전용 — _DragLineEdit 상속 (마우스 투과 + 더블클릭 편집)"""
    paste_multi = pyqtSignal(int, list)   # (시작 data 행, [cas, ...])

    def __init__(self, row):
        super().__init__()
        self._row = row


# ── PubChem 비동기 조회 Worker ──────────────────────────────────

class _CasWorker(QThread):
    # @codesyncer-decision: 밀도 None 표현을 위해 float 대신 object 시그널 인자 사용.
    #   PubChem 미수록/파싱 실패 시 None을 그대로 전달 → 핸들러에서 빈칸 유지.
    result_ready = pyqtSignal(int, str, float, object, str)  # row, name, mw, density|None, smiles
    lookup_error = pyqtSignal(int, str)                  # row, message

    _UA = {"User-Agent": "FlowChemApp/1.0"}
    # @codesyncer-inference: PubChem Density는 free-text로 여러 형식 존재:
    #   1) "0.7893 g/cu cm at 20 °C" → _DENS_RE로 매칭
    #   2) "Relative density (water = 1): 0.79" → _DENS_REL_RE로 매칭 (fallback)
    #   1)을 우선 사용, 없으면 2)로 대체 (상대 밀도 ≈ g/mL, 물 기준)
    _DENS_RE = re.compile(r"(\d+(?:\.\d+)?)\s*g\s*/\s*(?:cm\s*[³3]|m\s*[lL]|cc|cu\s*cm)", re.IGNORECASE)
    _DENS_REL_RE = re.compile(r"[Rr]elative\s+density\s*\([^)]*\)\s*:\s*(\d+(?:\.\d+)?)")

    def __init__(self, row, cas):
        super().__init__()
        self.row = row
        self.cas = cas

    def _fetch_json(self, url):
        req = urllib.request.Request(url, headers=self._UA)
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read())

    def _lookup_density(self, cid):
        """PUG-View Density 섹션에서 첫 번째 g/mL 값 추출 — 실패 시 None."""
        try:
            url = (
                "https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/"
                f"{cid}/JSON?heading=Density"
            )
            data = self._fetch_json(url)
        except Exception as exc:
            print(f"[Density] CID {cid}: not available ({exc.__class__.__name__})")
            return None

        # 재귀로 모든 String 값 평탄화하여 정규식 매칭
        def _walk(node):
            if isinstance(node, dict):
                for v in node.values():
                    yield from _walk(v)
            elif isinstance(node, list):
                for v in node:
                    yield from _walk(v)
            elif isinstance(node, str):
                yield node

        # 1차: 절대 밀도 (g/cm³, g/mL, g/cu cm 등)
        for s in _walk(data):
            m = self._DENS_RE.search(s)
            if m:
                try:
                    return float(m.group(1))
                except ValueError:
                    continue

        # 2차 fallback: 상대 밀도 (Relative density (water=1): 0.79)
        for s in _walk(data):
            m = self._DENS_REL_RE.search(s)
            if m:
                try:
                    val = float(m.group(1))
                    if 0.1 < val < 25.0:  # 합리적 범위
                        return val
                except ValueError:
                    continue

        return None

    def run(self):
        try:
            encoded = urllib.parse.quote(self.cas)
            # 1단계: CAS/이름 → CID
            cids = self._fetch_json(
                "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/"
                f"{encoded}/cids/JSON"
            )
            cid = cids["IdentifierList"]["CID"][0]

            # 2단계: CID → IUPACName, MW, SMILES
            data = self._fetch_json(
                "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/"
                f"{cid}/property/IUPACName,MolecularWeight,CanonicalSMILES/JSON"
            )
            props = data["PropertyTable"]["Properties"][0]
            name  = props.get("IUPACName", "") or ""
            mw    = float(props.get("MolecularWeight", 0))
            # @codesyncer-decision: SMILES → 연구노트(schemes.flow·구조) + reagent_lib.
            #   신 API가 CanonicalSMILES 를 SMILES/ConnectivitySMILES 로 개명한 경우 폴백.
            smiles = (props.get("CanonicalSMILES") or props.get("SMILES")
                      or props.get("ConnectivitySMILES") or "")

            # 3단계: 밀도 (선택) — 실패해도 결과 자체는 반환
            density = self._lookup_density(cid)

            self.result_ready.emit(self.row, name, mw, density, smiles)

        except Exception as e:
            self.lookup_error.emit(self.row, str(e))


# ── Fill Handle (우하단 드래그 → 아래 행 채우기) ──────────────────

class _FillHandle(QWidget):
    """셀 우하단의 파란 사각형; 드래그하면 아래 행에 값 복사"""
    ROW_H = 40

    def __init__(self, get_val_fn, fill_fn, parent=None):
        super().__init__(parent)
        self._get_val = get_val_fn
        self._fill    = fill_fn
        self.setFixedSize(8, 8)
        self.setCursor(Qt.SizeFDiagCursor)
        self.setToolTip("드래그: 아래로 채우기")
        self._drag_y = None

    def paintEvent(self, e):
        pass  # 투명 — 커서 변화로만 인식

    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            self._drag_y = e.globalY()
            self.grabMouse()

    def mouseReleaseEvent(self, e):
        if self._drag_y is not None and e.button() == Qt.LeftButton:
            self.releaseMouse()
            n = max(0, round((e.globalY() - self._drag_y) / self.ROW_H))
            if n > 0:
                self._fill(self._get_val(), n)
            self._drag_y = None


class _CellEdit(QWidget):
    """_DragLineEdit + 우하단 fill handle 오버레이
    컨테이너 자체도 마우스를 viewport로 투과 — 패딩 영역 클릭/드래그도 테이블 선택 동작"""
    def __init__(self, placeholder="-", style=None,
                 align=Qt.AlignRight | Qt.AlignVCenter):
        super().__init__()
        _s = style if style is not None else _CELL_NUM
        self.edit = _DragLineEdit(self)
        self.edit.setPlaceholderText(placeholder)
        self.edit.setFrame(False)
        self.edit.setAlignment(align)
        self.edit.setStyleSheet(_s)
        self.edit.setReadOnly(True)
        self.edit.setFocusPolicy(Qt.NoFocus)
        self._handle = None

    def _find_table(self):
        p = self.parent()
        while p and not isinstance(p, QTableWidget):
            p = p.parent()
        return p

    def _cell_at(self, e):
        tbl = self._find_table()
        if not tbl:
            return None, None, None
        vp = tbl.viewport()
        pos = vp.mapFromGlobal(self.mapToGlobal(e.pos()))
        idx = tbl.indexAt(pos)
        if idx.isValid():
            return tbl, idx.row(), idx.column()
        return tbl, None, None

    def mousePressEvent(self, e):
        tbl, row, col = self._cell_at(e)
        if tbl and row is not None:
            tbl.setCurrentCell(row, col)
            tbl.setProperty("_drag_origin", f"{row},{col}")

    def mouseMoveEvent(self, e):
        tbl, row, col = self._cell_at(e)
        if tbl and row is not None:
            origin = tbl.property("_drag_origin")
            if origin:
                from PyQt5.QtWidgets import QTableWidgetSelectionRange
                r0, c0 = [int(x) for x in origin.split(",")]
                top, bottom = min(r0, row), max(r0, row)
                left, right = min(c0, col), max(c0, col)
                tbl.clearSelection()
                tbl.setRangeSelected(QTableWidgetSelectionRange(top, left, bottom, right), True)

    def mouseReleaseEvent(self, e):
        tbl = self._find_table()
        if tbl:
            tbl.setProperty("_drag_origin", None)

    def mouseDoubleClickEvent(self, e):
        self.edit.mouseDoubleClickEvent(e)

    def attach_fill(self, fill_fn):
        self._handle = _FillHandle(self.edit.text, fill_fn, self)
        self._reposition()

    def resizeEvent(self, e):
        super().resizeEvent(e)
        self._reposition()

    def _reposition(self):
        w, h = self.width(), self.height()
        self.edit.setGeometry(0, 0, w, h)
        if self._handle:
            self._handle.move(w - 8, h - 8)
            self._handle.raise_()


# ── RTD 계산 함수 ────────────────────────────────────────────────

# 용매 물성표 (μ: Pa·s, ρ: kg/m³, 25°C 기준)
_SOLVENTS = {
    "Hexane":        (0.00030, 655),
    "Diethyl ether": (0.00022, 713),
    "DCM":           (0.00043, 1325),
    "Chloroform":    (0.00057, 1489),
    "Toluene":       (0.00059, 867),
    "Acetonitrile":  (0.00037, 786),
    "Methanol":      (0.00060, 792),
    "Ethanol":       (0.00110, 789),
    "Isopropanol":   (0.00240, 786),
    "Water":         (0.00089, 997),
    "THF":           (0.00046, 887),
    "DMF":           (0.00080, 944),
    "DMSO":          (0.00200, 1100),
    "1,4-Dioxane":   (0.00130, 1033),
}


def _rtd_tube_volume(id_mm, length_m):
    """V (mL) = π r² L"""
    r_cm = (id_mm / 10.0) / 2.0
    return math.pi * r_cm ** 2 * (length_m * 100.0)


def _rtd_diffusion(visc_Pa_s, T_K=298.15, r_mol_m=0.5e-9):
    """Stokes–Einstein 확산계수 D (m²/s)"""
    k_B = 1.380649e-23
    return k_B * T_K / (6.0 * math.pi * visc_Pa_s * r_mol_m)


def _rtd_calc(id_mm, length_m, flow_ml_min, visc_Pa_s, rho_kg_m3, T_K=298.15):
    """
    반환: dict {V_ml, tau_min, u_m_s, Re, D_m2_s, D_ax_m2_s, Pe, Gz}
    Taylor–Aris 분산 모델 (층류 원형 튜브)
    """
    Q_m3s   = flow_ml_min * 1e-6 / 60.0
    r_tube  = (id_mm / 1000.0) / 2.0
    A       = math.pi * r_tube ** 2
    u       = Q_m3s / A                                    # 평균 유속 [m/s]
    V_ml    = _rtd_tube_volume(id_mm, length_m)
    tau_min = V_ml / flow_ml_min                           # 체류시간 [min]

    Re  = rho_kg_m3 * u * (id_mm / 1000.0) / visc_Pa_s    # Reynolds 수
    D   = _rtd_diffusion(visc_Pa_s, T_K)                   # 분자 확산계수
    D_ax = D + (u ** 2 * r_tube ** 2) / (48.0 * D)        # Taylor–Aris

    Pe  = u * length_m / D_ax                              # Péclet 수
    # Graetz 수: 1 이상이면 Taylor–Aris 완전 발달 조건 충족
    Gz  = length_m * D / (u * r_tube ** 2) if u > 0 else 0

    return dict(V_ml=V_ml, tau_min=tau_min, u_m_s=u,
                Re=Re, D_m2_s=D, D_ax_m2_s=D_ax, Pe=Pe, Gz=Gz)


def _rtd_curve(tau_min, Pe, n_points=400):
    """E(t) 곡선 계산. 반환: (t_list_min, E_list_1_per_min)"""
    t_arr, E_arr = [], []
    for i in range(1, n_points + 1):
        t = i * tau_min * 3.0 / n_points
        theta = t / tau_min
        try:
            coeff = math.sqrt(Pe / (4.0 * math.pi * theta ** 3))
            expo  = -Pe * (1.0 - theta) ** 2 / (4.0 * theta)
            E     = (coeff * math.exp(expo)) / tau_min
        except (ValueError, OverflowError):
            E = 0.0
        t_arr.append(t)
        E_arr.append(E)
    return t_arr, E_arr


# ── 그룹 정보 계산 헬퍼 ─────────────────────────────────────────

def _build_group_info(inlet_pumps):
    """Returns list of (group_letter, first_pump_name)"""
    if not inlet_pumps:
        return []
    ppg = max(1, len(inlet_pumps) // 3) if len(inlet_pumps) >= 3 else 1
    groups = []
    gi = 0
    for idx in range(0, len(inlet_pumps), ppg):
        groups.append((chr(ord('A') + gi), inlet_pumps[idx]))
        gi += 1
    return groups


# ── 메인 탭 위젯 ───────────────────────────────────────────────

class CalculatorTab(QWidget):
    def __init__(self, app):
        super().__init__()
        self.app       = app
        self._building = False
        self._workers  = set()

        # ── 그룹 정보 계산 ──
        inlet_pumps = list(getattr(app.cfg, 'INLET_PUMPS', []) or [])
        self._group_info = _build_group_info(inlet_pumps)   # [(letter, pump), ...]
        self._num_groups = max(1, len(self._group_info))
        self._num_rows   = self._num_groups * ROWS_PER_GROUP

        self._rows = [_Row() for _ in range(self._num_rows)]
        self._undo_stack = []  # Ctrl+Z 되돌리기용 스냅샷 스택

        self._num_items   = []
        self._cas_edits   = []
        self._name_edits  = []
        self._mw_edits    = []
        self._d_edits     = []   # 밀도 g/mL (선택)
        self._ml_edits    = []
        self._g_edits     = []
        self._m_edits     = []
        self._result_lbls = []
        self._grp_sep_labels = []   # Group 섹션 구분 라벨 (테마 갱신용)
        self._tbl = None            # 구 농도 QTableWidget (미사용 · eventFilter 안전용)
        self._grid = None           # WebGrid(Tabulator) 농도 테이블
        self._calc_excel_path = None      # 'Excel에서 편집' 임시 파일 (시퀀스 탭과 동일 UX)
        self._calc_excel_watcher = None   # 저장 감지 → 자동 반영
        self._header_items = []     # 테이블 헤더 아이템 (테마 갱신용)

        # ── 테마 갱신용 위젯 참조 ──
        self._conc_sub_lbl = None
        self._conc_btn_imp = None
        self._dil_sub_lbl = None
        self._dil_title_lbl = None
        self._dil_ctrl_labels = []
        self._rtd_inp_frame = None
        self._rtd_sec_lbl = None
        self._rtd_btn_calc = None
        self._rtd_sep = None
        self._rtd_splitter = None
        self._rtd_hint_lbl = None
        self._rtd_rlabels = []
        self._rtd_card_frames = []
        self._rtd_card_titles = []
        self._dil_num_items = []   # 희석 테이블 행 번호 아이템
        self._conc_btns_secondary = []
        self._dil_btns_secondary = []

        self._setup_ui()

    @property
    def _P(self):
        """Current theme palette."""
        return Dark if getattr(self.app, 'is_dark_mode', True) else Light

    # ── UI 구성 ────────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(T.SP_SM, T.SP_SM, T.SP_SM, T.SP_SM)
        root.setSpacing(0)

        tabs = QTabWidget()
        tabs.setObjectName("CalcSubTabs")

        conc_page = QWidget()
        self._build_conc_tab(conc_page)
        tabs.addTab(conc_page, "농도 계산기")
        tabs.setTabToolTip(0,
            "<b>농도 계산기</b><br>"
            "사용할 시약의 CAS 번호를 입력하고 '일괄 조회'를 누르면<br>"
            "시약명, 분자량(MW), 밀도가 자동으로 채워집니다.<br>"
            "부피(mL), 질량(g), 농도(M) 중 2개만 입력하면 나머지가 계산됩니다.<br>"
            "<i>더블클릭으로 셀 편집 | 드래그로 범위 선택 | Ctrl+C 복사 | Ctrl+V 붙여넣기</i>"
        )

        dil_page = QWidget()
        self._build_dil_tab(dil_page)
        tabs.addTab(dil_page, "희석 계산기")
        tabs.setTabToolTip(1,
            "<b>희석 계산기</b><br>"
            "'시약 불러오기'로 농도 계산기의 원액 정보를 가져옵니다.<br>"
            "목표 농도와 총 부피를 입력하면 필요한 원액량과 용매량이 계산됩니다.<br>"
            "예: 1M 원액 → 0.1M 100mL 만들기 → 원액 10mL + 용매 90mL"
        )

        rtd_page = QWidget()
        self._build_rtd_tab(rtd_page)
        tabs.addTab(rtd_page, "RTD 계산기")
        tabs.setTabToolTip(2,
            "<b>RTD 계산기 (체류시간 분포)</b><br>"
            "반응기 규격과 유속을 입력하면 시약이 반응기 안에<br>"
            "얼마나 머무르는지(체류시간) 예측합니다.<br>"
            "유동 상태(층류/난류)와 혼합 효율도 확인할 수 있습니다."
        )

        root.addWidget(tabs)

    def _build_conc_tab(self, page):
        root = QVBoxLayout(page)
        root.setContentsMargins(T.SP_MD, T.SP_MD, T.SP_MD, T.SP_MD)
        root.setSpacing(10)

        # @codesyncer-decision: 지시문 부제 제거 — 제목·사용법은 WebGrid 카드 헤더가
        #   담당(제목 "시약·스톡 계산" + ⓘ 사용법 툴팁). 전문툴은 헤더에 지시문 상주 X.
        #   촘촘한 QTableWidget → WebGrid(Tabulator) 교체(엑셀 복붙/자동계산). 구 코드 아래 보존.
        self._build_conc_grid(root)
        return

        # ── 테이블 (separator 포함) ── (미사용 · 보존)
        total_tbl_rows = self._num_groups * (ROWS_PER_GROUP + 1)
        tbl = QTableWidget(total_tbl_rows, 9)

        col_defs = [
            ("#",            None,        QHeaderView.Fixed,   46),
            ("CAS",          None,        QHeaderView.Fixed,   170),
            ("시약명",       None,        QHeaderView.Stretch, -1),
            ("MW\n(g/mol)",  None,        QHeaderView.Fixed,   86),
            ("밀도\n(g/mL)", None,        QHeaderView.Fixed,   78),
            ("부피\n(mL)",   _HDR_INPUT,  QHeaderView.Fixed,   72),
            ("질량\n(g)",    _HDR_INPUT,  QHeaderView.Fixed,   72),
            ("농도\n(M)",    _HDR_INPUT,  QHeaderView.Fixed,   72),
            ("계산  결과",   _HDR_RESULT, QHeaderView.Stretch, -1),
        ]

        for col, (text, bg, resize, width) in enumerate(col_defs):
            item = QTableWidgetItem(text)
            item.setTextAlignment(Qt.AlignCenter)
            if bg:
                item.setBackground(QColor(bg))
                # @codesyncer-risk: 헤더 배경이 다크 컬러 고정 → 라이트 모드에서 기본
                # 텍스트(검정)가 묻혀 가독성 0. 흰색으로 강제하여 양쪽 테마 모두 가독.
                item.setForeground(QColor("#FFFFFF"))
            tbl.setHorizontalHeaderItem(col, item)
            tbl.horizontalHeader().setSectionResizeMode(col, resize)
            if width > 0:
                tbl.setColumnWidth(col, width)

        tbl.verticalHeader().setVisible(False)
        tbl.verticalHeader().setDefaultSectionSize(40)
        tbl.setAlternatingRowColors(False)
        tbl.setSelectionMode(QTableWidget.ContiguousSelection)
        tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.installEventFilter(self)
        self._tbl = tbl

        # ── 그룹별 섹션 + 데이터 행 생성 ──
        for gi, (letter, pump) in enumerate(self._group_info):
            grp_bg, grp_fg = _GRP_COLORS[gi % len(_GRP_COLORS)]
            sep_tr = gi * (ROWS_PER_GROUP + 1)

            # ── 섹션 구분 행 ──
            tbl.setRowHeight(sep_tr, 28)
            tbl.setSpan(sep_tr, 0, 1, 9)
            sep_lbl = QLabel(f"  Group {letter}   (Port 2 ~ 11)")
            sep_lbl.setAttribute(Qt.WA_TransparentForMouseEvents)
            sep_lbl.setStyleSheet(
                f"font-weight: bold; font-size: 11px;"
                f"color: {grp_fg};"
                f"background: {grp_bg};"
                f"padding: 0 6px;"
            )
            tbl.setCellWidget(sep_tr, 0, sep_lbl)
            self._grp_sep_labels.append(sep_lbl)

            # ── 데이터 행 ──
            for pos in range(ROWS_PER_GROUP):
                dr = gi * ROWS_PER_GROUP + pos
                tr = _d2t(dr)

                # Col 0: 행 번호
                ni = QTableWidgetItem(f"{letter}-{pos + 1}")
                ni.setTextAlignment(Qt.AlignCenter)
                ni.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                ni.setForeground(QColor(self._P.ACCENT_BLUE))
                tbl.setItem(tr, 0, ni)
                self._num_items.append(ni)

                # Col 1: CAS (단순 입력 — 조회는 상단 '일괄 조회' 버튼)
                cas_cell = _CellEdit(placeholder="CAS No.", style=_CELL_NUM,
                                     align=Qt.AlignLeft | Qt.AlignVCenter)
                cas_cell.edit.returnPressed.connect(lambda row=dr: self._lookup_cas(row))
                self._tag_paste(cas_cell.edit, dr, 0)
                tbl.setCellWidget(tr, 1, cas_cell)
                self._cas_edits.append(cas_cell.edit)

                # Col 2: 시약명
                ne_cell = _CellEdit(placeholder="시약명 입력", style=_CELL_EDIT,
                                    align=Qt.AlignLeft | Qt.AlignVCenter)
                ne_cell.edit.textChanged.connect(lambda t, row=dr: self._on_name(row, t))
                ne_cell.attach_fill(
                    lambda val, n, row=dr: self._fill_down(self._name_edits, row, val, n))
                tbl.setCellWidget(tr, 2, ne_cell)
                self._name_edits.append(ne_cell.edit)
                self._tag_paste(ne_cell.edit, dr, 1)

                # Col 3: MW
                mw_cell = _CellEdit()
                mw_cell.edit.textChanged.connect(lambda t, row=dr: (self._on_mw(row, t), self._recalc(row)))
                mw_cell.attach_fill(
                    lambda val, n, row=dr: self._fill_down(self._mw_edits, row, val, n))
                tbl.setCellWidget(tr, 3, mw_cell)
                self._mw_edits.append(mw_cell.edit)
                self._tag_paste(mw_cell.edit, dr, 2)

                # Col 4: 밀도 (g/mL) — 액상 시약일 때만 입력 (선택)
                # @codesyncer-decision: 밀도 변경 시 _recalc 호출하여 결과에 "원액 mL" 즉시 반영
                # @codesyncer-inference: _on_field 내부에서 _building=True 시 early return하지만,
                #   람다 체이닝이므로 _recalc는 별도 호출됨 → _building 중에도 _recalc 실행됨.
                #   다만 _building=True는 _on_cas_result에서만 설정되고 거기서 직접 _recalc를 호출하므로 중복 호출은 무해.
                d_cell = _CellEdit(placeholder="-")
                d_cell.edit.textChanged.connect(
                    lambda t, row=dr: (self._on_field(row, 'density', t), self._recalc(row)))
                d_cell.attach_fill(
                    lambda val, n, row=dr: self._fill_down(self._d_edits, row, val, n))
                tbl.setCellWidget(tr, 4, d_cell)
                self._d_edits.append(d_cell.edit)
                self._tag_paste(d_cell.edit, dr, 3)

                # Col 5~7: 부피/질량/농도 (입력 시 자동 계산)
                for col, store, attr in ((5, self._ml_edits, 'ml'),
                                         (6, self._g_edits,  'g'),
                                         (7, self._m_edits,  'conc')):
                    ec = _CellEdit()
                    ec.edit.textChanged.connect(
                        lambda t, row=dr, f=attr: (self._on_field(row, f, t), self._recalc(row))
                    )
                    ec.attach_fill(
                        lambda val, n, row=dr, s=store: self._fill_down(s, row, val, n))
                    tbl.setCellWidget(tr, col, ec)
                    store.append(ec.edit)
                    self._tag_paste(ec.edit, dr, col - 1)  # 부피=4, 질량=5, 농도=6

                # Col 8: 결과 라벨 (계산은 상단 '일괄 계산' 버튼 또는 입력 시 자동)
                rl = QLabel("-")
                rl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                rl.setStyleSheet(_CELL_RES + " padding: 0 6px;")
                rl.setAttribute(Qt.WA_TransparentForMouseEvents)  # 드래그 투과
                tbl.setCellWidget(tr, 8, rl)
                self._result_lbls.append(rl)

        root.addWidget(tbl)

        # ── 버튼 바 ──────────────────────────────────────────
        bar = QHBoxLayout()
        bar.setSpacing(T.SP_SM)

        btn_clr = QPushButton("전체 초기화")
        btn_clr.setMinimumHeight(T.H_BTN)
        btn_clr.clicked.connect(self._clear_all)
        bar.addWidget(btn_clr)

        btn_bulk = QPushButton("CAS 일괄 조회")
        btn_bulk.setMinimumHeight(T.H_BTN)
        btn_bulk.setToolTip("CAS 번호가 입력된 모든 행을 한 번에 조회합니다")
        btn_bulk.clicked.connect(self._bulk_lookup)
        bar.addWidget(btn_bulk)

        btn_calc_all = QPushButton("일괄 계산")
        btn_calc_all.setMinimumHeight(T.H_BTN)
        btn_calc_all.setToolTip("모든 행을 한 번에 계산합니다")
        btn_calc_all.clicked.connect(self._bulk_recalc)
        bar.addWidget(btn_calc_all)

        btn_copy_all = QPushButton("전체 복사")
        btn_copy_all.setMinimumHeight(T.H_BTN)
        btn_copy_all.setToolTip("모든 행의 데이터를 클립보드에 복사합니다 (엑셀 붙여넣기 가능)")
        btn_copy_all.clicked.connect(self._copy_all_conc)
        bar.addWidget(btn_copy_all)

        self._conc_btns_secondary = [btn_clr, btn_bulk, btn_calc_all, btn_copy_all]
        _sec = get_secondary_button_style(self._P)
        for btn in self._conc_btns_secondary:
            btn.setStyleSheet(_sec)

        bar.addStretch()

        btn_imp = QPushButton("▶  Sequence Import")
        btn_imp.setMinimumHeight(T.H_BTN_LG)
        btn_imp.setStyleSheet(
            f"QPushButton{{background:{self._P.ACCENT_BLUE};color:white;"
            f"font-size:14px;font-weight:bold;padding:0 20px;border-radius:4px;}}"
            f"QPushButton:hover{{background:{self._P.ACCENT_BLUE_DARK};}}"
        )
        btn_imp.clicked.connect(self._open_import)
        bar.addWidget(btn_imp)
        self._conc_btn_imp = btn_imp

        root.addLayout(bar)

    # ── WebGrid(Tabulator) 농도 테이블 ─────────────────────────
    def _build_conc_grid(self, root):
        """QTableWidget 대체 — Tabulator 임베드(엑셀 복붙/드래그선택/자동계산)."""
        from ui.webgrid.webgrid import WebGrid
        # 다성분 stock(A안): {dr: [자식 시약 dict(cas,name,mw,smiles,density,vol,mass,conc)]}
        # 포트행 = 첫 시약(self._rows[dr]), 추가 시약 = 자식행. Import 시 레시피로 번들.
        self._port_children = {}
        self._grid = WebGrid(self._grid_rows(), theme=self._grid_theme())
        self._grid.cellChanged.connect(self._on_grid_cell)
        root.addWidget(self._grid, 1)

        bar = QHBoxLayout()
        bar.setSpacing(T.SP_SM)
        btn_excel = QPushButton("Excel에서 편집")
        btn_excel.setMinimumHeight(T.H_BTN)
        btn_excel.setToolTip("계산기 표 전체를 Excel 파일로 열어 편집합니다.\n"
                             "Excel에서 저장하면 표에 자동 반영됩니다 (시퀀스 탭과 동일 방식).")
        btn_excel.clicked.connect(self._open_calc_excel)
        bar.addWidget(btn_excel)
        btn_calc = QPushButton("전체 계산")
        btn_calc.setMinimumHeight(T.H_BTN)
        btn_calc.setToolTip("모든 행을 재계산합니다.\n"
                            "MW + 부피(mL)·질량(g)·농도(M) 중 2개 → 나머지 자동.\n"
                            "(셀 편집 시 자동 계산되지만, 강제 재계산이 필요할 때 사용)")
        btn_calc.clicked.connect(self._grid_recalc_all)
        bar.addWidget(btn_calc)
        btn_clr = QPushButton("전체 초기화")
        btn_clr.setMinimumHeight(T.H_BTN)
        btn_clr.clicked.connect(self._grid_clear)
        bar.addWidget(btn_clr)
        btn_bulk = QPushButton("CAS 일괄 조회")
        btn_bulk.setMinimumHeight(T.H_BTN)
        btn_bulk.setToolTip("CAS 번호가 입력된 모든 행을 PubChem 으로 일괄 조회")
        btn_bulk.clicked.connect(self._bulk_lookup)
        bar.addWidget(btn_bulk)
        # @codesyncer(전체 복사 복원): WebGrid 전환 때 버튼 바에서 누락 — 구 _copy_all_conc 는
        #   Qt 셀 리스트(현재 빈 리스트)를 읽어 무동작이라 그리드 데이터 기반으로 재구현.
        btn_copy = QPushButton("전체 복사")
        btn_copy.setMinimumHeight(T.H_BTN)
        btn_copy.setToolTip("모든 행의 데이터를 클립보드에 복사합니다 (엑셀에 그대로 붙여넣기)")
        btn_copy.clicked.connect(self._grid_copy_all)
        bar.addWidget(btn_copy)
        self._conc_btns_secondary = [btn_excel, btn_calc, btn_clr, btn_bulk, btn_copy]
        _sec = get_secondary_button_style(self._P)
        for b in self._conc_btns_secondary:
            b.setStyleSheet(_sec)
        bar.addStretch()
        btn_imp = QPushButton("▶  Sequence Import")
        btn_imp.setMinimumHeight(T.H_BTN_LG)
        # 툴바 색 일관성 — 앱 표준 primary 헬퍼(라운드/폰트/hover 토큰) 사용, 인라인 하드코딩 제거
        btn_imp.setStyleSheet(get_primary_button_style(self._P))
        btn_imp.clicked.connect(self._open_import)
        bar.addWidget(btn_imp)
        self._conc_btn_imp = btn_imp
        root.addLayout(bar)

    def _grid_rows(self):
        """self._rows(_Row) + 그룹정보 → WebGrid 행 dict 리스트 (+다성분 자식행)."""
        rows = []
        children_map = getattr(self, "_port_children", {}) or {}
        for gi, (letter, pump) in enumerate(self._group_info):
            for pos in range(ROWS_PER_GROUP):
                dr = gi * ROWS_PER_GROUP + pos
                r = self._rows[dr]
                row = {
                    "_id": str(dr), "grp": letter, "port": pos + 2,
                    "cas": getattr(r, "cas", "") or "", "mw": r.mw or None,
                    "smiles": getattr(r, "smiles", "") or "", "name": r.name or "",
                    "density": r.density, "vol": r.ml, "mass": r.g,
                    "conc": r.conc, "result": "",
                }
                kids = children_map.get(dr) or []
                if kids:
                    row["mixN"] = len(kids) + 1        # 포트행(첫 시약) 포함
                    ch = []
                    for k, c in enumerate(kids, start=1):
                        ch.append({
                            "_id": f"{dr}:c{k}", "_comp": True,
                            "grp": letter, "port": "",
                            "cas": c.get("cas", "") or "", "mw": c.get("mw"),
                            "smiles": c.get("smiles", "") or "",
                            "name": c.get("name", "") or "",
                            "density": c.get("density"), "vol": c.get("vol"),
                            "mass": c.get("mass"), "conc": c.get("conc"),
                            "result": "",
                        })
                    ch.append({"_id": f"{dr}:add", "_add": True, "grp": letter,
                               "port": "", "name": "", "result": ""})
                    row["_children"] = ch
                rows.append(row)
        return rows

    # ── 다성분 자식행 조작 (시퀀스 시약그리드와 동일 UX) ──────────
    def _refresh_calc_grid(self, expand=None, edit=None):
        if getattr(self, "_grid", None) is None:
            return
        if expand is not None or edit is not None:
            import json as _json
            try:
                self._grid.view.page().runJavaScript(
                    f"window._expandNext = {_json.dumps(expand)};"
                    f"window._editNext = {_json.dumps(edit)};")
            except Exception:
                pass
        self._grid.replace_rows(self._grid_rows())

    def _calc_add_comp(self, dr_str):
        try:
            dr = int(str(dr_str))
        except (TypeError, ValueError):
            return
        if not (0 <= dr < self._num_rows):
            return
        kids = self._port_children.setdefault(dr, [])
        kids.append({"cas": "", "name": "", "mw": None, "smiles": "",
                     "density": None, "vol": None, "mass": None, "conc": None})
        self._refresh_calc_grid(expand=str(dr), edit=f"{dr}:c{len(kids)}")

    def _calc_del_comp(self, cid):
        try:
            dr_s, _, k_s = str(cid).partition(":c")
            dr, k = int(dr_s), int(k_s)
        except (TypeError, ValueError):
            return
        kids = self._port_children.get(dr) or []
        if not (1 <= k <= len(kids)):
            return
        del kids[k - 1]
        if not kids:
            self._port_children.pop(dr, None)
        self._refresh_calc_grid(expand=str(dr) if self._port_children.get(dr) else None)

    def _calc_edit_comp(self, d):
        try:
            dr_s, _, k_s = str(d.get("_id", "")).partition(":c")
            dr, k = int(dr_s), int(k_s)
        except (TypeError, ValueError):
            return
        kids = self._port_children.get(dr) or []
        if not (1 <= k <= len(kids)):
            return
        c = kids[k - 1]
        c["cas"] = d.get("cas", "") or ""
        c["name"] = d.get("name", "") or ""
        c["smiles"] = d.get("smiles", "") or ""
        for f in ("mw", "density", "vol", "mass", "conc"):
            c[f] = self._gf(d.get(f))

    def _grid_theme(self, P=None, is_dark=None):
        """앱 팔레트 → WebGrid CSS 변수 dict (공유 grid_theme 위임)."""
        from ui.webgrid.webgrid import grid_theme
        if is_dark is None:
            is_dark = getattr(self.app, 'is_dark_mode', True)
        if P is None:
            P = Dark if is_dark else Light
        return grid_theme(P, is_dark)

    @staticmethod
    def _gf(v):
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    def _on_grid_cell(self, d):
        """그리드 편집(자동계산 값 포함) → self._rows + map_mgr 동기화."""
        # 다성분 자식행 라우팅 (＋추가/✕삭제/자식 편집) — 포트행 동기보다 먼저
        if d.get("_addComp"):
            self._calc_add_comp(d["_addComp"])
            return
        if d.get("_delComp"):
            self._calc_del_comp(d["_delComp"])
            return
        if d.get("_comp"):
            self._calc_edit_comp(d)
            return
        if d.get("_add"):
            return
        if d.get("_bulk"):
            for gd in self._grid.rows():
                if not (gd.get("_comp") or gd.get("_add")):
                    self._sync_row_from_grid(gd)
            return
        self._sync_row_from_grid(d)
        # @codesyncer: CAS 셀 편집 시 자동 PubChem 조회(구 QTableWidget Enter 자동조회 복원).
        if d.get("_field") == "cas":
            cas = (d.get("cas") or "").strip()
            if cas:
                try:
                    self._start_cas_lookup(int(d.get("_id")), cas)
                except (TypeError, ValueError):
                    pass

    def _start_cas_lookup(self, dr, cas):
        """단일 행 CAS → PubChem 비동기 조회 워커 시작."""
        try:
            worker = _CasWorker(dr, cas)
            worker.result_ready.connect(self._on_cas_result)
            worker.lookup_error.connect(self._on_cas_error)
            worker.finished.connect(lambda w=worker: self._workers.discard(w))
            self._workers.add(worker)
            worker.start()
        except Exception as e:
            print(f"[CAS] {e}")

    def _sync_row_from_grid(self, d):
        try:
            dr = int(d.get("_id"))
        except (TypeError, ValueError):
            return
        if not (0 <= dr < len(self._rows)):
            return
        r = self._rows[dr]
        r.cas = d.get("cas", "") or ""
        r.name = d.get("name", "") or ""
        r.smiles = d.get("smiles", "") or ""
        r.mw = self._gf(d.get("mw")) or 0.0
        r.density = self._gf(d.get("density"))
        r.ml = self._gf(d.get("vol"))
        r.g = self._gf(d.get("mass"))
        r.conc = self._gf(d.get("conc"))
        # @codesyncer(버그수정 2026-07-13): 농도표 QTableWidget→WebGrid 전환 후
        #   _recalc 가 결과라벨 부재로 조기 return → eff_conc(Import·희석이 읽는 유효농도
        #   진실원)가 영영 None → 이름·농도를 입력해도 "설정된 행 없음"으로 Import 차단.
        #   그리드 JS recalc 가 파생 몰농도(질량+부피+MW→M)를 d.conc 에 이미 써서
        #   되돌리므로 r.conc 가 곧 유효농도 — 이를 eff_conc 로 미러(단일 진실원 복원).
        r.eff_conc = r.conc if (r.conc is not None and r.conc > 0) else None
        # map_mgr 동기화 (농도 → inlet_map) + reagent_lib(SMILES)
        gi, pos = dr // ROWS_PER_GROUP, dr % ROWS_PER_GROUP
        if gi < len(self._group_info):
            _, pump = self._group_info[gi]
            port = pos + 2
            if self.app.map_mgr and r.name:
                try:
                    self.app.map_mgr.update_inlet(pump, port, r.name, r.conc or 0.0, r.smiles or None)
                except Exception as e:
                    print(f"[Calc→map] {e}")

    # ── Excel에서 편집 (시퀀스 탭 ReagentExcelManager 와 동일 UX) ──
    #   열기 → 사용자가 Excel에서 편집/저장 → 파일 감시로 자동 반영.
    #   반영은 (Group, Port) 키 매칭 → set_rows 부분 갱신 → recalc_all 로
    #   자동계산·_rows 미러·map_mgr 동기화가 기존 경로(_bulk 콜백)를 그대로 탄다.

    def _calc_excel_file(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        temp_dir = os.path.join(root, "temp")
        os.makedirs(temp_dir, exist_ok=True)
        return os.path.join(temp_dir, "calc_reagents.xlsx")

    def _open_calc_excel(self):
        """계산기 표 전체를 Excel 로 열어 편집 (저장 시 자동 반영)."""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "경고",
                "openpyxl 패키지가 설치되지 않았습니다.\npip install openpyxl 명령으로 설치하세요.")
            return
        if self._grid is None:
            return

        path = self._calc_excel_file()
        # 이미 Excel 이 잠그고 있으면 재생성 불가 → 안내만 (저장하면 자동 반영됨)
        if os.path.exists(path):
            try:
                with open(path, 'a'):
                    pass
            except PermissionError:
                QMessageBox.information(self, "Excel 편집",
                    "이미 Excel에서 열려 있습니다.\n저장하면 자동 반영됩니다.")
                return

        from core.reagent_excel import ReagentExcelManager
        header_font, header_fill, center_align, thin_border = ReagentExcelManager._excel_styles()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "농도 계산기"
        # @codesyncer(다성분 왕복, 2026-07-13): '성분' 컬럼 신설 — 부모(포트행)=1,
        #   추가 시약(자식행)=2..N. Excel 에서 성분 2+ 행을 추가/수정/삭제하면
        #   _apply_calc_excel 이 포트 단위로 자식 시약을 전량 교체 반영한다.
        headers = ["Group", "Port", "성분", "시약명", "CAS", "MW (g/mol)", "SMILES",
                   "밀도 (g/mL)", "부피 (mL)", "질량 (g)", "농도 (M)"]
        widths = [10, 8, 8, 22, 14, 12, 26, 12, 10, 10, 10]
        for c, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = w
        r = 2
        for gd in self._grid.rows():
            vals = [gd.get("grp"), gd.get("port"), 1, gd.get("name") or "",
                    gd.get("cas") or "", gd.get("mw"), gd.get("smiles") or "",
                    gd.get("density"), gd.get("vol"), gd.get("mass"), gd.get("conc")]
            for c, v in enumerate(vals, start=1):
                ws.cell(row=r, column=c, value=v).border = thin_border
            r += 1
            try:
                dr = int(gd.get("_id"))
            except (TypeError, ValueError):
                dr = None
            for k, ch in enumerate(self._port_children.get(dr, []) or [], start=2):
                vals = [gd.get("grp"), gd.get("port"), k, ch.get("name") or "",
                        ch.get("cas") or "", ch.get("mw"), ch.get("smiles") or "",
                        ch.get("density"), ch.get("vol"), ch.get("mass"), ch.get("conc")]
                for c, v in enumerate(vals, start=1):
                    ws.cell(row=r, column=c, value=v).border = thin_border
                r += 1

        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.information(self, "Excel 편집",
                "이미 Excel에서 열려 있습니다.\n저장하면 자동 반영됩니다.")
            return
        self._calc_excel_path = path

        if self._calc_excel_watcher is None:
            self._calc_excel_watcher = QFileSystemWatcher()
            self._calc_excel_watcher.fileChanged.connect(self._on_calc_excel_changed)
        if path not in self._calc_excel_watcher.files():
            self._calc_excel_watcher.addPath(path)
        os.startfile(path)   # Windows 전용 (reagent_excel 과 동일 전제)

    def _on_calc_excel_changed(self, path):
        """저장 감지 → 1초 후 반영(Excel 파일락 해제 대기) + 워처 재등록."""
        if path == self._calc_excel_path:
            QTimer.singleShot(1000, self._apply_calc_excel)
            QTimer.singleShot(1500, self._re_register_calc_excel_watcher)

    def _re_register_calc_excel_watcher(self):
        # Windows QFileSystemWatcher 는 변경 이벤트 후 해제되는 경우가 있음
        if (self._calc_excel_watcher is not None and self._calc_excel_path
                and os.path.exists(self._calc_excel_path)
                and self._calc_excel_path not in self._calc_excel_watcher.files()):
            self._calc_excel_watcher.addPath(self._calc_excel_path)

    def _apply_calc_excel(self):
        """Excel 내용 → 그리드/자식 시약 반영. (Group, Port[, 성분]) 키 매칭.

        @codesyncer(다성분 왕복, 2026-07-13): 성분 1=부모(포트행) → self._rows 직접
        반영(권위 소스), 성분 2+=자식 시약 → 포트 단위 '전량 교체'(Excel 에서 행을
        지우면 그리드에서도 제거). 반영 후 전체 리빌드(_refresh_calc_grid, 펼침 보존)
        → JS recalc → 지연 recalcAll(_bulk)로 파생 conc·eff_conc·map 동기(기존 경로).
        구형 파일(성분 컬럼 없음)은 부모만 갱신하는 하위호환.
        """
        if self._grid is None or not self._calc_excel_path:
            return
        if not os.path.exists(self._calc_excel_path):
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._calc_excel_path, data_only=True)
        except Exception as e:
            print(f"[CalcExcel] 로드 실패: {e}")
            return
        ws = wb.active
        hdr = [c.value for c in ws[1]]
        has_comp = len(hdr) > 2 and str(hdr[2] or "").strip() == "성분"
        off = 1 if has_comp else 0

        def s(v):
            return "" if v is None else str(v).strip()

        key2dr = {}
        for gi, (letter, _pump) in enumerate(self._group_info):
            for pos in range(ROWS_PER_GROUP):
                key2dr[(str(letter), pos + 2)] = gi * ROWS_PER_GROUP + pos

        parents, children, ports_seen = {}, {}, set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 10 + off or row[1] in (None, ""):
                continue
            try:
                dr = key2dr.get((str(row[0] or "").strip(), int(row[1])))
            except (TypeError, ValueError):
                continue
            if dr is None:
                continue
            comp = 1
            if has_comp:
                try:
                    comp = int(row[2] or 1)
                except (TypeError, ValueError):
                    comp = 1
            vals = {"name": s(row[2 + off]), "cas": s(row[3 + off]),
                    "mw": self._gf(row[4 + off]), "smiles": s(row[5 + off]),
                    "density": self._gf(row[6 + off]), "vol": self._gf(row[7 + off]),
                    "mass": self._gf(row[8 + off]), "conc": self._gf(row[9 + off])}
            ports_seen.add(dr)
            if comp <= 1:
                parents[dr] = vals
            else:
                children.setdefault(dr, []).append((comp, vals))
        if not ports_seen:
            return

        for dr, v in parents.items():
            r = self._rows[dr]
            r.name = v["name"]
            r.cas = v["cas"]
            r.smiles = v["smiles"]
            r.mw = v["mw"] or 0.0
            r.density = v["density"]
            r.ml = v["vol"]
            r.g = v["mass"]
            r.conc = v["conc"]
        if has_comp:
            for dr in ports_seen:
                ks = sorted(children.get(dr, []), key=lambda t: t[0])
                if ks:
                    self._port_children[dr] = [dict(v) for _c, v in ks]
                else:
                    self._port_children.pop(dr, None)

        self._refresh_calc_grid()
        QTimer.singleShot(300, self._grid_recalc_all)
        if hasattr(self.app, 'log_browser') and self.app.log_browser:
            n_kids = sum(len(v) for v in children.values())
            self.app.signals.sig_log.emit(
                f"계산기: Excel 변경 반영 (포트 {len(ports_seen)}"
                + (f" · 추가 시약 {n_kids}" if has_comp else "") + ")")

    def _grid_recalc_all(self):
        """모든 행 강제 재계산 (자동계산 외 명시적 '전체 계산' 버튼용)."""
        if getattr(self, '_grid', None) is not None:
            self._grid.recalc_all()

    def _grid_clear(self):
        for r in self._rows:
            r.cas = ""; r.name = ""; r.smiles = ""; r.mw = 0.0
            r.density = None; r.ml = None; r.g = None; r.conc = None
        self._port_children = {}
        self._refresh_calc_grid()

    def _build_dil_tab(self, page):
        """일괄 희석 계산기: 농도 계산기 데이터 → 목표 농도로 희석"""
        lay = QVBoxLayout(page)
        lay.setContentsMargins(T.SP_MD, T.SP_MD, T.SP_MD, T.SP_MD)
        lay.setSpacing(10)

        # 제목은 WebGrid 카드 헤더("일괄 희석")가 담당 — Qt 중복 제목 제거.
        # 구 참조(_apply_dil_styles)의 가드용 None 스텁.
        self._dil_title_lbl = None
        self._dil_sub_lbl = None

        # ── 전역 설정 바 ──────────────────────────────────────
        ctrl = QHBoxLayout()
        ctrl.setSpacing(6)

        P = self._P
        _ctrl_input_ss = (
            f"QLineEdit{{"
            f"background:{P.BG_INPUT};color:{P.TEXT_INPUT};"
            f"border:1px solid {P.BORDER_INPUT};border-radius:3px;"
            f"padding:2px 6px;font-size:12px;}}"
            f"QLineEdit:focus{{border-color:{P.ACCENT_BLUE};}}"
        )

        def _ctrl_line(placeholder, width=80):
            ed = QLineEdit()
            ed.setPlaceholderText(placeholder)
            ed.setFixedWidth(width)
            ed.setFixedHeight(26)
            ed.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            ed.setStyleSheet(_ctrl_input_ss)
            return ed

        def _ctrl_lbl(text):
            l = QLabel(text)
            l.setStyleSheet(f"font-size:12px;color:{self._P.TEXT_SECONDARY};")
            self._dil_ctrl_labels.append(l)
            return l

        ctrl.addWidget(_ctrl_lbl("목표 농도:"))
        self._dil_g_conc = _ctrl_line("0.000")
        ctrl.addWidget(self._dil_g_conc)
        ctrl.addWidget(_ctrl_lbl("M"))

        ctrl.addSpacing(T.SP_MD)

        ctrl.addWidget(_ctrl_lbl("총 부피:"))
        self._dil_g_vol = _ctrl_line("0.000")
        ctrl.addWidget(self._dil_g_vol)
        ctrl.addWidget(_ctrl_lbl("mL"))

        ctrl.addStretch()

        btn_refresh = QPushButton("↻  시약 불러오기")
        btn_refresh.setMinimumHeight(T.H_BTN)
        btn_refresh.setToolTip("농도 계산기에서 시약명과 원액 농도를 가져옵니다")
        btn_refresh.clicked.connect(self._dil_refresh)
        ctrl.addWidget(btn_refresh)

        lay.addLayout(ctrl)

        # ── 희석 WebGrid (Tabulator) — 농도 계산기와 동일 스택 ──
        # @codesyncer(마이그레이션 2026-07-13, 사용자 승인 "희석 계산기도 webgrid로"):
        #   촘촘한 QTableWidget+셀위젯 → WebGrid. 계산은 JS recalc
        #   (V1=C2·V2/C1, 용매=V2−V1, C2>C1 은 '원액 부족' 배지), 엑셀 복붙/범위선택/
        #   fill-down 은 Tabulator 내장으로 대체. 구 위젯배열은 빈 스텁으로 유지 —
        #   _apply_dil_styles·eventFilter 등 잔존 참조가 전부 no-op 이 되도록.
        self._dil_tbl = None
        self._dil_name_lbls = []
        self._dil_stock_lbls = []
        self._dil_target_eds = []
        self._dil_totalv_eds = []
        self._dil_v1_lbls = []
        self._dil_solv_lbls = []

        from ui.webgrid.webgrid import WebGrid
        self._dil_grid = WebGrid(self._dil_grid_rows(), html="dil_grid.html",
                                 theme=self._grid_theme())
        lay.addWidget(self._dil_grid, 1)

        # ── 버튼 바 ──
        bar = QHBoxLayout()
        bar.setSpacing(T.SP_SM)

        btn_apply = QPushButton("일괄 적용")
        btn_apply.setMinimumHeight(T.H_BTN)
        btn_apply.setToolTip("위의 목표 농도·총 부피를 원액이 있는 모든 행에 채웁니다")
        btn_apply.clicked.connect(self._dil_apply_global)
        bar.addWidget(btn_apply)

        btn_calc = QPushButton("전체 계산")
        btn_calc.setMinimumHeight(T.H_BTN)
        btn_calc.setToolTip("모든 행을 강제 재계산합니다 (셀 편집 시 자동 계산됨)")
        btn_calc.clicked.connect(self._dil_calc_all)
        bar.addWidget(btn_calc)

        btn_clr = QPushButton("입력 초기화")
        btn_clr.setMinimumHeight(T.H_BTN)
        btn_clr.setToolTip("목표 농도·총 부피 입력을 지웁니다 (결과는 자동 소거)")
        btn_clr.clicked.connect(self._dil_clear)
        bar.addWidget(btn_clr)

        btn_copy_dil = QPushButton("전체 복사")
        btn_copy_dil.setMinimumHeight(T.H_BTN)
        btn_copy_dil.setToolTip("희석 표 전체를 클립보드에 복사합니다 (엑셀에 그대로 붙여넣기)")
        btn_copy_dil.clicked.connect(self._copy_all_dil)
        bar.addWidget(btn_copy_dil)

        self._dil_btns_secondary = [btn_refresh, btn_apply, btn_calc, btn_clr, btn_copy_dil]
        _sec = get_secondary_button_style(self._P)
        for btn in self._dil_btns_secondary:
            btn.setStyleSheet(_sec)

        bar.addStretch()
        lay.addLayout(bar)

    def _dil_grid_rows(self):
        """농도 계산기 행 → 희석 그리드 행 (원액 C1 = eff_conc, V2 프리필 = 용액 부피)."""
        rows = []
        for gi, (letter, _pump) in enumerate(self._group_info):
            for pos in range(ROWS_PER_GROUP):
                dr = gi * ROWS_PER_GROUP + pos
                r = self._rows[dr]
                rows.append({
                    "_id": str(dr), "grp": letter, "port": pos + 2,
                    "name": r.name or "",
                    "c1": round(r.eff_conc, 6) if (r.eff_conc and r.eff_conc > 0) else None,
                    "c2": None,
                    "v2": round(r.ml, 4) if (r.ml and r.ml > 0) else None,
                    "v1": None, "solv": None, "status": "",
                })
        return rows

    # ── RTD 탭 ────────────────────────────────────────────────

    def _build_rtd_tab(self, page):
        """RTD 계산기 탭 — 좌: 입력+카드  우: 그래프+해석"""
        lay = QVBoxLayout(page)
        lay.setContentsMargins(T.SP_MD, T.SP_MD, T.SP_MD, T.SP_MD)
        lay.setSpacing(0)

        # ── 공통 스타일 ────────────────────────────────────────
        P = self._P
        _INPUT_H = 28
        _spin_ss = (
            f"QDoubleSpinBox{{background:{P.BG_INPUT};color:{P.TEXT_INPUT};"
            f"border:1px solid {P.BORDER_INPUT};border-radius:3px;"
            f"padding:2px 6px;font-size:12px;}}"
            f"QDoubleSpinBox:focus{{border-color:{P.ACCENT_BLUE};}}"
        )
        _combo_ss = (
            f"QComboBox{{background:{P.BG_INPUT};color:{P.TEXT_INPUT};"
            f"border:1px solid {P.BORDER_INPUT};border-radius:3px;"
            f"padding:2px 6px;font-size:12px;}}"
            f"QComboBox::drop-down{{border:none;}}"
            f"QComboBox QAbstractItemView{{background:{P.BG_SECONDARY};"
            f"color:{P.TEXT_PRIMARY};"
            f"selection-background-color:{P.BG_HOVER};}}"
        )

        _W = 120  # 입력 위젯 공통 너비

        def _spin(lo, hi, val, dec, suffix):
            s = QDoubleSpinBox()
            s.setRange(lo, hi); s.setValue(val)
            s.setDecimals(dec); s.setSuffix(suffix)
            s.setFixedWidth(_W); s.setFixedHeight(_INPUT_H)
            s.setAlignment(Qt.AlignCenter)
            s.setStyleSheet(_spin_ss)
            return s

        def _rlbl(t):
            l = QLabel(t)
            l.setFixedWidth(72)
            l.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            l.setStyleSheet(
                f"font-size:12px;color:{P.TEXT_SECONDARY};padding-right:6px;"
            )
            return l

        # ── 메인 스플리터 ─────────────────────────────────────
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        splitter.setStyleSheet(
            f"QSplitter::handle{{background:{P.BORDER_PRIMARY};width:1px;}}"
        )
        self._rtd_splitter = splitter

        # ════════════════════════════════
        # 왼쪽: 입력 + 계산 버튼 + 카드
        # ════════════════════════════════
        left_w = QWidget()
        left_w.setMinimumWidth(300)
        left_w.setMaximumWidth(400)
        left_lay = QVBoxLayout(left_w)
        left_lay.setContentsMargins(0, 0, 10, 0)
        left_lay.setSpacing(T.SP_SM)

        # 입력 프레임
        inp_frame = QFrame()
        inp_frame.setFrameShape(QFrame.StyledPanel)
        inp_frame.setStyleSheet(
            f"QFrame{{background:{P.BG_SECONDARY};"
            f"border:1px solid {P.BORDER_PRIMARY};border-radius:6px;}}"
        )
        inp_lay = QVBoxLayout(inp_frame)
        inp_lay.setContentsMargins(T.SP_MD, 10, T.SP_MD, 10)
        inp_lay.setSpacing(7)

        sec_lbl = QLabel("반응기 조건")
        sec_lbl.setStyleSheet(
            f"font-size:11px;font-weight:bold;color:{P.TEXT_SECONDARY};"
            f"background:transparent;border:none;letter-spacing:0.5px;"
        )
        inp_lay.addWidget(sec_lbl)
        self._rtd_inp_frame = inp_frame
        self._rtd_sec_lbl = sec_lbl

        sp = getattr(self.app.cfg, '_sys_params', {}) if hasattr(self.app, 'cfg') else {}
        self._rtd_sp_id   = _spin(0.1,  10.0, sp.get('reactor_id_mm', 1.0), 2, " mm")
        self._rtd_sp_len  = _spin(0.1, 100.0, sp.get('reactor_len_m',  5.0), 2, " m")
        self._rtd_sp_flow = _spin(0.01, 50.0, 0.5,  3, " mL/min")
        self._rtd_sp_temp = _spin(0.0, 200.0, 25.0, 1, " °C")

        self._rtd_cb_solv = QComboBox()
        self._rtd_cb_solv.setFixedWidth(_W)
        self._rtd_cb_solv.setFixedHeight(_INPUT_H)
        self._rtd_cb_solv.setStyleSheet(_combo_ss)
        for n in _SOLVENTS:
            self._rtd_cb_solv.addItem(n)
        self._rtd_cb_solv.setCurrentText("DCM")
        self._rtd_cb_solv.setEditable(True)
        self._rtd_cb_solv.lineEdit().setReadOnly(True)
        self._rtd_cb_solv.lineEdit().setAlignment(Qt.AlignCenter)

        g = QGridLayout()
        g.setHorizontalSpacing(8); g.setVerticalSpacing(6)
        g.setColumnMinimumWidth(0, 72)
        g.setColumnMinimumWidth(1, _W)
        for r, (lbl, wgt) in enumerate([
            (_rlbl("반응기 내경"), self._rtd_sp_id),
            (_rlbl("반응기 길이"), self._rtd_sp_len),
            (_rlbl("유량"),        self._rtd_sp_flow),
            (_rlbl("온도"),        self._rtd_sp_temp),
            (_rlbl("용매"),        self._rtd_cb_solv),
        ]):
            self._rtd_rlabels.append(lbl)
            g.addWidget(lbl, r, 0)
            g.addWidget(wgt, r, 1)
        g.setColumnStretch(2, 1)
        inp_lay.addLayout(g)

        btn_calc = QPushButton("RTD 계산")
        btn_calc.setFixedHeight(T.H_BTN)
        btn_calc.setStyleSheet(
            f"QPushButton{{background:{P.ACCENT_BLUE};color:white;"
            f"font-size:13px;font-weight:bold;border-radius:4px;border:none;}}"
            f"QPushButton:hover{{background:{P.ACCENT_BLUE_DARK};}}"
        )
        btn_calc.clicked.connect(self._calc_rtd)
        inp_lay.addWidget(btn_calc)
        left_lay.addWidget(inp_frame)
        self._rtd_btn_calc = btn_calc

        # 구분선
        sep = QFrame(); sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f"color:{P.BORDER_PRIMARY};")
        left_lay.addWidget(sep)
        self._rtd_sep = sep

        # 결과 카드 4개 (세로 스택)
        self._rtd_cards = {}
        _card_ss = (
            f"QFrame{{background:{P.HEADER_BG};"
            f"border:1px solid {P.BORDER_SECONDARY};border-radius:5px;}}"
        )
        for key, title in [
            ("tau",    "평균 반응 시간"),
            ("spread", "반응시간 범위 (80%)"),
            ("flow",   "흐름 균일도"),
            ("vol",    "반응기 부피"),
        ]:
            card = QFrame()
            card.setFrameShape(QFrame.StyledPanel)
            card.setStyleSheet(_card_ss)
            card.setFixedHeight(60)
            cl = QHBoxLayout(card)
            cl.setContentsMargins(14, T.SP_SM, 14, T.SP_SM)
            cl.setSpacing(0)
            t = QLabel(title)
            t.setStyleSheet(
                f"color:{P.TEXT_SECONDARY};font-size:12px;"
                f"background:transparent;border:none;"
            )
            v = QLabel("—")
            v.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            v.setStyleSheet(
                f"color:{P.TEXT_PRIMARY};font-size:16px;font-weight:bold;"
                f"background:transparent;border:none;"
            )
            cl.addWidget(t, 1)
            cl.addWidget(v)
            left_lay.addWidget(card)
            self._rtd_cards[key] = v
            self._rtd_card_frames.append(card)
            self._rtd_card_titles.append(t)

        left_lay.addStretch()
        splitter.addWidget(left_w)

        # ════════════════════════════════
        # 오른쪽: 그래프 + 해석 텍스트
        # ════════════════════════════════
        right_w = QWidget()
        right_lay = QVBoxLayout(right_w)
        right_lay.setContentsMargins(T.SP_SM, 0, 0, 0)
        right_lay.setSpacing(6)

        self._rtd_plot = pg.PlotWidget()
        self._rtd_plot.setBackground(P.BG_MONITOR)
        self._rtd_plot.setLabel("bottom", "시간 (min)",
                                color=P.TEXT_SECONDARY, size="10pt")
        self._rtd_plot.setLabel("left", "E(t)",
                                color=P.TEXT_SECONDARY, size="10pt")
        self._rtd_plot.showGrid(x=True, y=True, alpha=0.18)
        for ax in ("bottom", "left"):
            self._rtd_plot.getAxis(ax).setPen(
                pg.mkPen(P.BORDER_TABLE, width=1))
            self._rtd_plot.getAxis(ax).setTextPen(P.TEXT_SECONDARY)
        self._rtd_plot.setMinimumHeight(160)

        self._rtd_curve_item = self._rtd_plot.plot(
            pen=pg.mkPen(P.ACCENT_BLUE, width=2.5))
        self._rtd_tau_line = pg.InfiniteLine(
            angle=90, movable=False,
            pen=pg.mkPen(P.TEXT_TERTIARY, width=1, style=Qt.DashLine))
        self._rtd_plot.addItem(self._rtd_tau_line)
        self._rtd_region = pg.LinearRegionItem(
            values=(0, 0), movable=False,
            brush=pg.mkBrush(255, 160, 50, 35),
            pen=pg.mkPen(255, 160, 50, 100, width=1))
        self._rtd_region.setVisible(False)
        self._rtd_plot.addItem(self._rtd_region)
        right_lay.addWidget(self._rtd_plot, stretch=1)

        hint_lbl = QLabel(
            "좁고 뾰족한 곡선 = 균일한 반응시간  │  주황 영역 = 시약 80% 통과 구간"
        )
        hint_lbl.setStyleSheet(f"color:{P.TEXT_DISABLED};font-size:11px;")
        hint_lbl.setAlignment(Qt.AlignCenter)
        right_lay.addWidget(hint_lbl)
        self._rtd_hint_lbl = hint_lbl

        # 해석 텍스트 (그래프 아래)
        self._rtd_interp = QLabel("")
        self._rtd_interp.setWordWrap(True)
        self._rtd_interp.setAlignment(Qt.AlignTop | Qt.AlignHCenter)
        self._rtd_interp.setTextFormat(Qt.RichText)
        self._rtd_interp.setStyleSheet(
            f"background:{P.BG_TERTIARY};"
            f"border:1px solid {P.BORDER_PRIMARY};"
            f"border-radius:5px;padding:12px;font-size:13px;"
            f"color:{P.TEXT_PRIMARY};"
        )
        self._rtd_interp.hide()
        right_lay.addWidget(self._rtd_interp)

        splitter.addWidget(right_w)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        lay.addWidget(splitter, stretch=1)

    def _calc_rtd(self):
        id_mm   = self._rtd_sp_id.value()
        len_m   = self._rtd_sp_len.value()
        flow    = self._rtd_sp_flow.value()
        T_K     = self._rtd_sp_temp.value() + 273.15
        solvent = self._rtd_cb_solv.currentText()

        visc, rho = _SOLVENTS[solvent]
        res = _rtd_calc(id_mm, len_m, flow, visc, rho, T_K)
        tau, Pe, Re, Gz = res["tau_min"], res["Pe"], res["Re"], res["Gz"]
        V_ml = res["V_ml"]

        sigma_t   = tau * (2.0 / Pe) ** 0.5
        sigma_pct = sigma_t / tau * 100.0

        # t10, t90 (CDF 사다리꼴 적분)
        t_vals, E_vals = _rtd_curve(tau, Pe, n_points=800)
        dt  = t_vals[1] - t_vals[0]
        cdf = 0.0
        t10 = t90 = tau
        f10 = f90 = False
        for ti, ei in zip(t_vals, E_vals):
            cdf += ei * dt
            if not f10 and cdf >= 0.10: t10 = ti; f10 = True
            if not f90 and cdf >= 0.90: t90 = ti; f90 = True
        spread_min = t90 - t10

        # 등급
        P = self._P
        if Pe >= 100:
            grade = "우수";       grade_color = P.ACCENT_GREEN
            tips  = "흐름이 매우 균일합니다. 현재 조건을 유지하세요."
        elif Pe >= 30:
            E = DarkExtras if getattr(self.app, 'is_dark_mode', True) else LightExtras
            grade = "보통";       grade_color = E.WARNING
            tips  = "분산이 다소 있습니다. 유량을 낮추거나 튜브를 더 길게 하면 개선됩니다."
        else:
            grade = "주의 필요";  grade_color = P.ACCENT_RED
            tips  = ("분산이 큽니다. 반응 재현성이 낮을 수 있습니다. "
                     "유량을 낮추거나 내경이 더 좁은 튜브 사용을 권장합니다.")

        # 경고
        warns = []
        if Re > 2100:
            warns.append(f"Re = {Re:.0f} &gt; 2100: 난류 구간 — 계산 신뢰도 낮음")
        if Gz < 1.0:
            warns.append(f"Gz = {Gz:.3f} &lt; 1: 분산 발달 미충족 — 실제 분산이 더 클 수 있음")

        # 카드 업데이트
        _mono_ss = (
            f"color:{P.TEXT_PRIMARY};font-size:16px;font-weight:bold;"
            f"background:transparent;border:none;"
        )
        _accent_ss = (
            f"color:{P.ACCENT_BLUE};font-size:14px;font-weight:bold;"
            f"background:transparent;border:none;"
        )
        self._rtd_cards["tau"].setText(f"{tau:.2f} min")
        self._rtd_cards["tau"].setStyleSheet(_mono_ss)

        self._rtd_cards["spread"].setText(f"{t10:.2f} ~ {t90:.2f} min")
        self._rtd_cards["spread"].setStyleSheet(_accent_ss)

        self._rtd_cards["flow"].setText(grade)
        self._rtd_cards["flow"].setStyleSheet(
            f"color:{grade_color};font-size:16px;font-weight:bold;"
            f"background:transparent;border:none;")

        self._rtd_cards["vol"].setText(f"{V_ml:.3f} mL")
        self._rtd_cards["vol"].setStyleSheet(_mono_ss)

        # 해석 텍스트
        warn_html = "".join(
            f"<br><span style='color:{P.ACCENT_RED};font-size:12px;'>* {w}</span>"
            for w in warns
        )
        self._rtd_interp.setText(
            f"<b style='color:{grade_color};font-size:14px;'>흐름 등급: {grade}</b>"
            f"<br><br>"
            f"평균 체류시간 <b>{tau:.2f}분</b> 기준으로, 시약의 80%가 "
            f"<b style='color:{P.ACCENT_BLUE};'>{t10:.2f} ~ {t90:.2f}분</b> 사이에 "
            f"반응기를 통과합니다 (분포 폭 {spread_min:.2f}분, ±{sigma_pct:.1f}%)."
            f"<br><br>"
            f"<span style='color:{P.TEXT_SECONDARY};'>{tips}</span>"
            f"{warn_html}"
        )
        self._rtd_interp.show()

        # 그래프 업데이트
        self._rtd_curve_item.setData(t_vals, E_vals)
        self._rtd_tau_line.setValue(tau)
        self._rtd_region.setRegion((t10, t90))
        self._rtd_region.setVisible(True)

        for item in list(self._rtd_plot.items()):
            if isinstance(item, pg.TextItem):
                self._rtd_plot.removeItem(item)
        E_peak = max(E_vals) if E_vals else 1.0
        for text, pos, color in [
            (f"τ = {tau:.2f}", tau, P.TEXT_TERTIARY),
            ("10%",            t10, P.ACCENT_BLUE),
            ("90%",            t90, P.ACCENT_BLUE),
        ]:
            ti = pg.TextItem(text, color=color, anchor=(0.5, 1.15))
            ti.setPos(pos, E_peak * 0.04)
            self._rtd_plot.addItem(ti)

    # ── 희석 탭 메서드 ─────────────────────────────────────────

    def _dil_refresh(self):
        """농도 계산기 데이터 → 희석 그리드 (이름·원액 C1·V2 프리필 부분갱신)."""
        if getattr(self, "_dil_grid", None) is None:
            return
        ups = []
        for dr in range(self._num_rows):
            r = self._rows[dr]
            u = {"_id": str(dr), "name": r.name or "",
                 "c1": round(r.eff_conc, 6) if (r.eff_conc and r.eff_conc > 0) else None}
            if r.ml and r.ml > 0:
                u["v2"] = round(r.ml, 4)
            ups.append(u)
        self._dil_grid.set_rows(ups)   # JS updateRows 가 행별 recalc 수행

    def _dil_apply_global(self):
        """전역 목표 농도 / 총 부피를 원액(C1)이 있는 행에 일괄 적용."""
        if getattr(self, "_dil_grid", None) is None:
            return
        tc = _parse(self._dil_g_conc.text())
        tv = _parse(self._dil_g_vol.text())
        ups = []
        for gd in self._dil_grid.rows():
            if not gd.get("c1"):
                continue
            u = {"_id": gd["_id"]}
            if tc:
                u["c2"] = tc
            if tv:
                u["v2"] = tv
            if len(u) > 1:
                ups.append(u)
        if ups:
            self._dil_grid.set_rows(ups)

    def _dil_calc_row(self, dr):
        """V₁ = C₂ × V₂ / C₁,  용매 = V₂ - V₁"""
        P = self._P
        row_done_bg = "#0d2010" if getattr(self.app, 'is_dark_mode', True) else "#d6eedc"
        _done = _CELL_RES + f" font-weight:bold; color:{P.TEXT_PRIMARY}; background:{row_done_bg};"
        _err  = _CELL_RES + f" color:{P.ACCENT_RED};"
        _rst  = _CELL_RES

        def _clear():
            self._dil_v1_lbls[dr].setText("-");  self._dil_v1_lbls[dr].setStyleSheet(_rst)
            self._dil_solv_lbls[dr].setText("-"); self._dil_solv_lbls[dr].setStyleSheet(_rst)

        try:
            c1 = float(self._dil_stock_lbls[dr].text())
            c2 = float(self._dil_target_eds[dr].text().replace(',', '.'))
            v2 = float(self._dil_totalv_eds[dr].text().replace(',', '.'))
        except (ValueError, AttributeError):
            _clear(); return

        if c1 <= 0 or c2 <= 0 or v2 <= 0:
            _clear(); return

        v1   = c2 * v2 / c1
        solv = v2 - v1

        if v1 > v2:
            self._dil_v1_lbls[dr].setText("원액 부족")
            self._dil_v1_lbls[dr].setStyleSheet(_err)
            self._dil_solv_lbls[dr].setText("-")
            self._dil_solv_lbls[dr].setStyleSheet(_rst)
            return

        self._dil_v1_lbls[dr].setText(f"{v1:.4f}")
        self._dil_v1_lbls[dr].setStyleSheet(_done)
        self._dil_solv_lbls[dr].setText(f"{solv:.4f}")
        self._dil_solv_lbls[dr].setStyleSheet(_done)

    def _dil_calc_all(self):
        """모든 행 강제 재계산 — JS recalcAll (셀 편집 시엔 자동)."""
        if getattr(self, "_dil_grid", None) is not None:
            self._dil_grid.recalc_all()

    def _dil_clear(self):
        """목표 농도·총 부피 입력 초기화 (결과/상태는 recalc 가 자동 소거)."""
        if getattr(self, "_dil_grid", None) is None:
            return
        self._dil_grid.set_rows([{"_id": str(dr), "c2": None, "v2": None}
                                 for dr in range(self._num_rows)])

    # ── Fill-down 헬퍼 ────────────────────────────────────────

    def _fill_down(self, edit_list, start_row, value, n):
        """start_row+1 ~ start_row+n 행의 edit에 value 채우기"""
        for i in range(1, n + 1):
            target = start_row + i
            if target >= self._num_rows:
                break
            edit_list[target].setText(value)

    # ── CAS 조회 ───────────────────────────────────────────────

    def _lookup_cas(self, row):
        cas = self._cas_edits[row].text().strip()
        if not cas:
            return

        lbl = self._result_lbls[row]
        lbl.setText("조회 중 ...")
        lbl.setStyleSheet(_CELL_RES + f" color: {self._P.TEXT_SECONDARY};")

        worker = _CasWorker(row, cas)
        worker.result_ready.connect(self._on_cas_result)
        worker.lookup_error.connect(self._on_cas_error)
        worker.finished.connect(lambda w=worker: self._workers.discard(w))
        self._workers.add(worker)
        worker.start()

    def _on_cas_result(self, row, name, mw, density=None, smiles=""):
        """PubChem 결과 → reagent_lib 캐시 + self._rows + WebGrid 갱신. (row = 그리드 _id)"""
        print(f"[CAS Result] row={row}, name={name}, mw={mw}, density={density}, smiles={smiles}")
        r = self._rows[row] if 0 <= row < len(self._rows) else None
        cas = (getattr(r, "cas", "") or "") if r else ""
        # reagent_lib 캐시 (연구노트 Export 조회용)
        try:
            lib = getattr(self.app, "reagent_lib", None)
            if lib is None:
                lib = self.app.reagent_lib = {}
            if name:
                lib[name] = {"mw": mw if mw > 0 else 0.0,
                             "density": density if (density and density > 0) else None,
                             "smiles": smiles or "", "cas": cas}
        except Exception as _e:
            print(f"[CAS Result] reagent_lib 저장 경고(무시): {_e}")
        # _rows 갱신
        if r is not None:
            r.name = name
            r.mw = mw if mw > 0 else 0.0
            if density is not None and density > 0:
                r.density = density
            r.smiles = smiles or ""
        # 그리드 갱신 (부분)
        upd = {"_id": str(row), "name": name, "mw": mw if mw > 0 else None, "smiles": smiles or ""}
        if density is not None and density > 0:
            upd["density"] = density
        if getattr(self, "_grid", None):
            self._grid.set_rows([upd])

    def _on_cas_error(self, row, _msg):
        print(f"[CAS] 조회 실패 row={row}: {_msg}")

    def build_reagent_lib(self):
        """계산기 rows + CAS 캐시 → {name: {mw, density, smiles, cas}} (연구노트 Export용).
        수동입력 시약(CAS 미조회)도 name·mw·density 로 병합, SMILES 는 캐시 유지."""
        lib = getattr(self.app, "reagent_lib", None)
        if lib is None:
            lib = self.app.reagent_lib = {}
        for r in self._rows:
            nm = (getattr(r, "name", "") or "").strip()
            if not nm:
                continue
            cur = lib.get(nm, {})
            lib[nm] = {
                "mw": r.mw if getattr(r, "mw", 0) else cur.get("mw", 0.0),
                "density": r.density if getattr(r, "density", None) else cur.get("density"),
                "smiles": cur.get("smiles", ""),
                "cas": cur.get("cas", ""),
            }
        return lib

    # ── 시그널 핸들러 ──────────────────────────────────────────

    def _on_name(self, row, text):
        if self._building:
            return
        self._rows[row].name = text.strip()

    def _on_mw(self, row, text):
        if self._building:
            return
        v = _parse(text)
        self._rows[row].mw = v if v else 0.0

    def _on_field(self, row, field, text):
        if self._building:
            return
        setattr(self._rows[row], field, _parse(text))

    # ── 계산 ──────────────────────────────────────────────────

    def _recalc(self, row):
        if not self._result_lbls:   # 농도 테이블 WebGrid 대체 → Qt 결과 라벨 없음
            return
        r   = self._rows[row]
        mw  = r.mw
        ml, g, conc = r.ml, r.g, r.conc
        dens = r.density if (r.density is not None and r.density > 0) else None

        # @codesyncer-decision: 액상 시약일 때 "질량 X g" 옆에 "원액 Y mL" 를 병기.
        #   neat_mL = g / density (순수 원액 부피, 용액 부피와 다름).
        #   밀도가 비어있으면 기존 표기 그대로 유지 → 고체 시약 영향 없음.
        def _neat(gram):
            if dens is None or gram is None:
                return ""
            try:
                return f"  |  원액 = {gram / dens:.3f} mL"
            except ZeroDivisionError:
                return ""
        lbl = self._result_lbls[row]
        ni  = self._num_items[row]

        P = self._P
        _is_dark = getattr(self.app, 'is_dark_mode', True)
        _row_done = "#0d2010" if _is_dark else "#d6eedc"

        def _done(text):
            lbl.setText(text)
            lbl.setStyleSheet(
                _CELL_RES +
                f" font-weight: bold; color: {P.TEXT_PRIMARY}; background: {_row_done};"
            )
            ni.setBackground(QColor(_row_done))
            ni.setForeground(QColor(P.TEXT_PRIMARY))

        def _clear():
            lbl.setText("-")
            lbl.setStyleSheet(_CELL_RES)
            ni.setData(Qt.BackgroundRole, None)
            ni.setForeground(QColor(P.ACCENT_BLUE))

        def _error(text):
            lbl.setText(text)
            lbl.setStyleSheet(_CELL_RES + f" color: {P.ACCENT_RED};")
            ni.setData(Qt.BackgroundRole, None)
            ni.setForeground(QColor(P.ACCENT_BLUE))

        # @codesyncer-decision: 밀도 기반 계산 경로 추가
        # 기존: MW > 0 + (ml, g, conc) 중 2개 필요
        # 추가: 밀도가 있으면 ml↔g 상호 변환 가능 → 1개만 있어도 계산 가능
        #   - 밀도 + ml → g 자동 계산 (g = ml × density)
        #   - 밀도 + g → ml 자동 계산 (ml = g / density)
        #   이 변환 후 기존 2-입력 계산 로직에 진입

        # 사용자 입력값 기준으로 판단 (밀도 보충 전)
        orig_ml, orig_g, orig_conc = ml, g, conc
        filled_orig = sum(v is not None for v in (orig_ml, orig_g, orig_conc))

        # 밀도로 ml↔g 보충 (계산용 — 원본 구분 유지)
        if dens is not None:
            if ml is not None and g is None:
                g = ml * dens
            elif g is not None and ml is None:
                ml = g / dens

        filled = sum(v is not None for v in (ml, g, conc))

        # MW 없어도 밀도+ml 또는 밀도+g 만으로 변환 가능
        if mw <= 0 and dens is not None:
            if orig_ml is not None and orig_g is None:
                calc_g = orig_ml * dens
                _done(f"질량 = {calc_g:.4f} g  |  원액 = {orig_ml:.3f} mL  (밀도 기반)")
                r.eff_conc = None
                return
            elif orig_g is not None and orig_ml is None:
                calc_ml = orig_g / dens
                _done(f"원액 = {calc_ml:.3f} mL  |  질량 = {orig_g:.4f} g  (밀도 기반)")
                r.eff_conc = None
                return

        if mw <= 0 or filled < 2:
            r.eff_conc = conc
            _clear()
            return

        try:
            # 원래 사용자 입력 기준으로 분기 (밀도 보충값은 결과 표시용)
            if orig_ml is not None and orig_conc is not None and orig_g is None:
                calc = conc * mw * (ml / 1000.0)
                _done(f"질량 = {calc:.4f} g" + _neat(calc))
                r.eff_conc = conc

            elif orig_g is not None and orig_conc is not None and orig_ml is None:
                calc = (g / (conc * mw)) * 1000.0
                _done(f"용액 부피 = {calc:.2f} mL" + _neat(g))
                r.eff_conc = conc

            elif orig_ml is not None and orig_g is not None and orig_conc is None:
                calc = (g / mw) / (ml / 1000.0)
                _done(f"농도 = {calc:.4f} M" + _neat(g))
                r.eff_conc = calc

            # 밀도로 보충된 경우: 원래 1개만 입력 + 밀도로 2개 됨
            elif dens is not None and filled_orig == 1:
                if orig_ml is not None and orig_conc is None and orig_g is None:
                    # ml + 밀도 → g 보충됨 → 농도 계산
                    calc = (g / mw) / (ml / 1000.0)
                    _done(f"농도 = {calc:.4f} M  |  질량 = {g:.4f} g")
                    r.eff_conc = calc
                elif orig_g is not None and orig_conc is None and orig_ml is None:
                    # g + 밀도 → ml 보충됨 → 농도 계산
                    calc = (g / mw) / (ml / 1000.0)
                    _done(f"농도 = {calc:.4f} M  |  원액 = {ml:.3f} mL")
                    r.eff_conc = calc
                elif orig_conc is not None and orig_ml is None and orig_g is None:
                    # conc만 → 밀도로는 ml/g 보충 불가 → clear
                    r.eff_conc = conc
                    _clear()
                    return
                else:
                    r.eff_conc = conc
                    _done("✓  완료" + _neat(g))

            else:
                r.eff_conc = conc
                _done("✓  완료" + _neat(g))

        except ZeroDivisionError:
            r.eff_conc = None
            _error("0 입력 불가")

    # ── 엑셀 붙여넣기 분배 ────────────────────────────────────

    def _paste_cas(self, start_row, lines):
        # 하위 호환용 — CAS 단일 열 붙여넣기
        self._paste_grid(start_row, 0, "\n".join(lines))

    # ── 엑셀 그리드 (행×열) 붙여넣기 ──────────────────────────
    # @codesyncer-decision: 엑셀 TSV(\t=열, \n=행)를 포커스 셀 기준으로 오른쪽·아래로 분배.
    #   컬럼 순서는 UI 헤더와 동일하게 고정(0=CAS 1=시약명 2=MW 3=mL 4=g 5=M).
    #   순서가 다른 엑셀 시트는 사용자가 열을 맞춰 복사한다는 전제.
    # @codesyncer-inference: 엑셀은 복사 시 끝에 빈 줄(\r\n)을 붙이므로 trailing 빈 줄만 제거,
    #   중간 빈 셀(\t\t)은 "비움"으로 의도된 것으로 해석하여 보존한다.
    # @codesyncer-risk: Ctrl+V를 eventFilter에서 True로 소비하므로, 단일 셀 단순 붙여넣기까지
    #   가로채면 UX가 어색해질 수 있음 → 클립보드에 \t 또는 개행이 있을 때만 가로채고,
    #   그 외에는 QLineEdit 기본 동작에 위임.

    def _tag_paste(self, edit, row, col):
        """편집창에 (row, col)을 부여하고 그리드 붙여넣기 필터 설치"""
        edit.setProperty("grid_row", row)
        edit.setProperty("grid_col", col)
        edit.installEventFilter(self)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyPress:
            mods = event.modifiers()
            key = event.key()

            if mods & Qt.ControlModifier:
                # Ctrl+V: 그리드 붙여넣기
                if key == Qt.Key_V:
                    text = QApplication.clipboard().text()
                    if text and ("\t" in text or "\n" in text or "\r" in text):
                        # 방법 1: 셀 위젯에서 (grid_row, grid_col) 기준 — 농도 테이블 전용
                        row = obj.property("grid_row")
                        col = obj.property("grid_col")
                        if row is not None and col is not None:
                            self._save_undo_snapshot()
                            self._paste_grid(int(row), int(col), text, table=self._tbl)
                            return True
                        # 방법 2: 테이블에서 선택된 셀 기준
                        if obj in (self._tbl, getattr(self, '_dil_tbl', None)):
                            ranges = obj.selectedRanges()
                            if ranges:
                                sr = ranges[0]
                                # 테이블 행 → 데이터 행 변환
                                dr = self._tbl_row_to_data_row(sr.topRow())
                                if dr is not None:
                                    # 테이블 컬럼 → 그리드 컬럼 (col 0=# → skip)
                                    gc = max(0, sr.leftColumn() - 1)
                                    self._save_undo_snapshot()
                                    self._paste_grid(dr, gc, text, table=obj)
                                    return True

                # Ctrl+C: 테이블 선택 영역 복사
                if key == Qt.Key_C:
                    if obj in (self._tbl, getattr(self, '_dil_tbl', None)):
                        self._copy_table_selection(obj)
                        return True

                # Ctrl+Z: 되돌리기
                if key == Qt.Key_Z:
                    self._undo()
                    return True

                # Ctrl+A: 전체 선택
                if key == Qt.Key_A:
                    if obj in (self._tbl, getattr(self, '_dil_tbl', None)):
                        obj.selectAll()
                        return True

            # Delete: 선택 영역 지우기
            if key == Qt.Key_Delete:
                if obj in (self._tbl, getattr(self, '_dil_tbl', None)):
                    self._save_undo_snapshot()
                    self._delete_selection(obj)
                    return True

        return super().eventFilter(obj, event)

    def _copy_table_selection(self, table):
        """QTableWidget 선택 영역을 TSV로 클립보드에 복사 (헤더 포함)"""
        ranges = table.selectedRanges()
        if not ranges:
            return
        sr = ranges[0]
        lines = []

        # 헤더 행 추가
        header_row = []
        for c in range(sr.leftColumn(), sr.rightColumn() + 1):
            h_item = table.horizontalHeaderItem(c)
            header_row.append(h_item.text().replace("\n", " ") if h_item else "")
        lines.append("\t".join(header_row))

        # 데이터 행
        for r in range(sr.topRow(), sr.bottomRow() + 1):
            row_data = []
            for c in range(sr.leftColumn(), sr.rightColumn() + 1):
                w = table.cellWidget(r, c)
                if w:
                    text = self._extract_widget_text(w)
                else:
                    item = table.item(r, c)
                    text = item.text() if item else ""
                row_data.append(text)
            lines.append("\t".join(row_data))
        QApplication.clipboard().setText("\n".join(lines))

    def _extract_widget_text(self, widget):
        """셀 위젯에서 텍스트 추출 (QLineEdit, QLabel, 복합 위젯 등)"""
        # QLineEdit 직접
        if isinstance(widget, QLineEdit):
            return widget.text()
        # QLabel 직접
        if isinstance(widget, QLabel):
            return widget.text()
        # 복합 위젯: 내부 QLineEdit 또는 QLabel 탐색
        edit = widget.findChild(QLineEdit)
        if edit:
            return edit.text()
        lbl = widget.findChild(QLabel)
        if lbl:
            return lbl.text()
        return ""

    def _copy_all_conc(self):
        """농도 계산기 전체 데이터를 TSV로 클립보드에 복사"""
        headers = ["#", "CAS", "시약명", "MW(g/mol)", "밀도(g/mL)",
                   "부피(mL)", "질량(g)", "농도(M)", "계산 결과"]
        lines = ["\t".join(headers)]
        for dr in range(self._num_rows):
            gi = dr // ROWS_PER_GROUP
            pos = dr % ROWS_PER_GROUP
            letter = chr(65 + gi) if gi < 26 else f"G{gi}"
            row = [
                f"{letter}-{pos+1}",
                self._cas_edits[dr].text(),
                self._name_edits[dr].text(),
                self._mw_edits[dr].text(),
                self._d_edits[dr].text(),
                self._ml_edits[dr].text(),
                self._g_edits[dr].text(),
                self._m_edits[dr].text(),
                self._result_lbls[dr].text(),
            ]
            lines.append("\t".join(row))
        QApplication.clipboard().setText("\n".join(lines))

    def _grid_copy_all(self):
        """WebGrid 농도 계산기 전체 데이터를 TSV로 클립보드에 복사 (엑셀 붙여넣기용)"""
        headers = ["Group", "Port", "시약명", "CAS", "MW(g/mol)", "SMILES",
                   "밀도(g/mL)", "부피(mL)", "질량(g)", "농도(M)"]
        lines = ["\t".join(headers)]
        rows = self._grid.rows() if self._grid is not None else []
        for r in rows:
            vals = [r.get("grp", ""), r.get("port", ""), r.get("name", ""),
                    r.get("cas", ""), r.get("mw", ""), r.get("smiles", ""),
                    r.get("density", ""), r.get("vol", ""), r.get("mass", ""),
                    r.get("conc", "")]
            lines.append("\t".join("" if v is None else str(v) for v in vals))
        QApplication.clipboard().setText("\n".join(lines))

    def _copy_all_dil(self):
        """희석 그리드 전체를 TSV로 클립보드에 복사 (엑셀 붙여넣기용)."""
        if getattr(self, "_dil_grid", None) is None:
            return
        headers = ["#", "시약명", "원액 C1(M)", "목표 C2(M)", "총 부피(mL)",
                   "원액량(mL)", "용매량(mL)", "상태"]
        st_txt = {"ok": "완료", "short": "원액 부족"}

        def s(v):
            return "" if v in (None, "") else str(v)

        lines = ["\t".join(headers)]
        for gd in self._dil_grid.rows():
            try:
                dr = int(gd.get("_id"))
            except (TypeError, ValueError):
                continue
            gi, pos = dr // ROWS_PER_GROUP, dr % ROWS_PER_GROUP
            letter = chr(65 + gi) if gi < 26 else f"G{gi}"
            lines.append("\t".join([
                f"{letter}-{pos + 1}", s(gd.get("name")), s(gd.get("c1")),
                s(gd.get("c2")), s(gd.get("v2")), s(gd.get("v1")),
                s(gd.get("solv")), st_txt.get(gd.get("status") or "", "")]))
        QApplication.clipboard().setText("\n".join(lines))

    # ── Undo / Delete ──────────────────────────────────────────

    def _save_undo_snapshot(self):
        """현재 상태를 undo 스택에 저장 (최대 20개)"""
        # @codesyncer: 농도 테이블이 WebGrid 로 대체되어 Qt 셀 리스트가 비어 있음
        #   → 스냅샷 대상 없음(희석 테이블 Ctrl+V/Delete 시 IndexError 방지).
        if not self._cas_edits:
            return
        snapshot = []
        for row in range(self._num_rows):
            snapshot.append({
                'cas':     self._cas_edits[row].text(),
                'name':    self._name_edits[row].text(),
                'mw':      self._mw_edits[row].text(),
                'density': self._d_edits[row].text(),
                'ml':      self._ml_edits[row].text(),
                'g':       self._g_edits[row].text(),
                'conc':    self._m_edits[row].text(),
            })
        self._undo_stack.append(snapshot)
        if len(self._undo_stack) > 20:
            self._undo_stack.pop(0)

    def _undo(self):
        """Ctrl+Z: 마지막 스냅샷으로 복원"""
        if not self._undo_stack or not self._cas_edits:
            return
        snapshot = self._undo_stack.pop()
        self._building = True
        try:
            for row, data in enumerate(snapshot):
                if row >= self._num_rows:
                    break
                for key, edit_list in (
                    ('cas', self._cas_edits), ('name', self._name_edits),
                    ('mw', self._mw_edits), ('density', self._d_edits),
                    ('ml', self._ml_edits), ('g', self._g_edits),
                    ('conc', self._m_edits),
                ):
                    edit_list[row].blockSignals(True)
                    edit_list[row].setText(data[key])
                    edit_list[row].blockSignals(False)
                # _Row 동기화
                r = self._rows[row]
                r.name = data['name'].strip()
                mw_v = _parse(data['mw'])
                r.mw = mw_v if mw_v else 0.0
                r.density = _parse(data['density'])
                r.ml = _parse(data['ml'])
                r.g = _parse(data['g'])
                r.conc = _parse(data['conc'])
        finally:
            self._building = False
        # 전체 재계산
        for row in range(self._num_rows):
            self._recalc(row)

    def _delete_selection(self, table):
        """Delete 키: 테이블 선택 영역의 입력 셀 지우기"""
        ranges = table.selectedRanges()
        if not ranges:
            return
        # 농도 테이블인지 희석 테이블인지에 따라 스토어 결정
        if table is self._tbl:
            stores = [None, self._cas_edits, self._name_edits, self._mw_edits,
                      self._d_edits, self._ml_edits, self._g_edits, self._m_edits, None]
            is_dilution = False
        elif table is getattr(self, "_dil_tbl", None):
            # 희석 테이블 편집 가능 컬럼: 1=이름, 2=원액농도, 3=목표농도, 4=목표부피
            stores = [None, self._dil_name_lbls, self._dil_stock_lbls,
                      self._dil_target_eds, self._dil_totalv_eds, None, None]
            is_dilution = True
        else:
            return

        sr = ranges[0]
        self._building = True
        try:
            for r in range(sr.topRow(), sr.bottomRow() + 1):
                # separator 행 무시
                dr = self._tbl_row_to_data_row(r)
                if dr is None or dr < 0 or dr >= self._num_rows:
                    continue
                for c in range(sr.leftColumn(), sr.rightColumn() + 1):
                    if c < len(stores) and stores[c] is not None:
                        edit = stores[c][dr]
                        edit.blockSignals(True)
                        edit.clear()
                        edit.blockSignals(False)
                # 농도 테이블만 _Row 동기화
                if not is_dilution:
                    row = dr
                    r_obj = self._rows[row]
                    r_obj.name = self._name_edits[row].text().strip()
                    mw_v = _parse(self._mw_edits[row].text())
                    r_obj.mw = mw_v if mw_v else 0.0
                    r_obj.density = _parse(self._d_edits[row].text())
                    r_obj.ml = _parse(self._ml_edits[row].text())
                    r_obj.g = _parse(self._g_edits[row].text())
                    r_obj.conc = _parse(self._m_edits[row].text())
        finally:
            self._building = False
        # 영향 받은 행 재계산
        for r in range(sr.topRow(), sr.bottomRow() + 1):
            dr = self._tbl_row_to_data_row(r)
            if dr is not None and 0 <= dr < self._num_rows:
                if is_dilution:
                    self._dil_calc_row(dr)
                else:
                    self._recalc(dr)

    def _tbl_row_to_data_row(self, tbl_row):
        """테이블 행 번호 → 데이터 행 인덱스 (separator 행이면 None)"""
        group_size = ROWS_PER_GROUP + 1  # 데이터 행 + separator
        group_idx = tbl_row // group_size
        pos_in_group = tbl_row % group_size
        if pos_in_group == 0:
            return None  # separator 행
        return group_idx * ROWS_PER_GROUP + (pos_in_group - 1)

    def _paste_grid(self, start_row, start_col, text, table=None):
        """엑셀 클립보드(TSV)를 (start_row, start_col)부터 분배.

        Args:
            start_row: 시작 데이터 행 (0-based, separator 제외)
            start_col: 시작 그리드 열 (0-based, # 컬럼 제외)
            text: 클립보드 TSV 텍스트
            table: 대상 테이블 — self._tbl(농도) 또는 self._dil_tbl(희석). None이면 농도.
        """
        is_dilution = table is getattr(self, "_dil_tbl", None) and table is not None

        if is_dilution:
            # 희석 테이블 store 매핑 (# 컬럼 제외, 1-based 컬럼을 0-based grid로 -1 shift)
            #   tbl col 1 → gc 0: 이름 (name)
            #   tbl col 2 → gc 1: 원액 농도 (stock)
            #   tbl col 3 → gc 2: 목표 농도 (target conc)
            #   tbl col 4 → gc 3: 목표 부피 (target vol)
            #   tbl col 5 → gc 4: V1 (읽기전용)
            #   tbl col 6 → gc 5: 용매량 (읽기전용)
            stores = [
                self._dil_name_lbls,    # 0
                self._dil_stock_lbls,   # 1
                self._dil_target_eds,   # 2
                self._dil_totalv_eds,   # 3
                None,                   # 4: V1 읽기전용
                None,                   # 5: solvent 읽기전용
            ]
            numeric_cols = {1, 2, 3}   # stock, target_conc, target_vol
        else:
            stores = [
                self._cas_edits,    # 0
                self._name_edits,   # 1
                self._mw_edits,     # 2
                self._d_edits,      # 3  밀도 g/mL
                self._ml_edits,     # 4
                self._g_edits,      # 5
                self._m_edits,      # 6
            ]
            numeric_cols = {2, 3, 4, 5, 6}   # MW, density, vol, mass, conc

        # 줄 끝 정규화 — Excel은 보통 마지막에 \r\n 한 줄을 추가
        raw_lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        while raw_lines and raw_lines[-1] == "":
            raw_lines.pop()
        if not raw_lines:
            return

        # 1차 검증: 숫자 컬럼에 숫자 아닌 값이 있으면 에러창
        for i, line in enumerate(raw_lines):
            row = start_row + i
            if row >= self._num_rows:
                break  # 범위 초과 → 잘라내기
            cells = line.split("\t")
            for j, val in enumerate(cells):
                col = start_col + j
                if col >= len(stores):
                    break
                if col not in numeric_cols:
                    continue
                v = val.strip()
                if v == "":
                    continue
                try:
                    float(v.replace(",", "."))
                except ValueError:
                    QMessageBox.warning(
                        self, "붙여넣기 오류",
                        f"숫자 컬럼에는 숫자만 기입해 주세요.\n"
                        f"잘못된 값: '{v}' (행 {row + 1}, 열 {col + 1})"
                    )
                    return

        self._building = True
        try:
            for i, line in enumerate(raw_lines):
                row = start_row + i
                if row >= self._num_rows:
                    break
                cells = line.split("\t")
                for j, val in enumerate(cells):
                    col = start_col + j
                    if col >= len(stores):
                        break
                    store = stores[col]
                    if store is None:
                        continue  # 읽기전용 컬럼 스킵
                    edit = store[row]
                    edit.blockSignals(True)
                    edit.setText(val.strip())
                    edit.blockSignals(False)
        finally:
            self._building = False

        # 후처리: 영향 받은 행 재계산 / 저장
        for i in range(len(raw_lines)):
            row = start_row + i
            if row >= self._num_rows:
                break
            if is_dilution:
                # 희석 테이블: 해당 행만 재계산
                self._dil_calc_row(row)
            else:
                # 농도 테이블: CAS 조회 + _Row 동기화 + 재계산
                if start_col == 0 and self._cas_edits[row].text().strip():
                    self._lookup_cas(row)
                r = self._rows[row]
                r.name = self._name_edits[row].text().strip()
                mw_v   = _parse(self._mw_edits[row].text())
                r.mw   = mw_v if mw_v else 0.0
                r.density = _parse(self._d_edits[row].text())
                r.ml   = _parse(self._ml_edits[row].text())
                r.g    = _parse(self._g_edits[row].text())
                r.conc = _parse(self._m_edits[row].text())
                self._recalc(row)

    # ── 일괄 CAS 조회 ─────────────────────────────────────────

    def _bulk_lookup(self):
        """WebGrid 의 CAS 입력 행들을 PubChem 으로 일괄 조회 → 결과를 그리드에 반영."""
        if not getattr(self, "_grid", None):
            return
        for gd in self._grid.rows():
            cas = (gd.get("cas") or "").strip()
            if not cas or (gd.get("mw") and gd.get("smiles")):
                continue
            try:
                self._start_cas_lookup(int(gd.get("_id")), cas)
            except (TypeError, ValueError):
                continue

    # ── 일괄 계산 ────────────────────────────────────────────

    def _bulk_recalc(self):
        # 위젯 → _Row 강제 동기화 후 재계산
        for row in range(self._num_rows):
            r = self._rows[row]
            r.name = self._name_edits[row].text().strip()
            mw_v = _parse(self._mw_edits[row].text())
            r.mw = mw_v if mw_v else 0.0
            r.density = _parse(self._d_edits[row].text())
            r.ml = _parse(self._ml_edits[row].text())
            r.g = _parse(self._g_edits[row].text())
            r.conc = _parse(self._m_edits[row].text())
            self._recalc(row)

    # ── 초기화 ────────────────────────────────────────────────

    def _clear_all(self):
        self._building = True
        for row in range(self._num_rows):
            for w in (self._cas_edits[row], self._name_edits[row],
                      self._mw_edits[row], self._d_edits[row],
                      self._ml_edits[row], self._g_edits[row],
                      self._m_edits[row]):
                w.blockSignals(True)
                w.clear()
                w.blockSignals(False)

            self._result_lbls[row].setText("-")
            self._result_lbls[row].setStyleSheet(_CELL_RES)

            ni = self._num_items[row]
            ni.setData(Qt.BackgroundRole, None)
            ni.setForeground(QColor(self._P.ACCENT_BLUE))

            self._rows[row] = _Row()
        self._building = False

    # ── Import ────────────────────────────────────────────────

    def _open_import(self):
        valid = [
            (dr, r) for dr, r in enumerate(self._rows)
            if r.name and r.eff_conc is not None and r.eff_conc > 0
        ]
        if not valid:
            QMessageBox.warning(self, "Import 불가",
                                "이름과 농도(M)가 설정된 행이 없습니다.")
            return
        _ImportDialog(valid, self._group_info, self.app, parent=self,
                      port_children=getattr(self, "_port_children", {})).exec_()

    # ─── Theme Support ─────────────────────────────────────────
    # 각 탭별 독립 메서드로 분리 → apply_theme 에서 일괄 호출

    def apply_theme(self, is_dark=True):
        """테마 전환 시 Calculator 탭 색상 갱신 (탭별 위임)"""
        P = Dark if is_dark else Light
        self._apply_table_styles(P, is_dark)
        self._apply_conc_styles(P)
        self._apply_dil_styles(P, is_dark)
        self._apply_rtd_styles(P, is_dark)
        # WebGrid(Tabulator) 농도/희석 테이블 팔레트 재주입
        if getattr(self, '_grid', None) is not None:
            self._grid.set_theme(self._grid_theme(P, is_dark))
        if getattr(self, '_dil_grid', None) is not None:
            self._dil_grid.set_theme(self._grid_theme(P, is_dark))

    # ── 공통 (농도 · 희석 테이블 공유) ──

    def _apply_table_styles(self, P, is_dark):
        """그룹 구분행, 행 번호, 테이블 헤더 배경 — 농도/희석 공통"""
        grp_bg = "#1e252e" if is_dark else "#e8edf4"
        grp_fg = "#e0e0e0" if is_dark else "#344054"
        for lbl in self._grp_sep_labels:
            lbl.setStyleSheet(
                f"font-weight:bold;font-size:11px;"
                f"color:{grp_fg};background:{grp_bg};padding:0 6px;"
            )
        for ni in self._num_items:
            ni.setForeground(QColor(P.ACCENT_BLUE))
        for ni in self._dil_num_items:
            ni.setForeground(QColor(grp_fg))

        hdr_in = "#1e3a5f" if is_dark else "#dce6f5"
        hdr_out = "#1a3a20" if is_dark else "#d6eedc"
        _IN_SET = {"#1e3a5f", "#dce6f5"}
        _OUT_SET = {"#1a3a20", "#d6eedc"}
        for tbl in (getattr(self, '_tbl', None), getattr(self, '_dil_tbl', None)):
            if not tbl:
                continue
            for col in range(tbl.columnCount()):
                item = tbl.horizontalHeaderItem(col)
                if not item:
                    continue
                bg = item.background().color().name()
                if bg in _IN_SET:
                    item.setBackground(QColor(hdr_in))
                elif bg in _OUT_SET:
                    item.setBackground(QColor(hdr_out))

    # ── 농도 계산기 ──

    def _apply_conc_styles(self, P):
        if self._conc_sub_lbl:
            self._conc_sub_lbl.setStyleSheet(
                f"color:{P.TEXT_SECONDARY};font-size:12px;"
            )
        if self._conc_btn_imp:
            self._conc_btn_imp.setStyleSheet(get_primary_button_style(P))
        _sec = get_secondary_button_style(P)
        for btn in self._conc_btns_secondary:
            btn.setStyleSheet(_sec)

        # 결과 라벨 테마 갱신 — 구 QTableWidget 경로에서만.
        # WebGrid 는 JS 가 자체 재계산하고 apply_theme 에서 팔레트를 재주입한다.
        if getattr(self, '_tbl', None) is not None:
            self._bulk_recalc()

    # ── 희석 계산기 ──

    def _apply_dil_styles(self, P, is_dark):
        if getattr(self, '_dil_title_lbl', None):
            self._dil_title_lbl.setStyleSheet(
                f"color:{P.TEXT_PRIMARY};font-size:13px;font-weight:{T.FW_SEMI};"
            )
        if self._dil_sub_lbl:
            self._dil_sub_lbl.setStyleSheet(
                f"color:{P.TEXT_SECONDARY};font-size:12px;"
            )
        for lbl in self._dil_ctrl_labels:
            lbl.setStyleSheet(f"font-size:12px;color:{P.TEXT_SECONDARY};")

        # 전역 입력 LineEdit
        _line_ss = (
            f"QLineEdit{{background:{P.BG_INPUT};color:{P.TEXT_INPUT};"
            f"border:1px solid {P.BORDER_INPUT};border-radius:3px;"
            f"padding:2px 6px;font-size:12px;}}"
            f"QLineEdit:focus{{border-color:{P.ACCENT_BLUE};}}"
        )
        for ed in (getattr(self, '_dil_g_conc', None),
                    getattr(self, '_dil_g_vol', None)):
            if ed:
                ed.setStyleSheet(_line_ss)

        # 입력 셀 (목표농도 · 총부피)
        cell_bg = "#1a2540" if is_dark else "#e8eef6"
        cell_bd = "#3a5a9a" if is_dark else "#7aa0d4"
        _cell_ss = (
            f"font-family:Consolas,monospace;font-size:{_CELL_FS};"
            f"background:{cell_bg};border:none;"
            f"border-bottom:1px solid {cell_bd};padding:0 6px;"
        )
        for ed in getattr(self, '_dil_target_eds', []):
            ed.setStyleSheet(_cell_ss)
        for ed in getattr(self, '_dil_totalv_eds', []):
            ed.setStyleSheet(_cell_ss)

        _sec = get_secondary_button_style(P)
        for btn in self._dil_btns_secondary:
            btn.setStyleSheet(_sec)

        # 희석 결과 라벨 테마 갱신
        self._dil_calc_all()

    # ── RTD 계산기 ──

    def _apply_rtd_styles(self, P, is_dark):
        # 스플리터
        if self._rtd_splitter:
            self._rtd_splitter.setStyleSheet(
                f"QSplitter::handle{{background:{P.BORDER_PRIMARY};width:1px;}}"
            )

        # 입력 프레임 + 섹션 라벨
        if self._rtd_inp_frame:
            self._rtd_inp_frame.setStyleSheet(
                f"QFrame{{background:{P.BG_SECONDARY};"
                f"border:1px solid {P.BORDER_PRIMARY};border-radius:6px;}}"
            )
        if self._rtd_sec_lbl:
            self._rtd_sec_lbl.setStyleSheet(
                f"font-size:11px;font-weight:bold;color:{P.TEXT_SECONDARY};"
                f"background:transparent;border:none;letter-spacing:0.5px;"
            )

        # 행 라벨 (반응기 내경, 길이 등)
        for lbl in self._rtd_rlabels:
            lbl.setStyleSheet(
                f"font-size:12px;color:{P.TEXT_SECONDARY};padding-right:6px;"
            )

        # 스핀박스 · 콤보 공통 스타일
        _spin = (
            f"QDoubleSpinBox{{background:{P.BG_INPUT};color:{P.TEXT_INPUT};"
            f"border:1px solid {P.BORDER_INPUT};border-radius:3px;"
            f"padding:2px 6px;font-size:12px;}}"
            f"QDoubleSpinBox:focus{{border-color:{P.ACCENT_BLUE};}}"
        )
        _combo = (
            f"QComboBox{{background:{P.BG_INPUT};color:{P.TEXT_INPUT};"
            f"border:1px solid {P.BORDER_INPUT};border-radius:3px;"
            f"padding:2px 6px;font-size:12px;}}"
            f"QComboBox::drop-down{{border:none;}}"
            f"QComboBox QAbstractItemView{{background:{P.BG_SECONDARY};"
            f"color:{P.TEXT_PRIMARY};selection-background-color:{P.BG_HOVER};}}"
        )
        for name in ('id', 'len', 'flow', 'temp'):
            sp = getattr(self, f'_rtd_sp_{name}', None)
            if sp:
                sp.setStyleSheet(_spin)
        if getattr(self, '_rtd_cb_solv', None):
            self._rtd_cb_solv.setStyleSheet(_combo)

        # 계산 버튼 · 구분선
        if self._rtd_btn_calc:
            self._rtd_btn_calc.setStyleSheet(
                f"QPushButton{{background:{P.ACCENT_BLUE};color:white;"
                f"font-size:13px;font-weight:bold;border-radius:4px;border:none;}}"
                f"QPushButton:hover{{background:{P.ACCENT_BLUE_DARK};}}"
            )
        if self._rtd_sep:
            self._rtd_sep.setStyleSheet(f"color:{P.BORDER_PRIMARY};")

        # 결과 카드 (프레임 + 타이틀 + 값)
        _card = (
            f"QFrame{{background:{P.HEADER_BG};"
            f"border:1px solid {P.BORDER_SECONDARY};border-radius:5px;}}"
        )
        _title = (
            f"color:{P.TEXT_SECONDARY};font-size:12px;"
            f"background:transparent;border:none;"
        )
        _value = (
            f"color:{P.TEXT_PRIMARY};font-size:16px;font-weight:bold;"
            f"background:transparent;border:none;"
        )
        for card in self._rtd_card_frames:
            card.setStyleSheet(_card)
        for t in self._rtd_card_titles:
            t.setStyleSheet(_title)
        for v in self._rtd_cards.values():
            if v.text() == "—":
                v.setStyleSheet(_value)

        # 힌트 · 해석 · 플롯
        if self._rtd_hint_lbl:
            self._rtd_hint_lbl.setStyleSheet(
                f"color:{P.TEXT_DISABLED};font-size:11px;"
            )
        if hasattr(self, '_rtd_interp'):
            self._rtd_interp.setStyleSheet(
                f"background:{P.BG_TERTIARY};border:1px solid {P.BORDER_PRIMARY};"
                f"border-radius:5px;padding:12px;font-size:13px;color:{P.TEXT_PRIMARY};"
            )
        if hasattr(self, '_rtd_plot'):
            plot_bg = P.BG_MONITOR if is_dark else P.BG_SECONDARY
            self._rtd_plot.setBackground(plot_bg)
            self._rtd_plot.setLabel("bottom", "시간 (min)",
                                    color=P.TEXT_SECONDARY, size="10pt")
            self._rtd_plot.setLabel("left", "E(t)",
                                    color=P.TEXT_SECONDARY, size="10pt")
            for ax in ("left", "bottom"):
                self._rtd_plot.getAxis(ax).setTextPen(P.TEXT_SECONDARY)
                self._rtd_plot.getAxis(ax).setPen(
                    pg.mkPen(P.BORDER_LIGHT if is_dark else P.BORDER_PRIMARY, width=2))


# ── Import 다이얼로그 (자동 배정 확인) ─────────────────────────────

class _ImportDialog(QDialog):
    def __init__(self, valid_rows, group_info, app, parent=None, port_children=None):
        super().__init__(parent)
        self.app        = app
        self.valid_rows = valid_rows   # [(data_idx, _Row), ...]
        self.group_info = group_info   # [(letter, pump), ...]
        # 다성분: {dr: [자식 시약 dict]} — Import 시 (pump, port) 레시피로 번들
        self.port_children = dict(port_children or {})
        self.setWindowTitle("시퀀스 자동 Import")
        self.setMinimumWidth(560)
        self._setup_ui()

    def _auto_assign(self, dr):
        """data row index → (pump, port) 자동 계산"""
        gi   = dr // ROWS_PER_GROUP
        port = dr % ROWS_PER_GROUP + 2
        pump = self.group_info[gi][1] if gi < len(self.group_info) else ""
        return pump, port

    def _setup_ui(self):
        P = Dark if getattr(self.app, 'is_dark_mode', True) else Light
        is_dark = getattr(self.app, 'is_dark_mode', True)
        grp_fg = "#e0e0e0" if is_dark else "#344054"

        root = QVBoxLayout(self)
        root.setSpacing(10)

        lbl = QLabel(
            "행 위치에 따라 그룹·포트가 자동 배정됩니다.  "
            "아래 목록을 확인 후 Import를 실행하세요."
        )
        lbl.setWordWrap(True)
        lbl.setStyleSheet("font-size: 12px;")
        root.addWidget(lbl)

        grid = QGridLayout()
        grid.setSpacing(6)
        for col, h in enumerate(["행", "시약명", "농도", "→  그룹 / 펌프", "포트"]):
            hl = QLabel(h)
            hl.setStyleSheet("font-weight: bold; font-size: 12px;")
            grid.addWidget(hl, 0, col)

        for i, (dr, r) in enumerate(self.valid_rows):
            gi     = dr // ROWS_PER_GROUP
            letter = self.group_info[gi][0] if gi < len(self.group_info) else "?"
            pump, port = self._auto_assign(dr)

            row_lbl = QLabel(f"{letter}-{dr % ROWS_PER_GROUP + 1}")
            row_lbl.setStyleSheet(f"font-weight: bold; color: {grp_fg}; font-size: 12px;")
            grid.addWidget(row_lbl, i + 1, 0)
            _n_kids = len([c for c in self.port_children.get(dr, [])
                           if (c.get("name") or "").strip()])
            _name_txt = r.name + (f"  (혼합 ×{_n_kids + 1})" if _n_kids else "")
            grid.addWidget(QLabel(_name_txt), i + 1, 1)
            grid.addWidget(QLabel(f"{r.eff_conc:.4f} M"), i + 1, 2)

            grp_lbl = QLabel(f"Group {letter}  ·  {pump}")
            grp_lbl.setStyleSheet(f"color: {grp_fg}; font-size: 12px;")
            grid.addWidget(grp_lbl, i + 1, 3)
            grid.addWidget(QLabel(f"Port {port}"), i + 1, 4)

        root.addLayout(grid)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        root.addWidget(sep)

        btn_bar = QHBoxLayout()
        btn_cancel = QPushButton("취소")
        btn_cancel.clicked.connect(self.reject)
        btn_ok = QPushButton("Import 실행")
        btn_ok.setStyleSheet(
            f"background: {P.ACCENT_BLUE}; color: white; "
            f"font-weight: bold; padding: 4px 16px;"
        )
        btn_ok.clicked.connect(self._do_import)
        btn_bar.addWidget(btn_cancel)
        btn_bar.addStretch()
        btn_bar.addWidget(btn_ok)
        root.addLayout(btn_bar)

    def _do_import(self):
        map_mgr = self.app.map_mgr
        if not map_mgr:
            QMessageBox.warning(self, "오류", "Map Manager가 초기화되지 않았습니다.")
            return

        seq      = getattr(self.app, 'seq_tab', None)
        imported = []

        for dr, r in self.valid_rows:
            pump, port = self._auto_assign(dr)
            if not pump:
                continue
            conc_val = round(r.eff_conc, 4)
            map_mgr.update_inlet(pump, port, r.name, conc_val)
            imported.append((pump, port, r.name, conc_val))

        if seq:
            for pump, port, name, conc in imported:
                tbl = seq.reagent_tables.get(pump)
                if tbl:
                    tbl_row = port - 1
                    tbl.blockSignals(True)
                    if tbl.item(tbl_row, 1):
                        tbl.item(tbl_row, 1).setText(name)
                    if tbl.item(tbl_row, 2):
                        tbl.item(tbl_row, 2).setText(str(conc))
                    tbl.blockSignals(False)
            seq.refresh_port_combos()

        # ── 다성분 번들 → 시퀀스탭 레시피 (연구노트 export 까지 연동) ──
        n_mix = 0
        if seq and hasattr(seq, "_apply_stock_recipe"):
            from engine.stock_stoich import compute_stock
            inlet = list(getattr(self.app.cfg, "INLET_PUMPS", []) or [])
            for dr, r in self.valid_rows:
                pump, port = self._auto_assign(dr)
                if not pump or pump not in inlet:
                    continue
                gi_seq = inlet.index(pump)
                kids = [c for c in self.port_children.get(dr, [])
                        if (c.get("name") or "").strip()]
                if kids:
                    comps = [{"reagent": r.name, "molarity": round(r.eff_conc, 4),
                              "smiles": getattr(r, "smiles", "") or "",
                              "mw": r.mw or 0.0, "density": r.density or 0.0,
                              "limiting": True}]
                    for c in kids:
                        comps.append({"reagent": c.get("name", ""),
                                      "molarity": _parse(str(c.get("conc") or "")) or 0.0,
                                      "smiles": c.get("smiles", "") or "",
                                      "mw": c.get("mw") or 0.0,
                                      "density": c.get("density") or 0.0})
                    recipe = {"name": "", "total_volume_ml": r.ml or 0.0,
                              "solvents": [], "components": comps}
                    try:
                        seq._apply_stock_recipe(pump, gi_seq, port, compute_stock(recipe))
                        n_mix += 1
                    except Exception as e:
                        print(f"[Import→레시피] {pump} P{port}: {e}")
                elif (pump, port) in getattr(seq, "stock_recipes", {}):
                    # 단일시약으로 재배정 → 기존 레시피 해제 (Import = 계산기 상태 반영)
                    try:
                        seq._apply_stock_recipe(pump, gi_seq, port, {})
                    except Exception:
                        pass

        msg = f"{len(imported)}개 시약이 자동 배정되었습니다."
        if n_mix:
            msg += f"\n(혼합 stock {n_mix}개 포트는 레시피로 함께 배정 — 연구노트에 성분별 반영)"
        QMessageBox.information(self, "완료", msg)
        self.accept()

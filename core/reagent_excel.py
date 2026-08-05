"""
Reagent & Sequence Excel Manager - Excel 기반 시약/시퀀스 편집/감시/자동반영
@codesyncer-decision: main.py에서 분리 - Excel 시약+시퀀스 관리 단일 책임

사용법:
    excel_mgr = ReagentExcelManager(app)
    btn.clicked.connect(excel_mgr.open_reagents_excel)
"""

import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem
from PyQt5.QtCore import Qt, QTimer, QFileSystemWatcher


class ReagentExcelManager:
    """Excel 파일을 통한 시약 설정 및 시퀀스 편집/자동 반영을 담당합니다."""

    def __init__(self, app):
        self.app = app
        self._reagent_excel_path = None
        self._excel_watcher = QFileSystemWatcher()
        self._excel_watcher.fileChanged.connect(self._on_excel_file_changed)

    # ─── Excel 스타일 헬퍼 ─────────────────────────────────────

    @staticmethod
    def _excel_styles():
        """공용 Excel 셀 스타일 반환"""
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        return header_font, header_fill, center_align, thin_border

    # ─── Excel 열기 (시약 + 시퀀스 2시트) ──────────────────────

    def open_reagents_excel(self):
        """시약 설정 + 시퀀스 설정을 Excel 파일로 열어 편집 (저장 시 자동 반영)
        @codesyncer-inference: os.startfile()은 Windows 전용 API
          - macOS/Linux 지원 시 subprocess.Popen(['xdg-open', path]) 등으로 대체 필요"""
        if not HAS_OPENPYXL:
            QMessageBox.warning(self.app, "경고",
                "openpyxl 패키지가 설치되지 않았습니다.\npip install openpyxl 명령으로 설치하세요.")
            return

        try:
            temp_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "temp")
            os.makedirs(temp_dir, exist_ok=True)
            self._reagent_excel_path = os.path.join(temp_dir, "reagent_settings.xlsx")

            if os.path.exists(self._reagent_excel_path):
                try:
                    with open(self._reagent_excel_path, 'a'):
                        pass
                except PermissionError:
                    QMessageBox.information(self.app, "알림",
                        "Excel 파일이 이미 열려 있습니다.\n열린 Excel 창에서 편집하세요.")
                    return

            if self._excel_watcher.files():
                self._excel_watcher.removePaths(self._excel_watcher.files())

            wb = openpyxl.Workbook()
            header_font, header_fill, center_align, thin_border = self._excel_styles()

            pumps = [p for p in self.app.reagent_tables.keys() if p in self.app.cfg.INLET_PUMPS]

            # ── Sheet 1: 시약 설정 ──
            ws1 = wb.active
            ws1.title = "시약 설정"
            self._write_reagent_sheet(ws1, pumps, header_font, header_fill, center_align, thin_border)

            # ── Sheet 2: 시퀀스 설정 ──
            ws2 = wb.create_sheet("시퀀스 설정")
            self._write_sequence_sheet(ws2, pumps, header_font, header_fill, center_align, thin_border)

            wb.save(self._reagent_excel_path)
            self._excel_watcher.addPath(self._reagent_excel_path)
            os.startfile(self._reagent_excel_path)

            if hasattr(self.app, 'log_browser') and self.app.log_browser:
                self.app.signals.sig_log.emit("Excel 파일 열림 - 시약/시퀀스 모두 저장하면 자동 반영됩니다")

        except Exception as e:
            QMessageBox.critical(self.app, "오류", f"Excel 파일 생성/열기 실패:\n{e}")

    def _write_reagent_sheet(self, ws, pumps, header_font, header_fill, center_align, thin_border):
        """Sheet 1: 시약 설정 작성"""
        headers = ["포트"]
        for p in pumps:
            headers.append(f"{p} 시약명")
            headers.append(f"{p} 농도(M)")
            headers.append(f"{p} SMILES")
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for port in range(1, 13):
            row_data = [port]
            for p in pumps:
                tbl = self.app.reagent_tables[p]
                row_idx = port - 1
                name = tbl.item(row_idx, 1).text() if tbl.item(row_idx, 1) else "Empty"
                conc = tbl.item(row_idx, 2).text() if tbl.item(row_idx, 2) else "1.0"
                smiles = tbl.item(row_idx, 3).text() if tbl.item(row_idx, 3) else ""
                row_data.append(name)
                row_data.append(float(conc) if conc else 1.0)
                row_data.append(smiles)
            ws.append(row_data)

        for row in ws.iter_rows(min_row=2, max_row=13, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length * 1.5 + 2, 12)

        note_row = 15
        note_cell = ws.cell(row=note_row, column=1,
                            value="※ 시약 설정과 시퀀스 설정 모두 저장(Ctrl+S) 시 자동 반영됩니다.")
        note_cell.font = Font(italic=True, color="808080")

    def _write_sequence_sheet(self, ws, pumps, header_font, header_fill, center_align, thin_border):
        """Sheet 2: 시퀀스 설정 작성 — 현재 프로그램의 Step 데이터를 Excel에 기록"""
        # 헤더 구성: Step | 조건4개 | 포트 모아서 | 당량 모아서
        headers = ["Step", "반응온도(℃)", "반응시간(min)", "반응부피(mL)", "분획부피(mL)"]
        for p in pumps:
            headers.append(f"{p} 포트")
        for p in pumps:
            headers.append(f"{p} 당량(eq)")
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 현재 시퀀스 데이터 기록
        seq_data = []
        if hasattr(self.app, 'seq_tab'):
            seq_data = self.app.seq_tab.sequence_data

        if seq_data:
            for step_idx, step in enumerate(seq_data):
                row_data = [
                    step_idx + 1,
                    step.get('temp', 25.0),
                    step.get('rt', 10.0),
                    step.get('vol', 5.0),
                    step.get('tube_vol', 1.5),
                ]
                for p in pumps:
                    pump_data = step.get('pumps', {}).get(p, {'port': 2, 'eq': 1.0})
                    row_data.append(pump_data.get('port', 2))
                for p in pumps:
                    pump_data = step.get('pumps', {}).get(p, {'port': 2, 'eq': 1.0})
                    row_data.append(round(pump_data.get('eq', 1.0), 2))
                ws.append(row_data)

            # 데이터 행 스타일 적용
            data_rows = len(seq_data)
            for row in ws.iter_rows(min_row=2, max_row=1 + data_rows, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.alignment = center_align
                    cell.border = thin_border
        else:
            # Step이 없으면 안내 행
            guide_cell = ws.cell(row=2, column=1, value="← 행을 추가하여 Step을 만드세요")
            guide_cell.font = Font(italic=True, color="808080")

        # 컬럼 너비 자동 조정
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length * 1.5 + 2, 14)

        # 안내 노트 — 헤더 셀 코멘트로 삽입 (데이터 영역 방해 없음)
        from openpyxl.comments import Comment
        ws.cell(row=1, column=1).comment = Comment(
            "행 추가 = Step 추가, 행 삭제 = Step 제거.\n"
            "조건 값이 있는 행이 자동 반영됩니다.\n"
            "포트: 1~12 정수, 당량: 소수점 가능 (예: 1.50)",
            "FlowChem"
        )

    # ─── 파일 감시 ─────────────────────────────────────────────

    def _on_excel_file_changed(self, path):
        """Excel 파일 변경 감지 시 자동 반영
        @codesyncer-decision: 1초 지연 후 적용 - Excel이 파일 잠금 해제할 시간 확보
        @codesyncer-decision: 1.5초 후 watcher 재등록 - Windows에서 파일 변경 시 watcher 해제됨
        """
        if path == self._reagent_excel_path:
            QTimer.singleShot(1000, self._apply_excel_silently)
            QTimer.singleShot(1500, self._re_register_excel_watcher)

    def _re_register_excel_watcher(self):
        """파일 감시 재등록 (Windows 호환성)
        @codesyncer-inference: Windows QFileSystemWatcher는 파일 변경 이벤트 후 자동 해제됨
          - 검증: Qt 5.x 공식 문서 "file is renamed or removed from disk" 참조"""
        if self._reagent_excel_path and os.path.exists(self._reagent_excel_path):
            if self._reagent_excel_path not in self._excel_watcher.files():
                self._excel_watcher.addPath(self._reagent_excel_path)

    # ─── 시약 시트 반영 (공용) ─────────────────────────────────

    def _apply_reagent_sheet(self, wb):
        """Sheet 1(시약 설정) 파싱 후 reagent_tables + map_mgr 업데이트.
        반환: 반영된 펌프 수"""
        ws = wb["시약 설정"] if "시약 설정" in wb.sheetnames else wb.active

        headers = [cell.value for cell in ws[1]]
        pump_columns = {}

        for i, h in enumerate(headers):
            if h and " 시약명" in str(h):
                pump_name = str(h).replace(" 시약명", "")
                if i + 1 < len(headers) and headers[i + 1] and " 농도" in str(headers[i + 1]):
                    # SMILES 컬럼(i+2)은 있을 수도/없을 수도(구버전 파일) → None 허용
                    smiles_idx = None
                    if i + 2 < len(headers) and headers[i + 2] and "SMILES" in str(headers[i + 2]):
                        smiles_idx = i + 2
                    pump_columns[pump_name] = (i, i + 1, smiles_idx)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=13), start=1):
            port = row_idx

            for pump_name, (name_idx, conc_idx, smiles_idx) in pump_columns.items():
                if pump_name in self.app.reagent_tables:
                    tbl = self.app.reagent_tables[pump_name]
                    name_val = row[name_idx].value if row[name_idx].value else "Empty"
                    conc_val = row[conc_idx].value if row[conc_idx].value else 1.0
                    smiles_val = ""
                    if smiles_idx is not None and row[smiles_idx].value:
                        smiles_val = str(row[smiles_idx].value).strip()

                    # @codesyncer-decision: Port 1(Solvent), Port 12(Waste)는 시스템 예약 → 편집 차단
                    if port == 1 or port == 12:
                        continue

                    row_in_table = port - 1
                    tbl.blockSignals(True)
                    tbl.setItem(row_in_table, 1, QTableWidgetItem(str(name_val)))
                    conc_item = QTableWidgetItem(str(conc_val))
                    conc_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                    tbl.setItem(row_in_table, 2, conc_item)
                    smi_item = QTableWidgetItem(smiles_val)
                    smi_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                    tbl.setItem(row_in_table, 3, smi_item)
                    tbl.blockSignals(False)

                    if self.app.map_mgr:
                        self.app.map_mgr.update_inlet(pump_name, port, str(name_val), float(conc_val), smiles_val)

        return len(pump_columns)

    # ─── 시퀀스 시트 반영 (공용) ───────────────────────────────

    def _apply_sequence_sheet(self, wb):
        """Sheet 2(시퀀스 설정) 파싱 후 seq_tab의 StepCard 재구성.
        반환: 반영된 Step 수"""
        if "시퀀스 설정" not in wb.sheetnames:
            return 0

        seq_tab = getattr(self.app, 'seq_tab', None)
        if not seq_tab:
            return 0

        ws = wb["시퀀스 설정"]
        headers = [cell.value for cell in ws[1]]

        # 헤더에서 펌프 컬럼 매핑 (포트/당량이 분리 배치)
        port_cols = {}  # {pump_name: col_idx}
        eq_cols = {}    # {pump_name: col_idx}
        for i, h in enumerate(headers):
            if h and " 포트" in str(h):
                port_cols[str(h).replace(" 포트", "")] = i
            elif h and " 당량" in str(h):
                eq_cols[str(h).replace(" 당량(eq)", "")] = i
        # 포트+당량 모두 있는 펌프만 매핑
        pump_col_map = {}
        for pump_name in port_cols:
            if pump_name in eq_cols:
                pump_col_map[pump_name] = (port_cols[pump_name], eq_cols[pump_name])

        # 유효한 행 수집: Step 열에 숫자가 있거나, 조건 컬럼(온도/시간/부피) 중 값이 있는 행
        step_rows = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            # 빈 행 건너뛰기
            if len(row) < 2:
                continue
            step_val = row[0].value
            has_step_num = False
            if step_val is not None:
                try:
                    int(step_val)
                    has_step_num = True
                except (ValueError, TypeError):
                    pass
            # Step 번호가 없어도 조건 컬럼(1~4: 온도/시간/부피/분획)에 값이 있으면 유효
            has_data = any(
                row[c].value is not None
                for c in range(1, min(5, len(row)))
            )
            if has_step_num or has_data:
                step_rows.append(row)

        if not step_rows:
            return 0

        # 기존 Step 모두 제거 후 Excel 데이터로 재구성
        seq_tab.clear_all_steps()

        for row in step_rows:
            seq_tab.add_step()
            step_idx = len(seq_tab.sequence_data) - 1
            card = seq_tab.step_cards[step_idx]
            step_data = seq_tab.sequence_data[step_idx]

            # 반응 조건 파싱 (헤더 인덱스 기반)
            def _safe_float(cell, default):
                try:
                    return float(cell.value) if cell.value is not None else default
                except (ValueError, TypeError):
                    return default

            temp = _safe_float(row[1], 25.0)
            rt = _safe_float(row[2], 10.0)
            vol = _safe_float(row[3], 5.0)
            tube_vol = _safe_float(row[4], 1.5)

            step_data['temp'] = temp
            step_data['rt'] = rt
            step_data['vol'] = vol
            step_data['tube_vol'] = tube_vol

            card.sp_temp.setValue(temp)
            card.sp_rt.setValue(rt)
            card.sp_vol.setValue(vol)
            card.sp_tube.setValue(tube_vol)

            # 펌프별 포트/당량 파싱
            for pump_name, (port_col, eq_col) in pump_col_map.items():
                if pump_name not in step_data.get('pumps', {}):
                    continue

                port_val = row[port_col].value
                eq_val = row[eq_col].value

                try:
                    port_int = int(port_val) if port_val is not None else 2
                    port_int = max(1, min(12, port_int))
                except (ValueError, TypeError):
                    port_int = 2

                try:
                    eq_float = float(eq_val) if eq_val is not None else 1.0
                except (ValueError, TypeError):
                    eq_float = 1.0

                step_data['pumps'][pump_name]['port'] = port_int
                step_data['pumps'][pump_name]['eq'] = eq_float

                pw = card.pump_widgets.get(pump_name)
                if pw:
                    cb = pw['cb_port']
                    idx = cb.findData(port_int)
                    if idx >= 0:
                        cb.blockSignals(True)
                        cb.setCurrentIndex(idx)
                        cb.blockSignals(False)
                    pw['sp_eq'].setValue(eq_float)

        seq_tab._mark_flow_dirty("excel_sequence_loaded")
        return len(step_rows)

    # ─── 자동 반영 ─────────────────────────────────────────────

    def _apply_excel_silently(self):
        """Excel 변경사항 자동 적용 (메시지 없이)"""
        if not self._reagent_excel_path or not os.path.exists(self._reagent_excel_path):
            return

        try:
            wb = openpyxl.load_workbook(self._reagent_excel_path)

            # 1. 시약 시트 먼저 반영 (포트콤보에 시약명 필요)
            num_pumps = self._apply_reagent_sheet(wb)

            # 2. 시퀀스 시트 반영
            num_steps = self._apply_sequence_sheet(wb)

            # 3. 포트 콤보 최종 갱신 (시약명+농도 → 시퀀스 카드에 반영)
            if hasattr(self.app, 'seq_tab'):
                self.app.seq_tab.refresh_port_combos()

            if hasattr(self.app, 'log_browser') and self.app.log_browser:
                msg = f"✓ Excel 자동 반영 완료 (시약 {num_pumps}펌프"
                if num_steps > 0:
                    msg += f", 시퀀스 {num_steps}스텝"
                msg += ")"
                self.app.signals.sig_log.emit(msg)

        except PermissionError:
            QTimer.singleShot(1000, self._apply_excel_silently)
        except Exception as e:
            print(f"[Excel] 자동 반영 실패: {e}")

    # ─── 수동 반영 ─────────────────────────────────────────────

    def apply_reagents_excel(self):
        """Excel 파일의 변경사항을 적용 (수동)"""
        if not HAS_OPENPYXL:
            QMessageBox.warning(self.app, "경고",
                "openpyxl 패키지가 설치되지 않았습니다.\npip install openpyxl 명령으로 설치하세요.")
            return

        if not self._reagent_excel_path or not os.path.exists(self._reagent_excel_path):
            QMessageBox.warning(self.app, "경고", "먼저 'Excel 편집' 버튼을 눌러 파일을 생성하세요.")
            return

        try:
            wb = openpyxl.load_workbook(self._reagent_excel_path)

            num_pumps = self._apply_reagent_sheet(wb)
            num_steps = self._apply_sequence_sheet(wb)

            if hasattr(self.app, 'seq_tab'):
                self.app.seq_tab.refresh_port_combos()

            msg = f"시약 설정 ({num_pumps}펌프)"
            if num_steps > 0:
                msg += f" + 시퀀스 {num_steps}개 Step"
            msg += " 적용 완료"
            QMessageBox.information(self.app, "완료", msg)

        except Exception as e:
            QMessageBox.critical(self.app, "오류", f"Excel 파일 읽기 실패:\n{e}")
